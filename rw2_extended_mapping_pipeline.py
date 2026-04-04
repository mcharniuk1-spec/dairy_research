#!/usr/bin/env python3
"""
RW2 thesis-grade product-first pipeline.

Execution overview implemented:
- Auto-discovery and fail-fast schema checks
- Product/standardized_type classification before all analytics
- Unit harmonization with unified `price` naming
- Daily series variants (real/linear/pchip) with imputation flags
- Descriptive stats, correlations, diagnostic tests
- Product-first forecasts (prices + Silpo discounts)
- One output workbook: rw2_full_output.xlsx
- Short structured terminal report
"""

from __future__ import annotations

import argparse
import re
import shutil
import tempfile
import warnings
import zlib
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl.styles import Alignment, Font, PatternFill
from statsmodels.stats.diagnostic import acorr_ljungbox, het_breuschpagan, het_white
from statsmodels.stats.stattools import jarque_bera
from statsmodels.tsa.seasonal import STL
from statsmodels.tsa.stattools import adfuller, coint, kpss
from statsmodels.tsa.vector_ar.vecm import VECM, select_coint_rank, select_order

warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.filterwarnings("ignore", category=UserWarning)

try:
    from scipy.interpolate import PchipInterpolator

    HAS_PCHIP = True
except Exception:
    HAS_PCHIP = False


DEFAULT_MODEL_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = DEFAULT_MODEL_DIR / "full_uah.xlsx"
DEFAULT_OUTPUT = DEFAULT_MODEL_DIR / "results" / "rw2_full_output.xlsx"


DATASET_ORDER = [
    "FarmGateUA_initial",
    "FarmGateUA_filled",
    "ProducerUA",
    "ConsumerUA",
    "EU",
    "ProZorro",
    "Silpo",
    "Novus",
    "CME",
]


STANDARDIZED_RULES: List[Tuple[str, List[str]]] = [
    ("butter", ["масло", "butter"]),
    ("sour_cream", ["сметан", "sour cream"]),
    ("cream", ["вершк", "cream"]),
    ("cottage_cheese", ["кисломолоч", "творог", "cottage"]),
    ("hard_cheese", ["сир", "cheddar", "edam", "emmental", "gouda", "hard cheese"]),
    (
        "milk",
        [
            "молоко",
            "milk",
            "кефір",
            "кeфір",
            "ryazh",
            "ряжан",
            "smp",
            "wmp",
            "drinking milk",
            "згущ",
        ],
    ),
    ("yogurt_dessert", ["йогурт", "yogurt", "десерт", "dessert"]),
]


PRODUCT_RULES: List[Tuple[str, List[str]]] = [
    ("Масло вершкове", ["масло", "butter"]),
    ("Сметана", ["сметан", "sour cream"]),
    ("Вершки", ["вершк", "cream"]),
    ("Сир кисломолочний", ["кисломолоч", "творог", "cottage"]),
    ("Сир твердий", ["сир", "cheddar", "edam", "emmental", "gouda", "hard cheese"]),
    ("Йогурт", ["йогурт", "yogurt"]),
    ("Кефір", ["кефір", "кeфір", "kefir"]),
    ("Молоко питне", ["молоко", "drinking milk", "milk"]),
]

KG_UNITS = {"кг", "kg", "кілограм", "kilogram", "килограмм"}
G_UNITS = {"г", "гр", "g", "gram", "grams"}
L_UNITS = {"л", "літр", "l", "liter", "litre"}
ML_UNITS = {"мл", "ml"}
PCS_UNITS = {"шт", "штука", "pcs", "piece", "уп", "упаковка"}


@dataclass
class Config:
    model_dir: Path
    input_file: Path
    output_file: Path
    pdf_output: Optional[Path] = None
    min_obs_tests: int = 24
    min_obs_forecast: int = 45


def ntext(v: object) -> str:
    if pd.isna(v):
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower())


def normalize_col(c: str) -> str:
    s = ntext(c)
    s = s.replace("_", " ").replace("-", " ")
    s = s.replace("(", " ").replace(")", " ")
    return re.sub(r"\s+", " ", s).strip()


def find_column(df: pd.DataFrame, aliases: Iterable[str]) -> Optional[str]:
    if df.empty:
        return None
    rev = {normalize_col(c): c for c in df.columns}
    for a in aliases:
        a_n = normalize_col(a)
        if a_n in rev:
            return rev[a_n]
    # partial fallback
    for a in aliases:
        a_n = normalize_col(a)
        for c_norm, c_real in rev.items():
            if a_n in c_norm:
                return c_real
    return None


def require_columns(df: pd.DataFrame, requirements: Dict[str, List[str]], ctx: str) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    missing: List[str] = []
    for key, aliases in requirements.items():
        col = find_column(df, aliases)
        if col is None:
            missing.append(f"{key}: {aliases}")
        else:
            mapping[key] = col
    if missing:
        available = ", ".join([str(c) for c in df.columns])
        raise ValueError(
            f"[{ctx}] Missing critical columns:\n- "
            + "\n- ".join(missing)
            + f"\nAvailable columns: {available}"
        )
    return mapping


def detect_standardized_type(*fields: object) -> str:
    text = " | ".join(ntext(f) for f in fields if not pd.isna(f))
    for std_name, keys in STANDARDIZED_RULES:
        if any(k in text for k in keys):
            return std_name
    return "other"


def detect_product_name(*fields: object) -> str:
    text = " | ".join(ntext(f) for f in fields if not pd.isna(f))
    for p_name, keys in PRODUCT_RULES:
        if any(k in text for k in keys):
            return p_name
    return "Інше/невідомо"


def parse_pack_size_kg_or_l(title: object) -> Tuple[Optional[float], Optional[str]]:
    s = ntext(title)
    if not s:
        return None, None

    # Match patterns like 900г, 1.5 кг, 930 мл, 1л
    m = re.search(r"(\d+(?:[\.,]\d+)?)\s*(кг|г|л|мл|kg|g|l|ml)", s)
    if not m:
        return None, None

    qty = float(m.group(1).replace(",", "."))
    unit = m.group(2)

    if unit in {"кг", "kg"}:
        return qty, "kg"
    if unit in {"г", "g"}:
        return qty / 1000.0, "kg"
    if unit in {"л", "l"}:
        return qty, "liter"
    if unit in {"мл", "ml"}:
        return qty / 1000.0, "liter"
    return None, None


def discover_input_files(model_dir: Path) -> List[Path]:
    return sorted([p for p in model_dir.glob("*.xlsx") if p.is_file()])


def discover_input_workbook(model_dir: Path, input_file: Path) -> Path:
    if input_file.exists():
        return input_file
    xlsx = discover_input_files(model_dir)
    preferred = [p for p in xlsx if p.name.lower() == "full_uah.xlsx"]
    if preferred:
        return preferred[0]
    if xlsx:
        return xlsx[0]
    raise FileNotFoundError(f"No .xlsx files found in {model_dir}")


def build_daily_variants(
    g: pd.DataFrame,
    date_col: str = "date",
    value_col: str = "price",
    include_interpolation: bool = True,
) -> pd.DataFrame:
    s = g[[date_col, value_col]].copy()
    s[date_col] = pd.to_datetime(s[date_col], errors="coerce")
    s[value_col] = pd.to_numeric(s[value_col], errors="coerce")
    s = s.dropna(subset=[date_col]).groupby(date_col, as_index=False)[value_col].mean().sort_values(date_col)
    if s.empty:
        return pd.DataFrame(columns=["date", "price_real", "price_linear", "price_pchip", "imputed_flag_linear", "imputed_flag_pchip"])

    idx = pd.date_range(s[date_col].min(), s[date_col].max(), freq="D")
    out = pd.DataFrame({"date": idx})
    out = out.merge(s.rename(columns={value_col: "price_real"}), on="date", how="left")

    if include_interpolation:
        out["price_linear"] = out["price_real"].interpolate(method="linear", limit_direction="both")

        if HAS_PCHIP and out["price_real"].notna().sum() >= 3:
            xs = out.loc[out["price_real"].notna(), "date"]
            ys = out.loc[out["price_real"].notna(), "price_real"]
            x0 = xs.min()
            xp = (xs - x0).dt.days.to_numpy(dtype=float)
            yp = ys.to_numpy(dtype=float)
            xg = (out["date"] - x0).dt.days.to_numpy(dtype=float)
            try:
                pchip = PchipInterpolator(xp, yp, extrapolate=True)
                out["price_pchip"] = pchip(xg)
                out["price_pchip"] = pd.Series(out["price_pchip"]).replace([np.inf, -np.inf], np.nan).interpolate(limit_direction="both")
            except Exception:
                out["price_pchip"] = out["price_linear"]
        else:
            out["price_pchip"] = out["price_linear"]

        out["imputed_flag_linear"] = ((out["price_real"].isna()) & (out["price_linear"].notna())).astype(int)
        out["imputed_flag_pchip"] = ((out["price_real"].isna()) & (out["price_pchip"].notna())).astype(int)
    else:
        out["price_linear"] = np.nan
        out["price_pchip"] = np.nan
        out["imputed_flag_linear"] = 0
        out["imputed_flag_pchip"] = 0
    return out


def _desc_one(series: pd.Series, imputed_flag: Optional[pd.Series] = None) -> Dict[str, float]:
    s = pd.to_numeric(series, errors="coerce")
    n_total = len(s)
    s_valid = s.dropna()
    if len(s_valid) == 0:
        return {
            "count": 0,
            "missing": n_total,
            "imputed_share": float(np.nan),
            "mean": np.nan,
            "median": np.nan,
            "std": np.nan,
            "min": np.nan,
            "max": np.nan,
            "q05": np.nan,
            "q25": np.nan,
            "q75": np.nan,
            "q95": np.nan,
            "skew": np.nan,
            "kurtosis": np.nan,
            "cv": np.nan,
            "rolling_vol_7": np.nan,
            "rolling_vol_30": np.nan,
        }

    returns = np.log(s_valid).diff().dropna()
    vol7 = returns.rolling(7).std().mean() if len(returns) >= 7 else np.nan
    vol30 = returns.rolling(30).std().mean() if len(returns) >= 30 else np.nan
    std = float(s_valid.std(ddof=1)) if len(s_valid) > 1 else 0.0
    mean = float(s_valid.mean())
    cv = std / mean if mean != 0 else np.nan
    imp = float(imputed_flag.mean()) if imputed_flag is not None else np.nan
    return {
        "count": int(s_valid.count()),
        "missing": int(n_total - s_valid.count()),
        "imputed_share": imp,
        "mean": mean,
        "median": float(s_valid.median()),
        "std": std,
        "min": float(s_valid.min()),
        "max": float(s_valid.max()),
        "q05": float(s_valid.quantile(0.05)),
        "q25": float(s_valid.quantile(0.25)),
        "q75": float(s_valid.quantile(0.75)),
        "q95": float(s_valid.quantile(0.95)),
        "skew": float(s_valid.skew()),
        "kurtosis": float(s_valid.kurtosis()),
        "cv": float(cv),
        "rolling_vol_7": float(vol7) if pd.notna(vol7) else np.nan,
        "rolling_vol_30": float(vol30) if pd.notna(vol30) else np.nan,
    }


def daily_variants_for_dataset(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    keep_cols = [
        c
        for c in [
            "date",
            "product",
            "standardized_type",
            "brand",
            "region",
            "price",
            "observed_price",
            "baseline_price",
            "price_variant",
            "reconstruction_variant",
            "subcategory",
            "unit_family",
            "comparison_family",
            "liter_equiv_allowed",
            "maturation_type",
            "admissible_for_level_model",
            "admissibility_reason",
            "mapping_quality_flag",
            "matched_pattern",
            "fat_band",
            "pack_band",
            "segment_key",
            "shock_dummy",
            "unit_ok",
            "price_real_input",
            "price_linear_input",
            "price_pchip_input",
        ]
        if c in df.columns
    ]
    d = df[keep_cols].copy()
    if "brand" not in d.columns:
        d["brand"] = ""
    if "region" not in d.columns:
        d["region"] = ""
    if "unit_ok" not in d.columns:
        d["unit_ok"] = 1

    meta_cols = [
        c
        for c in [
            "subcategory",
            "unit_family",
            "comparison_family",
            "liter_equiv_allowed",
            "maturation_type",
            "admissible_for_level_model",
            "admissibility_reason",
            "mapping_quality_flag",
            "matched_pattern",
            "fat_band",
            "pack_band",
            "segment_key",
            "shock_dummy",
        ]
        if c in d.columns
    ]
    groups = []
    gcols = ["product", "standardized_type", "brand", "region"]
    for keys, g in d.groupby(gcols, dropna=False):
        if source in {"ProducerUA", "ConsumerUA", "FarmGateUA_initial", "FarmGateUA_filled"} and ("price_linear_input" in g.columns or "price_pchip_input" in g.columns):
            base = g.copy()
            base["date"] = pd.to_datetime(base["date"], errors="coerce")
            base = base.dropna(subset=["date"])
            if base.empty:
                continue
            idx = pd.date_range(base["date"].min(), base["date"].max(), freq="D")
            dv = pd.DataFrame({"date": idx})
            price_real_src = "price_real_input" if "price_real_input" in base.columns else "price"
            pr = base.groupby("date", as_index=False)[price_real_src].mean().rename(columns={price_real_src: "price_real"})
            pl = (
                base.groupby("date", as_index=False)["price_linear_input"].mean().rename(columns={"price_linear_input": "price_linear"})
                if "price_linear_input" in base.columns
                else pd.DataFrame(columns=["date", "price_linear"])
            )
            pp = (
                base.groupby("date", as_index=False)["price_pchip_input"].mean().rename(columns={"price_pchip_input": "price_pchip"})
                if "price_pchip_input" in base.columns
                else pd.DataFrame(columns=["date", "price_pchip"])
            )
            dv = dv.merge(pr, on="date", how="left").merge(pl, on="date", how="left").merge(pp, on="date", how="left")
            if "price_linear" not in dv.columns:
                dv["price_linear"] = np.nan
            if "price_pchip" not in dv.columns:
                dv["price_pchip"] = np.nan
            if dv["price_real"].isna().all():
                dv["price_real"] = dv["price_linear"].where(dv["price_linear"].notna(), dv["price_pchip"])
            dv["price_linear"] = dv["price_linear"].where(dv["price_linear"].notna(), dv["price_real"].interpolate(method="linear", limit_direction="both"))
            dv["price_pchip"] = dv["price_pchip"].where(dv["price_pchip"].notna(), dv["price_linear"])
            dv["imputed_flag_linear"] = ((dv["price_real"].isna()) & (dv["price_linear"].notna())).astype(int)
            dv["imputed_flag_pchip"] = ((dv["price_real"].isna()) & (dv["price_pchip"].notna())).astype(int)
        else:
            dv = build_daily_variants(
                g,
                date_col="date",
                value_col="price",
                include_interpolation=False,
            )
        if dv.empty:
            continue
        key_vals = keys if isinstance(keys, tuple) else (keys,)
        for k, v in zip(gcols, key_vals):
            dv[k] = v
        dv["source"] = source
        dv["unit_ok"] = int(pd.to_numeric(g.get("unit_ok"), errors="coerce").fillna(0).max())
        if "admissible_for_level_model" in g.columns:
            dv["admissible_for_level_model"] = int(pd.to_numeric(g["admissible_for_level_model"], errors="coerce").fillna(0).max())
        for mc in meta_cols:
            non_null = g[mc].dropna()
            dv[mc] = non_null.iloc[0] if not non_null.empty else np.nan
        groups.append(dv)

    if not groups:
        return pd.DataFrame()

    out = pd.concat(groups, ignore_index=True)
    cols = [
        "source",
        "date",
        "product",
        "standardized_type",
        "brand",
        "region",
        "subcategory",
        "unit_family",
        "comparison_family",
        "liter_equiv_allowed",
        "maturation_type",
        "mapping_quality_flag",
        "matched_pattern",
        "fat_band",
        "pack_band",
        "segment_key",
        "shock_dummy",
        "admissible_for_level_model",
        "admissibility_reason",
        "unit_ok",
        "price_real",
        "price_linear",
        "price_pchip",
        "imputed_flag_linear",
        "imputed_flag_pchip",
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols]


def build_descriptive_stats(daily_df: pd.DataFrame, source: str) -> pd.DataFrame:
    if daily_df.empty:
        return pd.DataFrame()

    rows = []
    for keys, g in daily_df.groupby(["source", "product", "standardized_type"], dropna=False):
        sname, product, std = keys
        for var_name, col, imp in [
            ("real", "price_real", None),
            ("linear", "price_linear", "imputed_flag_linear"),
            ("pchip", "price_pchip", "imputed_flag_pchip"),
        ]:
            stats = _desc_one(g[col], g[imp] if imp else None)
            rows.append({"source": sname, "product": product, "standardized_type": std, "series_variant": var_name, **stats})
    out = pd.DataFrame(rows)
    return out.sort_values(["source", "product", "series_variant"]) if not out.empty else out


def long_price_series(daily_df: pd.DataFrame) -> pd.DataFrame:
    if daily_df.empty:
        return pd.DataFrame()
    out = daily_df.melt(
        id_vars=[c for c in ["source", "date", "product", "standardized_type", "brand", "region", "subcategory", "unit_family", "comparison_family", "segment_key", "admissible_for_level_model", "unit_ok", "imputed_flag_linear", "imputed_flag_pchip"] if c in daily_df.columns],
        value_vars=[c for c in ["price_real", "price_linear", "price_pchip"] if c in daily_df.columns],
        var_name="series_variant",
        value_name="price",
    )
    out["series_variant"] = out["series_variant"].str.replace("price_", "", regex=False)
    return out.dropna(subset=["price"]).reset_index(drop=True)


def build_series_catalog(long_df: pd.DataFrame) -> pd.DataFrame:
    if long_df.empty:
        return pd.DataFrame()
    rows = []
    for (src, prod, std, var), g in long_df.groupby(["source", "product", "standardized_type", "series_variant"], dropna=False):
        rows.append(
            {
                "source": src,
                "product": prod,
                "standardized_type": std,
                "series_variant": var,
                "date_min": pd.to_datetime(g["date"]).min(),
                "date_max": pd.to_datetime(g["date"]).max(),
                "n_obs": int(g["price"].notna().sum()),
                "missing_share": float(g["price"].isna().mean()),
                "imputed_share": float(g["imputed_flag_pchip"].mean()) if "imputed_flag_pchip" in g.columns else np.nan,
                "unit_ok_share": float(pd.to_numeric(g.get("unit_ok"), errors="coerce").mean()) if "unit_ok" in g.columns else np.nan,
                "block": "A_long_run_core" if src in {"ProducerUA", "ConsumerUA", "EU"} else "B_retail_intersection",
            }
        )
    return pd.DataFrame(rows).sort_values(["standardized_type", "source", "product", "series_variant"])


def winsorize_daily_variants(all_daily: pd.DataFrame, lower_q: float = 0.01, upper_q: float = 0.99) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if all_daily.empty:
        return all_daily, pd.DataFrame()
    out = all_daily.copy()
    summary_rows = []
    for col in [c for c in ["price_real", "price_linear", "price_pchip"] if c in out.columns]:
        out[f"{col}_outlier_flag"] = 0
        out[f"{col}_winsor"] = out[col]
        for (src, prod), idx in out.groupby(["source", "product"], dropna=False).groups.items():
            s = pd.to_numeric(out.loc[idx, col], errors="coerce")
            valid = s.dropna()
            if len(valid) < 20:
                continue
            lo = float(valid.quantile(lower_q))
            hi = float(valid.quantile(upper_q))
            clipped = s.clip(lower=lo, upper=hi)
            flag = ((s < lo) | (s > hi)).fillna(False).astype(int)
            out.loc[idx, f"{col}_winsor"] = clipped
            out.loc[idx, f"{col}_outlier_flag"] = flag
            summary_rows.append(
                {
                    "source": src,
                    "product": prod,
                    "series_variant": col.replace("price_", ""),
                    "q01": lo,
                    "q99": hi,
                    "share_trimmed": float(flag.mean()),
                    "n_obs": int(valid.shape[0]),
                }
            )
    # Promote winsorized as working columns for modeling
    for col in ["price_real", "price_linear", "price_pchip"]:
        if f"{col}_winsor" in out.columns:
            out[col] = out[f"{col}_winsor"]
    return out, pd.DataFrame(summary_rows)


def build_standardisation_policy(
    datasets_clean: Dict[str, pd.DataFrame],
) -> pd.DataFrame:
    rows = []
    for src, df in datasets_clean.items():
        if df.empty:
            rows.append(
                {
                    "source": src,
                    "rows": 0,
                    "unit_ok_share": np.nan,
                    "policy_gate_level_models": "blocked",
                    "currency_path": "UAH-native (no FX conversion)",
                    "note": "No data.",
                }
            )
            continue
        unit_ok_share = float(pd.to_numeric(df.get("unit_ok"), errors="coerce").fillna(0).mean()) if "unit_ok" in df.columns else np.nan
        gate = "allowed" if (pd.notna(unit_ok_share) and unit_ok_share >= 0.95) else "restricted_or_index_like"
        note = "Use only unit_ok==1 for level/cointegration models."
        if src in {"Silpo", "Novus"} and gate != "allowed":
            note = "Retail has partial unit uncertainty; use index-like or baseline logic with explicit caveat."
        rows.append(
            {
                "source": src,
                "rows": len(df),
                "unit_ok_share": unit_ok_share,
                "policy_gate_level_models": gate,
                "currency_path": "UAH-native (no FX conversion)",
                "note": note,
            }
        )
    return pd.DataFrame(rows)


def _effective_daily_prices(all_daily: pd.DataFrame) -> pd.DataFrame:
    if all_daily.empty:
        return pd.DataFrame()
    d = all_daily.copy()
    d["price_eff"] = d["price_pchip"].where(d["source"].isin(["ProducerUA", "ConsumerUA"]) & d["price_pchip"].notna(), d["price_real"])
    return d


def build_lag_matrix_all(all_daily: pd.DataFrame) -> pd.DataFrame:
    d = _effective_daily_prices(all_daily)
    if d.empty:
        return pd.DataFrame()
    pairs = [
        ("ProducerUA", "ConsumerUA"),
        ("EU", "ProducerUA"),
        ("ProducerUA", "ProZorro"),
        ("ProducerUA", "Silpo"),
        ("ProducerUA", "Novus"),
        ("Novus", "Silpo"),
        ("EU", "Silpo"),
        ("EU", "ProZorro"),
    ]
    rows = []
    lag_range = range(1, 31)
    for (prod, std), g in d.groupby(["product", "standardized_type"], dropna=False):
        pivot = g.pivot_table(index="date", columns="source", values="price_eff", aggfunc="mean").sort_index()
        for left, right in pairs:
            if left not in pivot.columns or right not in pivot.columns:
                continue
            s1 = pivot[left]
            s2 = pivot[right]
            best_lag = None
            best_corr = None
            for lag in lag_range:
                pair = pd.concat([s1, s2.shift(lag)], axis=1).dropna()
                if len(pair) < 15:
                    continue
                corr = float(pair.iloc[:, 0].corr(pair.iloc[:, 1]))
                rows.append(
                    {
                        "product": prod,
                        "standardized_type": std,
                        "pair_left": left,
                        "pair_right": right,
                        "lag_days": lag,
                        "corr": corr,
                    }
                )
                if best_corr is None or abs(corr) > abs(best_corr):
                    best_corr = corr
                    best_lag = lag
            if best_lag is not None:
                rows.append(
                    {
                        "product": prod,
                        "standardized_type": std,
                        "pair_left": left,
                        "pair_right": right,
                        "lag_days": best_lag,
                        "corr": best_corr,
                        "is_best_lag": 1,
                    }
                )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["is_best_lag"] = out.get("is_best_lag", 0).fillna(0).astype(int)
    return out.sort_values(["standardized_type", "product", "pair_left", "pair_right", "lag_days", "is_best_lag"])


def build_brand_io_metrics(silpo_clean: pd.DataFrame, novus_clean: pd.DataFrame) -> pd.DataFrame:
    retail = pd.concat([silpo_clean, novus_clean], ignore_index=True) if (not silpo_clean.empty or not novus_clean.empty) else pd.DataFrame()
    if retail.empty:
        return pd.DataFrame()
    r = retail.copy()
    r["date"] = pd.to_datetime(r["date"], errors="coerce")
    r = r.dropna(subset=["date"]).copy()
    r["month"] = r["date"].dt.to_period("M").dt.to_timestamp()
    r["brand"] = r["brand"].fillna("unknown").astype(str)
    private_patterns = ["premiya", "premia", "повна чаша", "private", "власн"]

    rows = []
    for (src, std, month), g in r.groupby(["source", "standardized_type", "month"], dropna=False):
        shares = g["brand"].value_counts(normalize=True)
        hhi = float(np.sum(np.square(shares.values))) if not shares.empty else np.nan
        top3 = float(shares.head(3).sum()) if not shares.empty else np.nan
        sku_count = int(g["title"].nunique()) if "title" in g.columns else int(g["product"].nunique())
        priv = g["brand"].str.lower().map(lambda x: any(p in x for p in private_patterns))
        private_share = float(priv.mean()) if len(priv) > 0 else np.nan
        rows.append(
            {
                "source": src,
                "standardized_type": std,
                "month": month,
                "hhi_brand": hhi,
                "top3_share": top3,
                "private_label_share": private_share,
                "sku_count": sku_count,
                "n_obs": int(len(g)),
            }
        )
    return pd.DataFrame(rows).sort_values(["source", "standardized_type", "month"])


def build_brand_economic_metrics(
    silpo_clean: pd.DataFrame,
    novus_clean: pd.DataFrame,
    all_daily: pd.DataFrame,
) -> pd.DataFrame:
    retail = pd.concat([silpo_clean, novus_clean], ignore_index=True) if (not silpo_clean.empty or not novus_clean.empty) else pd.DataFrame()
    if retail.empty:
        return pd.DataFrame()
    r = retail.copy()
    r["date"] = pd.to_datetime(r["date"], errors="coerce")
    r = r.dropna(subset=["date"]).copy()
    r["month"] = r["date"].dt.to_period("M").dt.to_timestamp()
    r["brand"] = r["brand"].fillna("unknown").astype(str)
    r["price"] = pd.to_numeric(r["price"], errors="coerce")
    rows = []
    eff = _effective_daily_prices(all_daily)
    bm = (
        eff[eff["source"].isin(["EU", "CME"])]
        .groupby(["date", "standardized_type"], as_index=False)["price_eff"]
        .mean()
        .rename(columns={"price_eff": "benchmark_price"})
    ) if not eff.empty else pd.DataFrame(columns=["date", "standardized_type", "benchmark_price"])
    for (src, std, month, brand), g in r.groupby(["source", "standardized_type", "month", "brand"], dropna=False):
        cat = r[(r["source"] == src) & (r["standardized_type"] == std) & (r["month"] == month)]
        brand_med = float(g["price"].median()) if g["price"].notna().any() else np.nan
        cat_med = float(cat["price"].median()) if cat["price"].notna().any() else np.nan
        premium = float(np.log(brand_med) - np.log(cat_med)) if (pd.notna(brand_med) and pd.notna(cat_med) and brand_med > 0 and cat_med > 0) else np.nan
        promo_intensity = float(pd.to_numeric(g.get("discount_present"), errors="coerce").fillna(0).mean()) if "discount_present" in g.columns else np.nan
        dlog = np.log(pd.to_numeric(g["price"], errors="coerce")).diff()
        vol = float(dlog.std()) if dlog.notna().sum() > 1 else np.nan

        # brand-level pass-through proxy
        coef = np.nan
        pval = np.nan
        gg = g.groupby("date", as_index=False)["price"].median().sort_values("date")
        mg = gg.merge(bm[bm["standardized_type"] == std], on="date", how="inner")
        mg["d_brand"] = np.log(mg["price"]).diff()
        mg["d_bm"] = np.log(mg["benchmark_price"]).diff()
        mg = mg.dropna()
        if len(mg) > 15:
            try:
                fit = sm.OLS(mg["d_brand"], sm.add_constant(mg[["d_bm"]])).fit(cov_type="HC1")
                coef = float(fit.params.get("d_bm", np.nan))
                pval = float(fit.pvalues.get("d_bm", np.nan))
            except Exception:
                pass

        rows.append(
            {
                "source": src,
                "standardized_type": std,
                "month": month,
                "brand": brand,
                "brand_premium_log": premium,
                "promo_intensity": promo_intensity,
                "price_volatility": vol,
                "brand_pass_through_coef": coef,
                "brand_pass_through_pvalue": pval,
                "n_obs": int(len(g)),
            }
        )
    return pd.DataFrame(rows).sort_values(["source", "standardized_type", "month", "brand"])


def build_prozorro_regional_fe(prozorro_clean: pd.DataFrame, model_series: pd.DataFrame) -> pd.DataFrame:
    if prozorro_clean.empty or model_series.empty:
        return pd.DataFrame()
    p = prozorro_clean.copy()
    p["week"] = pd.to_datetime(p["date"]).dt.to_period("W-MON").dt.start_time
    rows = []
    for std, g in p.groupby("standardized_type", dropna=False):
        reg = g.groupby(["region", "week"], as_index=False)["price"].median().rename(columns={"price": "price_pz"})
        bm = model_series[(model_series["standardized_type"] == std) & (model_series["source"].isin(["EU", "ProducerUA"]))].groupby("week", as_index=False)["price"].mean().rename(columns={"price": "price_bm"})
        d = reg.merge(bm, on="week", how="inner")
        d = d[(d["price_pz"] > 0) & (d["price_bm"] > 0)].sort_values(["region", "week"])
        d["dlog_price"] = d.groupby("region")["price_pz"].transform(lambda s: np.log(s).diff())
        d["dlog_bm"] = np.log(d["price_bm"]).diff()
        d = d.dropna(subset=["dlog_price", "dlog_bm"])
        if len(d) < 40:
            continue
        X = pd.concat([d[["dlog_bm"]], pd.get_dummies(d["region"], prefix="r", drop_first=True)], axis=1)
        X = sm.add_constant(X).astype(float)
        y = pd.to_numeric(d["dlog_price"], errors="coerce")
        fit = sm.OLS(y, X).fit(cov_type="HC1")
        rows.append(
            {
                "standardized_type": std,
                "n_obs": int(len(d)),
                "coef_dlog_benchmark": float(fit.params.get("dlog_bm", np.nan)),
                "p_dlog_benchmark": float(fit.pvalues.get("dlog_bm", np.nan)),
                "r2": float(fit.rsquared),
                "region_fe_count": int(sum([1 for c in X.columns if str(c).startswith("r_")])),
                "interpretation_note": "Regional FE model for institutional procurement heterogeneity.",
            }
        )
    return pd.DataFrame(rows)


def build_silpo_discount_modules(
    silpo_clean: pd.DataFrame,
    novus_clean: pd.DataFrame,
    all_daily: pd.DataFrame,
    brand_io: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if silpo_clean.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    s = silpo_clean.copy()
    s["date"] = pd.to_datetime(s["date"], errors="coerce")
    s = s.dropna(subset=["date"]).copy()
    s["month"] = s["date"].dt.to_period("M").dt.to_timestamp()
    s["baseline_price"] = np.where(
        pd.to_numeric(s["price_old"], errors="coerce").notna() & (pd.to_numeric(s["price_old"], errors="coerce") > 0),
        pd.to_numeric(s["price_old"], errors="coerce"),
        pd.to_numeric(s["price_current"], errors="coerce"),
    )
    s["discount_present"] = pd.to_numeric(s["discount_present"], errors="coerce").fillna(0).clip(0, 1)
    s["discount_depth"] = pd.to_numeric(s["discount_depth"], errors="coerce")

    eff = _effective_daily_prices(all_daily)
    bm = (
        eff[eff["source"].isin(["EU", "ProducerUA"])]
        .groupby(["date", "standardized_type", "source"], as_index=False)["price_eff"]
        .mean()
        .pivot_table(index=["date", "standardized_type"], columns="source", values="price_eff", aggfunc="mean")
        .reset_index()
    ) if not eff.empty else pd.DataFrame(columns=["date", "standardized_type", "EU", "ProducerUA"])
    nov = (
        novus_clean.groupby(["date", "standardized_type"], as_index=False)["price"]
        .median()
        .rename(columns={"price": "Novus"})
    ) if not novus_clean.empty else pd.DataFrame(columns=["date", "standardized_type", "Novus"])
    s = s.merge(bm, on=["date", "standardized_type"], how="left").merge(nov, on=["date", "standardized_type"], how="left")
    for c in ["EU", "ProducerUA", "Novus", "baseline_price", "price_current"]:
        if c in s.columns:
            s[f"dlog_{c}"] = np.log(pd.to_numeric(s[c], errors="coerce")).diff()
    s["dow"] = s["date"].dt.dayofweek

    # Attach IO metric hhi
    if not brand_io.empty:
        key = brand_io[["source", "standardized_type", "month", "hhi_brand"]].copy()
        key = key[key["source"] == "Silpo"][["standardized_type", "month", "hhi_brand"]]
        s = s.merge(key, on=["standardized_type", "month"], how="left")
    else:
        s["hhi_brand"] = np.nan

    occ_rows, depth_rows, trans_rows = [], [], []
    for std, g in s.groupby("standardized_type", dropna=False):
        g = g.sort_values("date").copy()
        top_brands = g["brand"].value_counts().head(8).index.tolist() if "brand" in g.columns else []
        if "brand" in g.columns:
            g["brand_grp"] = np.where(g["brand"].isin(top_brands), g["brand"], "other")
            brand_d = pd.get_dummies(g["brand_grp"], prefix="b", drop_first=True)
        else:
            brand_d = pd.DataFrame(index=g.index)
        X_base = pd.concat(
            [
                g[[c for c in ["dlog_EU", "dlog_ProducerUA", "dlog_Novus", "dow", "hhi_brand"] if c in g.columns]],
                brand_d,
            ],
            axis=1,
        ).apply(pd.to_numeric, errors="coerce")
        X_base = sm.add_constant(X_base).fillna(0.0)

        # Occurrence
        y_occ = (g["discount_present"] > 0.05).astype(int)
        if len(g) >= 40 and y_occ.nunique() > 1:
            try:
                fit_occ = sm.Logit(y_occ, X_base).fit(disp=False)
                occ_rows.append(
                    {
                        "standardized_type": std,
                        "n_obs": int(len(g)),
                        "coef_dlog_EU": float(fit_occ.params.get("dlog_EU", np.nan)),
                        "p_dlog_EU": float(fit_occ.pvalues.get("dlog_EU", np.nan)),
                        "coef_dlog_ProducerUA": float(fit_occ.params.get("dlog_ProducerUA", np.nan)),
                        "p_dlog_ProducerUA": float(fit_occ.pvalues.get("dlog_ProducerUA", np.nan)),
                        "coef_dlog_Novus": float(fit_occ.params.get("dlog_Novus", np.nan)),
                        "p_dlog_Novus": float(fit_occ.pvalues.get("dlog_Novus", np.nan)),
                        "coef_hhi_brand": float(fit_occ.params.get("hhi_brand", np.nan)),
                        "p_hhi_brand": float(fit_occ.pvalues.get("hhi_brand", np.nan)),
                        "model": "logit",
                        "interpretation_note": "Occurrence model for discount regime decision.",
                    }
                )
            except Exception:
                pass

        # Depth conditional
        gd = g[y_occ == 1].copy()
        if len(gd) >= 30:
            Xd = X_base.loc[gd.index]
            yd = pd.to_numeric(gd["discount_depth"], errors="coerce")
            m = pd.concat([yd.rename("y"), Xd], axis=1).dropna()
            if len(m) >= 20:
                try:
                    fit_d = sm.OLS(m["y"], m.drop(columns=["y"])).fit(cov_type="HC1")
                    depth_rows.append(
                        {
                            "standardized_type": std,
                            "n_obs": int(len(m)),
                            "coef_dlog_EU": float(fit_d.params.get("dlog_EU", np.nan)),
                            "p_dlog_EU": float(fit_d.pvalues.get("dlog_EU", np.nan)),
                            "coef_dlog_ProducerUA": float(fit_d.params.get("dlog_ProducerUA", np.nan)),
                            "p_dlog_ProducerUA": float(fit_d.pvalues.get("dlog_ProducerUA", np.nan)),
                            "coef_dlog_Novus": float(fit_d.params.get("dlog_Novus", np.nan)),
                            "p_dlog_Novus": float(fit_d.pvalues.get("dlog_Novus", np.nan)),
                            "coef_hhi_brand": float(fit_d.params.get("hhi_brand", np.nan)),
                            "p_hhi_brand": float(fit_d.pvalues.get("hhi_brand", np.nan)),
                            "r2": float(fit_d.rsquared),
                            "model": "ols_hc1",
                            "interpretation_note": "Depth model conditional on active discount.",
                        }
                    )
                except Exception:
                    pass

        # Transmission with/without promo controls
        agg = (
            g.groupby("date", as_index=False)
            .agg(
                baseline_price=("baseline_price", "median"),
                effective_price=("price_current", "median"),
                discount_present=("discount_present", "mean"),
                discount_depth=("discount_depth", "mean"),
                dlog_EU=("dlog_EU", "mean"),
                dlog_ProducerUA=("dlog_ProducerUA", "mean"),
            )
            .sort_values("date")
        )
        agg["dlog_base"] = np.log(pd.to_numeric(agg["baseline_price"], errors="coerce")).diff()
        agg["dlog_eff"] = np.log(pd.to_numeric(agg["effective_price"], errors="coerce")).diff()
        agg["lag_base"] = agg["dlog_base"].shift(1)
        a1 = agg[["dlog_base", "dlog_EU", "dlog_ProducerUA", "lag_base"]].dropna()
        if len(a1) >= 20:
            fit1 = sm.OLS(a1["dlog_base"], sm.add_constant(a1[["dlog_EU", "dlog_ProducerUA", "lag_base"]])).fit(cov_type="HAC", cov_kwds={"maxlags": 2})
            a2 = agg[["dlog_base", "dlog_EU", "dlog_ProducerUA", "lag_base", "discount_present", "discount_depth"]].dropna()
            if len(a2) >= 20:
                fit2 = sm.OLS(a2["dlog_base"], sm.add_constant(a2[["dlog_EU", "dlog_ProducerUA", "lag_base", "discount_present", "discount_depth"]])).fit(cov_type="HAC", cov_kwds={"maxlags": 2})
                trans_rows.append(
                    {
                        "standardized_type": std,
                        "n_obs": int(len(a2)),
                        "coef_EU_no_promo": float(fit1.params.get("dlog_EU", np.nan)),
                        "coef_EU_with_promo": float(fit2.params.get("dlog_EU", np.nan)),
                        "coef_Producer_no_promo": float(fit1.params.get("dlog_ProducerUA", np.nan)),
                        "coef_Producer_with_promo": float(fit2.params.get("dlog_ProducerUA", np.nan)),
                        "delta_EU": float(fit2.params.get("dlog_EU", np.nan) - fit1.params.get("dlog_EU", np.nan)),
                        "delta_Producer": float(fit2.params.get("dlog_ProducerUA", np.nan) - fit1.params.get("dlog_ProducerUA", np.nan)),
                        "promo_coef_present": float(fit2.params.get("discount_present", np.nan)),
                        "promo_coef_depth": float(fit2.params.get("discount_depth", np.nan)),
                        "interpretation_note": "Difference between no-promo and promo-controlled pass-through.",
                    }
                )

    occ_df = pd.DataFrame(occ_rows)
    depth_df = pd.DataFrame(depth_rows)
    trans_df = pd.DataFrame(trans_rows)
    if occ_df.empty:
        occ_df = pd.DataFrame(
            [
                {
                    "standardized_type": "n/a",
                    "n_obs": 0,
                    "model": "logit",
                    "interpretation_note": "Insufficient variation/data for occurrence model in current window.",
                }
            ]
        )
    if depth_df.empty:
        depth_df = pd.DataFrame(
            [
                {
                    "standardized_type": "n/a",
                    "n_obs": 0,
                    "model": "ols_hc1",
                    "interpretation_note": "Insufficient discounted observations for depth model.",
                }
            ]
        )
    if trans_df.empty:
        trans_df = pd.DataFrame(
            [
                {
                    "standardized_type": "n/a",
                    "n_obs": 0,
                    "interpretation_note": "Insufficient data for promo-controlled transmission comparison.",
                }
            ]
        )
    return occ_df, depth_df, trans_df


def build_msem_summary(all_daily: pd.DataFrame) -> pd.DataFrame:
    d = _effective_daily_prices(all_daily)
    if d.empty:
        return pd.DataFrame()
    window_start = pd.Timestamp("2025-10-01")
    window_end = pd.Timestamp("2026-01-31")
    d = d[(pd.to_datetime(d["date"]) >= window_start) & (pd.to_datetime(d["date"]) <= window_end)].copy()
    rows = []
    sources_keep = ["Silpo", "Novus", "ProZorro", "ProducerUA", "ConsumerUA", "EU"]
    for std, g in d.groupby("standardized_type", dropna=False):
        wide = (
            g[g["source"].isin(sources_keep)]
            .pivot_table(index="date", columns="source", values="price_eff", aggfunc="mean")
            .sort_index()
        )
        wide = wide.ffill().bfill().dropna(axis=1, how="all")
        if wide.shape[1] < 3 or len(wide) < 25:
            continue
        z = (wide - wide.mean()) / wide.std(ddof=0).replace(0, np.nan)
        z = z.dropna()
        if z.empty:
            continue
        X = z.to_numpy()
        U, S, Vt = np.linalg.svd(X, full_matrices=False)
        loadings = Vt[0, :]
        factor = U[:, 0] * S[0]
        rows.append(
            {
                "standardized_type": std,
                "window": "2025-10-01..2026-01-31",
                "n_obs": int(z.shape[0]),
                "n_series": int(z.shape[1]),
                "sources_used": ", ".join(z.columns.tolist()),
                "factor_volatility": float(np.std(factor)),
                "loadings": "; ".join([f"{c}:{v:.3f}" for c, v in zip(z.columns.tolist(), loadings)]),
                "interpretation_note": "MSEM-like latent factor for short-window co-movement (robustness module).",
            }
        )
    return pd.DataFrame(rows)


def build_matrices_heatmap_ready(
    model_plan: pd.DataFrame,
    brand_io: pd.DataFrame,
    ardl: pd.DataFrame,
    neio_proxy: pd.DataFrame,
    lag_matrix: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    if not model_plan.empty:
        for _, r in model_plan.iterrows():
            std = r["standardized_type"]
            for src in ["ProducerUA", "ConsumerUA", "EU", "CME", "ProZorro", "Silpo", "Novus"]:
                rows.append(
                    {
                        "matrix": "coverage",
                        "standardized_type": std,
                        "row_key": src,
                        "col_key": "available",
                        "value": 1.0 if src in str(r.get("available_series", "")) else 0.0,
                    }
                )
    if not brand_io.empty:
        bi = brand_io.groupby("standardized_type", as_index=False)["hhi_brand"].mean()
        for _, r in bi.iterrows():
            rows.append(
                {
                    "matrix": "brand_io",
                    "standardized_type": r["standardized_type"],
                    "row_key": "hhi_brand",
                    "col_key": "mean",
                    "value": float(r["hhi_brand"]),
                }
            )
    if not ardl.empty:
        aa = (
            ardl.groupby("standardized_type", as_index=False)["long_run_coef"]
            .agg(lambda s: float(np.nanmean(np.abs(pd.to_numeric(s, errors="coerce")))))
            .rename(columns={"long_run_coef": "v"})
        )
        for _, r in aa.iterrows():
            rows.append(
                {
                    "matrix": "transmission_strength",
                    "standardized_type": r["standardized_type"],
                    "row_key": "abs_long_run_coef",
                    "col_key": "mean",
                    "value": float(r["v"]),
                }
            )
    if not neio_proxy.empty:
        for _, r in neio_proxy.iterrows():
            rows.append(
                {
                    "matrix": "neio_proxy",
                    "standardized_type": r["standardized_type"],
                    "row_key": "asymmetry_strength_index",
                    "col_key": "value",
                    "value": float(r.get("asymmetry_strength_index", np.nan)),
                }
            )
    if not lag_matrix.empty:
        best = lag_matrix[lag_matrix.get("is_best_lag", 0) == 1]
        for _, r in best.iterrows():
            rows.append(
                {
                    "matrix": "lag_best",
                    "standardized_type": r["standardized_type"],
                    "row_key": f"{r['pair_left']}->{r['pair_right']}",
                    "col_key": "lag_days",
                    "value": float(r["lag_days"]),
                }
            )
    return pd.DataFrame(rows)


def build_corr_matrix(corr_df: pd.DataFrame) -> pd.DataFrame:
    if corr_df.empty:
        return pd.DataFrame()
    d = corr_df[(corr_df["corr_type"] == "between_sources") & (corr_df["lag"] == 0)].copy()
    if d.empty:
        return pd.DataFrame()
    srcs = sorted(set(d["source_left"].dropna().unique().tolist()) | set(d["source_right"].dropna().unique().tolist()))
    mat = pd.DataFrame(np.nan, index=srcs, columns=srcs, dtype=float)
    for s in srcs:
        mat.loc[s, s] = 1.0
    agg = d.groupby(["source_left", "source_right"], as_index=False)["pearson"].mean()
    for _, r in agg.iterrows():
        a, b, v = r["source_left"], r["source_right"], float(r["pearson"])
        mat.loc[a, b] = v
        mat.loc[b, a] = v
    mat.index.name = "source"
    return mat.reset_index()


def build_stargazer_like(
    ardl: pd.DataFrame,
    ecm: pd.DataFrame,
    nardl: pd.DataFrame,
    vecm: pd.DataFrame,
    proz_fe: pd.DataFrame,
    silpo_trans: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    if not ardl.empty:
        for _, r in ardl.iterrows():
            rows.append(
                {
                    "model": "ARDL",
                    "standardized_type": r.get("standardized_type"),
                    "y_series": r.get("y_series_source"),
                    "x_series": r.get("x_series_sources"),
                    "sample_period": r.get("sample_period"),
                    "coef_main": r.get("short_run_coef"),
                    "coef_long_run": r.get("long_run_coef"),
                    "p_or_rank": r.get("coint_or_bounds_p"),
                    "diagnostics_or_note": r.get("diagnostics_flags"),
                }
            )
    if not ecm.empty:
        for _, r in ecm.iterrows():
            rows.append(
                {
                    "model": "ECM",
                    "standardized_type": r.get("standardized_type"),
                    "y_series": r.get("y_series_source"),
                    "x_series": r.get("x_series_sources"),
                    "sample_period": r.get("sample_period"),
                    "coef_main": r.get("short_run_coef"),
                    "coef_long_run": r.get("long_run_coef"),
                    "p_or_rank": r.get("ect_pvalue"),
                    "diagnostics_or_note": f"ECT={r.get('ect_coef')}",
                }
            )
    if not nardl.empty:
        for _, r in nardl.iterrows():
            rows.append(
                {
                    "model": "NARDL",
                    "standardized_type": r.get("standardized_type"),
                    "y_series": r.get("y_series_source"),
                    "x_series": r.get("x_series_sources"),
                    "sample_period": r.get("sample_period"),
                    "coef_main": r.get("short_run_coef"),
                    "coef_long_run": r.get("long_run_coef"),
                    "p_or_rank": r.get("asymmetry_long_p"),
                    "diagnostics_or_note": f"asym_short_p={r.get('asymmetry_short_p')}",
                }
            )
    if not vecm.empty:
        for _, r in vecm.iterrows():
            rows.append(
                {
                    "model": "VECM",
                    "standardized_type": r.get("standardized_type"),
                    "y_series": r.get("y_series_source"),
                    "x_series": r.get("x_series_sources"),
                    "sample_period": r.get("sample_period"),
                    "coef_main": r.get("adjustment_alpha_abs_mean"),
                    "coef_long_run": np.nan,
                    "p_or_rank": r.get("cointegration_rank"),
                    "diagnostics_or_note": r.get("system"),
                }
            )
    if not proz_fe.empty:
        for _, r in proz_fe.iterrows():
            rows.append(
                {
                    "model": "Prozorro_FE",
                    "standardized_type": r.get("standardized_type"),
                    "y_series": "ProZorro unit price",
                    "x_series": "Benchmark shocks + region FE",
                    "sample_period": "",
                    "coef_main": r.get("coef_dlog_benchmark"),
                    "coef_long_run": np.nan,
                    "p_or_rank": r.get("p_dlog_benchmark"),
                    "diagnostics_or_note": f"R2={r.get('r2')}",
                }
            )
    if not silpo_trans.empty:
        for _, r in silpo_trans.iterrows():
            rows.append(
                {
                    "model": "Silpo_PromoCtrl",
                    "standardized_type": r.get("standardized_type"),
                    "y_series": "Silpo baseline dlog",
                    "x_series": "EU/Producer shocks",
                    "sample_period": "",
                    "coef_main": r.get("coef_EU_with_promo"),
                    "coef_long_run": r.get("coef_Producer_with_promo"),
                    "p_or_rank": np.nan,
                    "diagnostics_or_note": f"delta_EU={r.get('delta_EU')}; delta_Producer={r.get('delta_Producer')}",
                }
            )
    return pd.DataFrame(rows)


def _export_separate_csvs(base_dir: Path, named_frames: Dict[str, pd.DataFrame]) -> None:
    base_dir.mkdir(parents=True, exist_ok=True)
    for name, df in named_frames.items():
        safe = re.sub(r"[^A-Za-z0-9_\\-]+", "_", name).strip("_")
        (df if df is not None else pd.DataFrame()).to_csv(base_dir / f"{safe}.csv", index=False)


def _excel_safe_sheet_name(*parts: str, max_len: int = 31) -> str:
    raw = "_".join([str(p) for p in parts if str(p)])
    cleaned = re.sub(r"[\\/*?:\[\]]+", "_", raw).strip("_")
    cleaned = re.sub(r"\s+", "_", cleaned)
    if len(cleaned) <= max_len:
        return cleaned
    suffix = f"{zlib.crc32(cleaned.encode('utf-8')) & 0xffff:04x}"
    keep = max_len - len(suffix) - 1
    return f"{cleaned[:keep]}_{suffix}"


def _find_date_col(df: pd.DataFrame) -> Optional[str]:
    if df is None or df.empty:
        return None
    dcol = find_column(df, ["date", "дата", "week", "month"])
    if dcol:
        return dcol
    for c in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[c]):
            return c
    return None


def _find_price_col(df: pd.DataFrame) -> Optional[str]:
    if df is None or df.empty:
        return None
    pcol = find_column(
        df,
        [
            "price",
            "price_eff",
            "unit_price",
            "price_current",
            "price_real",
            "value",
            "coef",
            "long_run_coef",
        ],
    )
    if pcol:
        return pcol
    num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    return num_cols[0] if num_cols else None


def _series_from_table(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=float)
    date_col = _find_date_col(df)
    price_col = _find_price_col(df)
    if date_col is None or price_col is None:
        return pd.Series(dtype=float)
    d = df[[date_col, price_col]].copy()
    d.columns = ["date", "price"]
    d["date"] = pd.to_datetime(d["date"], errors="coerce")
    d["price"] = pd.to_numeric(d["price"], errors="coerce")
    d = d.dropna(subset=["date", "price"]).sort_values("date")
    if d.empty:
        return pd.Series(dtype=float)
    s = d.groupby("date", as_index=True)["price"].median().sort_index()
    return s


def _step_test_bundle(df: pd.DataFrame, min_obs: int = 24) -> Dict[str, object]:
    s = _series_from_table(df)
    if s.empty or len(s) < min_obs:
        return {
            "n_obs": int(len(s)),
            "adf_p": np.nan,
            "kpss_p": np.nan,
            "ljungbox_p": np.nan,
            "bp_p": np.nan,
            "white_p": np.nan,
            "jb_p": np.nan,
            "stability_flag": 0,
            "integration_class": "insufficient",
            "recommended_action": "Collect longer series / aggregate frequency",
            "recommended_model_family": "insufficient_data",
            "avoid": "Strong inference on short/fragmented sample",
            "diag_note": "Step sample too short for full diagnostics.",
        }
    w = s.resample("W").mean().dropna()
    tests = run_tests_for_series(w, min_obs=min_obs)
    row = pd.Series(tests)
    diag = interpret_diagnostics(row)
    return {**tests, **diag}


def build_transformation_outputs(step_frames: List[Dict[str, object]], min_obs: int = 24) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    ledger_rows: List[Dict[str, object]] = []
    stats_rows: List[Dict[str, object]] = []
    tests_rows: List[Dict[str, object]] = []
    for item in step_frames:
        dataset = str(item.get("dataset_name", "unknown"))
        sheet = str(item.get("sheet_name", dataset))
        step_id = str(item.get("step_id", "step"))
        step_desc = str(item.get("step_description", ""))
        before_df = item.get("before_df", pd.DataFrame())
        after_df = item.get("after_df", pd.DataFrame())
        before = before_df if isinstance(before_df, pd.DataFrame) else pd.DataFrame()
        after = after_df if isinstance(after_df, pd.DataFrame) else pd.DataFrame()

        b_series = _series_from_table(before)
        a_series = _series_from_table(after)
        b_missing = float(before.isna().mean().mean()) if not before.empty else np.nan
        a_missing = float(after.isna().mean().mean()) if not after.empty else np.nan
        outlier_share = np.nan
        if not a_series.empty:
            q1 = float(a_series.quantile(0.01))
            q99 = float(a_series.quantile(0.99))
            outlier_share = float(((a_series < q1) | (a_series > q99)).mean())
        unit_ok_share = float(pd.to_numeric(after.get("unit_ok"), errors="coerce").mean()) if "unit_ok" in after.columns else np.nan

        tests = _step_test_bundle(after, min_obs=min_obs)
        note = str(tests.get("diag_note", ""))
        rec_action = str(tests.get("recommended_action", ""))
        note_short = f"{note} {rec_action}".strip()

        ledger_rows.append(
            {
                "dataset_name": dataset,
                "sheet_name": sheet,
                "step_id": step_id,
                "step_description": step_desc,
                "rows_before": int(len(before)),
                "rows_after": int(len(after)),
                "missing_before": b_missing,
                "missing_after": a_missing,
                "outlier_share": outlier_share,
                "unit_ok_share": unit_ok_share,
                "date_range_before": f"{b_series.index.min().date()}..{b_series.index.max().date()}" if not b_series.empty else "",
                "date_range_after": f"{a_series.index.min().date()}..{a_series.index.max().date()}" if not a_series.empty else "",
                "interpretation_note": note_short[:300],
            }
        )

        if not a_series.empty:
            stats = _desc_one(a_series)
            stats_rows.append(
                {
                    "dataset_name": dataset,
                    "step_id": step_id,
                    "step_description": step_desc,
                    **stats,
                }
            )
        tests_rows.append(
            {
                "dataset_name": dataset,
                "step_id": step_id,
                "step_description": step_desc,
                **tests,
            }
        )

    return pd.DataFrame(ledger_rows), pd.DataFrame(stats_rows), pd.DataFrame(tests_rows)


def build_log_variant_table(all_daily: pd.DataFrame) -> pd.DataFrame:
    if all_daily.empty:
        return pd.DataFrame()
    rows: List[pd.DataFrame] = []
    for variant, col in [("real", "price_real"), ("linear", "price_linear"), ("pchip", "price_pchip")]:
        if col not in all_daily.columns:
            continue
        d = all_daily[["source", "product", "standardized_type", "date", col]].copy()
        d = d.rename(columns={col: "price"}).dropna(subset=["price"])
        if d.empty:
            continue
        d = d.sort_values(["source", "product", "date"]).copy()
        d["series_variant"] = variant
        d["log_price"] = np.where(pd.to_numeric(d["price"], errors="coerce") > 0, np.log(pd.to_numeric(d["price"], errors="coerce")), np.nan)
        d["dlog_price"] = d.groupby(["source", "product"], dropna=False)["log_price"].diff()
        rows.append(d)
    out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    return out.sort_values(["source", "product", "series_variant", "date"]) if not out.empty else out


def build_decomposition_tables(all_daily: pd.DataFrame, min_obs: int = 35) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, pd.DataFrame], pd.DataFrame]:
    d = _effective_daily_prices(all_daily)
    if d.empty:
        return pd.DataFrame(), pd.DataFrame(), {}, pd.DataFrame()
    d = d[["source", "product", "standardized_type", "date", "price_eff"]].copy()
    d["price_eff"] = pd.to_numeric(d["price_eff"], errors="coerce")
    d = d.dropna(subset=["price_eff", "date"])
    d = d[d["price_eff"] > 0].copy()

    comp_rows: List[pd.DataFrame] = []
    summary_rows: List[Dict[str, object]] = []
    sheet_tables: Dict[str, pd.DataFrame] = {}
    index_rows: List[Dict[str, object]] = []

    for (src, prod, std), g in d.groupby(["source", "product", "standardized_type"], dropna=False):
        s = g.groupby("date", as_index=True)["price_eff"].median().sort_index()
        if len(s) < min_obs:
            continue
        try:
            log_s = np.log(s)
            stl = STL(log_s, period=7, robust=True)
            res = stl.fit()
            comp = pd.DataFrame(
                {
                    "date": s.index,
                    "source": src,
                    "product": prod,
                    "standardized_type": std,
                    "observed": s.values,
                    "log_observed": log_s.values,
                    "trend": res.trend.values,
                    "seasonal": res.seasonal.values,
                    "resid": res.resid.values,
                }
            )
            comp_rows.append(comp)
            summary_rows.append(
                {
                    "source": src,
                    "product": prod,
                    "standardized_type": std,
                    "n_obs": int(len(comp)),
                    "var_trend": float(np.nanvar(comp["trend"])),
                    "var_seasonal": float(np.nanvar(comp["seasonal"])),
                    "var_resid": float(np.nanvar(comp["resid"])),
                    "seasonal_strength": float(np.nanvar(comp["seasonal"]) / max(np.nanvar(comp["log_observed"]), 1e-9)),
                }
            )
            sheet_name = _excel_safe_sheet_name("Decomposition", src, str(prod))
            sheet_tables[sheet_name] = comp
            index_rows.append({"source": src, "product": prod, "standardized_type": std, "sheet_name": sheet_name})
        except Exception:
            continue

    comp_all = pd.concat(comp_rows, ignore_index=True) if comp_rows else pd.DataFrame()
    summary_df = pd.DataFrame(summary_rows)
    index_df = pd.DataFrame(index_rows)
    return comp_all, summary_df, sheet_tables, index_df


def build_before_after_ln_tables(all_daily: pd.DataFrame, min_obs: int = 10) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame], pd.DataFrame]:
    d = _effective_daily_prices(all_daily)
    if d.empty:
        return pd.DataFrame(), {}, pd.DataFrame()
    d = d[["source", "product", "standardized_type", "date", "price_eff"]].copy()
    d = d.dropna(subset=["date", "price_eff"]).sort_values(["source", "product", "date"])
    rows = []
    sheets: Dict[str, pd.DataFrame] = {}
    idx_rows: List[Dict[str, object]] = []
    for (src, prod, std), g in d.groupby(["source", "product", "standardized_type"], dropna=False):
        s = g.groupby("date", as_index=False)["price_eff"].median().sort_values("date")
        if len(s) < min_obs:
            continue
        s["log_price"] = np.where(s["price_eff"] > 0, np.log(s["price_eff"]), np.nan)
        s["trend_before_ln"] = s["price_eff"].rolling(7, min_periods=3).mean()
        s["trend_after_ln"] = s["log_price"].rolling(7, min_periods=3).mean()
        s["source"] = src
        s["product"] = prod
        s["standardized_type"] = std
        rows.append(s)
        sheet_name = _excel_safe_sheet_name("BeforeAfterLN", src, str(prod))
        sheets[sheet_name] = s.rename(columns={"price_eff": "price"})
        idx_rows.append({"source": src, "product": prod, "standardized_type": std, "sheet_name": sheet_name})
    out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    return out, sheets, pd.DataFrame(idx_rows)


def build_overlay_tables(all_daily: pd.DataFrame, min_obs: int = 10) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame], pd.DataFrame]:
    d = _effective_daily_prices(all_daily)
    if d.empty:
        return pd.DataFrame(), {}, pd.DataFrame()
    d = d[["source", "product", "standardized_type", "date", "price_eff"]].copy()
    d = d.dropna(subset=["date", "price_eff"])
    rows = []
    sheets: Dict[str, pd.DataFrame] = {}
    idx_rows: List[Dict[str, object]] = []
    for (prod, std), g in d.groupby(["product", "standardized_type"], dropna=False):
        p = g.pivot_table(index="date", columns="source", values="price_eff", aggfunc="mean").sort_index()
        if len(p) < min_obs:
            continue
        p = p.reset_index()
        p["product"] = prod
        p["standardized_type"] = std
        rows.append(p)
        sheet_name = _excel_safe_sheet_name("Charts_Overlay", str(prod), "intersection")
        sheets[sheet_name] = p
        idx_rows.append({"product": prod, "standardized_type": std, "window": "intersection", "sheet_name": sheet_name})
    out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    return out, sheets, pd.DataFrame(idx_rows)


def _corr_pairs_from_pivot(pivot: pd.DataFrame, group_meta: Dict[str, object], metric: str) -> pd.DataFrame:
    cols = [c for c in pivot.columns if c != "date"]
    rows = []
    for i in range(len(cols)):
        for j in range(i + 1, len(cols)):
            a, b = cols[i], cols[j]
            pair = pivot[[a, b]].dropna()
            if len(pair) < 10:
                continue
            pear = float(pair[a].corr(pair[b], method="pearson"))
            spea = float(pair[a].corr(pair[b], method="spearman"))
            rows.append({**group_meta, "left": a, "right": b, "metric": metric, "pearson": pear, "spearman": spea, "n_obs": int(len(pair))})
    return pd.DataFrame(rows)


def build_correlation_modules(all_daily: pd.DataFrame, prozorro_clean: pd.DataFrame, silpo_clean: pd.DataFrame, novus_clean: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    outputs: Dict[str, pd.DataFrame] = {}
    eff = _effective_daily_prices(all_daily)
    if not eff.empty:
        # Corr by products within each dataset.
        for src in DATASET_ORDER:
            g = eff[eff["source"] == src].copy()
            if g.empty:
                outputs[f"Corr_Products_{src}"] = pd.DataFrame()
                continue
            p = g.pivot_table(index="date", columns="product", values="price_eff", aggfunc="mean").sort_index()
            logp = np.log(p.where(p > 0))
            dlogp = logp.diff()
            c1 = _corr_pairs_from_pivot(logp.reset_index(), {"dataset": src}, "log_price")
            c2 = _corr_pairs_from_pivot(dlogp.reset_index(), {"dataset": src}, "dlog_price")
            outputs[f"Corr_Products_{src}"] = pd.concat([c1, c2], ignore_index=True)

        # Corr sources by product.
        rows = []
        for (prod, std), g in eff.groupby(["product", "standardized_type"], dropna=False):
            p = g.pivot_table(index="date", columns="source", values="price_eff", aggfunc="mean").sort_index()
            logp = np.log(p.where(p > 0))
            dlogp = logp.diff()
            rows.append(_corr_pairs_from_pivot(logp.reset_index(), {"product": prod, "standardized_type": std}, "log_price"))
            rows.append(_corr_pairs_from_pivot(dlogp.reset_index(), {"product": prod, "standardized_type": std}, "dlog_price"))
        outputs["Corr_Sources_ByProduct"] = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

    # Corr regions for ProZorro.
    if not prozorro_clean.empty:
        pr = prozorro_clean.copy()
        pr["price"] = pd.to_numeric(pr["price"], errors="coerce")
        rr = []
        for prod, g in pr.groupby("product", dropna=False):
            p = g.pivot_table(index="date", columns="region", values="price", aggfunc="median").sort_index()
            logp = np.log(p.where(p > 0))
            dlogp = logp.diff()
            rr.append(_corr_pairs_from_pivot(logp.reset_index(), {"product": prod}, "log_price"))
            rr.append(_corr_pairs_from_pivot(dlogp.reset_index(), {"product": prod}, "dlog_price"))
        outputs["Corr_Regions_Prozorro"] = pd.concat(rr, ignore_index=True) if rr else pd.DataFrame()
    else:
        outputs["Corr_Regions_Prozorro"] = pd.DataFrame()

    # Corr brands for retail.
    retail = pd.concat([silpo_clean, novus_clean], ignore_index=True) if (not silpo_clean.empty or not novus_clean.empty) else pd.DataFrame()
    if not retail.empty:
        br = retail.copy()
        br["price"] = pd.to_numeric(br["price"], errors="coerce")
        rr = []
        for (src, prod), g in br.groupby(["source", "product"], dropna=False):
            p = g.pivot_table(index="date", columns="brand", values="price", aggfunc="median").sort_index()
            if p.shape[1] < 2:
                continue
            logp = np.log(p.where(p > 0))
            dlogp = logp.diff()
            rr.append(_corr_pairs_from_pivot(logp.reset_index(), {"dataset": src, "product": prod}, "log_price"))
            rr.append(_corr_pairs_from_pivot(dlogp.reset_index(), {"dataset": src, "product": prod}, "dlog_price"))
        outputs["Corr_Brands_Retail"] = pd.concat(rr, ignore_index=True) if rr else pd.DataFrame()
    else:
        outputs["Corr_Brands_Retail"] = pd.DataFrame()

    return outputs


def build_lag_outputs(lag_matrix_all: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if lag_matrix_all.empty:
        return pd.DataFrame(), pd.DataFrame()
    best = lag_matrix_all[lag_matrix_all.get("is_best_lag", 0) == 1].copy()
    profile = lag_matrix_all[lag_matrix_all.get("is_best_lag", 0) != 1].copy()
    return best.sort_values(["standardized_type", "product", "pair_left", "pair_right"]), profile.sort_values(
        ["standardized_type", "product", "pair_left", "pair_right", "lag_days"]
    )


def _get_best_lag(lag_best: pd.DataFrame, product: str, left: str, right: str, default: int = 1) -> int:
    if lag_best is None or lag_best.empty:
        return default
    m = lag_best[
        (lag_best["product"] == product)
        & (lag_best["pair_left"] == left)
        & (lag_best["pair_right"] == right)
    ]
    if m.empty:
        return default
    return int(pd.to_numeric(m["lag_days"], errors="coerce").dropna().iloc[0]) if pd.to_numeric(m["lag_days"], errors="coerce").notna().any() else default


def build_short_run_models(all_daily: pd.DataFrame, lag_best: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    eff = _effective_daily_prices(all_daily)
    if eff.empty:
        return pd.DataFrame(), pd.DataFrame()
    details_rows: List[Dict[str, object]] = []
    summary_rows: List[Dict[str, object]] = []
    for (prod, std), g in eff.groupby(["product", "standardized_type"], dropna=False):
        p = g.pivot_table(index="date", columns="source", values="price_eff", aggfunc="mean").sort_index()
        if p.empty:
            continue
        lp = np.log(p.where(p > 0))
        dlp = lp.diff()
        for y_source in ["Silpo", "Novus"]:
            if y_source not in dlp.columns:
                continue
            x_cols = []
            frame = pd.DataFrame({"y": dlp[y_source], "lag_y": dlp[y_source].shift(1)})
            for x_source in ["ProducerUA", "EU", "CME", "ProZorro", "Silpo", "Novus"]:
                if x_source == y_source or x_source not in dlp.columns:
                    continue
                lag = _get_best_lag(lag_best, prod, y_source, x_source, default=1)
                cname = f"x_{x_source}_lag{lag}"
                frame[cname] = dlp[x_source].shift(lag)
                x_cols.append(cname)
            m = frame.dropna()
            if len(m) < 30 or not x_cols:
                continue
            try:
                X = sm.add_constant(m[["lag_y"] + x_cols])
                fit = sm.OLS(m["y"], X).fit(cov_type="HAC", cov_kwds={"maxlags": 3})
                for cname in x_cols:
                    details_rows.append(
                        {
                            "product": prod,
                            "standardized_type": std,
                            "y_source": y_source,
                            "term": cname,
                            "coef": float(fit.params.get(cname, np.nan)),
                            "pvalue": float(fit.pvalues.get(cname, np.nan)),
                            "tvalue": float(fit.tvalues.get(cname, np.nan)),
                            "n_obs": int(len(m)),
                        }
                    )
                summary_rows.append(
                    {
                        "product": prod,
                        "standardized_type": std,
                        "y_source": y_source,
                        "n_obs": int(len(m)),
                        "r2": float(fit.rsquared),
                        "adj_r2": float(fit.rsquared_adj),
                        "producer_effect": float(fit.params.filter(like="x_ProducerUA_").iloc[0]) if any("x_ProducerUA_" in c for c in x_cols) else np.nan,
                        "eu_effect": float(fit.params.filter(like="x_EU_").iloc[0]) if any("x_EU_" in c for c in x_cols) else np.nan,
                        "cme_effect": float(fit.params.filter(like="x_CME_").iloc[0]) if any("x_CME_" in c for c in x_cols) else np.nan,
                        "competitor_effect": float(fit.params.filter(like=f"x_{'Novus' if y_source == 'Silpo' else 'Silpo'}_").iloc[0]) if any(
                            f"x_{'Novus' if y_source == 'Silpo' else 'Silpo'}_" in c for c in x_cols
                        ) else np.nan,
                        "interpretation_note": "Short-run retail pass-through with lag-scan-guided regressors.",
                    }
                )
            except Exception:
                continue
    return pd.DataFrame(summary_rows), pd.DataFrame(details_rows)


def build_chain_effects(all_daily: pd.DataFrame, lag_best: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    eff = _effective_daily_prices(all_daily)
    if eff.empty:
        return pd.DataFrame(), pd.DataFrame()
    summary_rows: List[Dict[str, object]] = []
    details_rows: List[Dict[str, object]] = []
    for (prod, std), g in eff.groupby(["product", "standardized_type"], dropna=False):
        p = g.pivot_table(index="date", columns="source", values="price_eff", aggfunc="mean").sort_index()
        if p.empty:
            continue
        lp = np.log(p.where(p > 0))
        dlp = lp.diff()

        def pair_coef(y_src: str, x_src: str) -> Tuple[float, float, int]:
            if y_src not in dlp.columns or x_src not in dlp.columns:
                return np.nan, np.nan, 0
            lag = _get_best_lag(lag_best, prod, y_src, x_src, default=1)
            m = pd.DataFrame({"y": dlp[y_src], "x": dlp[x_src].shift(lag), "lag_y": dlp[y_src].shift(1)}).dropna()
            if len(m) < 25:
                return np.nan, np.nan, len(m)
            fit = sm.OLS(m["y"], sm.add_constant(m[["x", "lag_y"]])).fit(cov_type="HAC", cov_kwds={"maxlags": 3})
            return float(fit.params.get("x", np.nan)), float(fit.pvalues.get("x", np.nan)), int(len(m))

        # Retail proxy: Silpo first, else Novus.
        retail = "Silpo" if "Silpo" in dlp.columns else ("Novus" if "Novus" in dlp.columns else None)
        if retail is None:
            continue
        rp_coef, rp_p, rp_n = pair_coef(retail, "ProZorro")
        pp_coef, pp_p, pp_n = pair_coef("ProZorro", "ProducerUA")
        rr_coef, rr_p, rr_n = pair_coef(retail, "ProducerUA")
        eu_coef, eu_p, _ = pair_coef(retail, "EU")
        cme_coef, cme_p, _ = pair_coef(retail, "CME")
        summary_rows.append(
            {
                "product": prod,
                "standardized_type": std,
                "retail_source": retail,
                "coef_retail_from_prozorro": rp_coef,
                "p_retail_from_prozorro": rp_p,
                "coef_prozorro_from_producer": pp_coef,
                "p_prozorro_from_producer": pp_p,
                "coef_retail_from_producer": rr_coef,
                "p_retail_from_producer": rr_p,
                "coef_retail_from_eu": eu_coef,
                "p_retail_from_eu": eu_p,
                "coef_retail_from_cme": cme_coef,
                "p_retail_from_cme": cme_p,
                "n_obs_chain_min": int(min([n for n in [rp_n, pp_n, rr_n] if n > 0], default=0)),
            }
        )
        details_rows.extend(
            [
                {"product": prod, "standardized_type": std, "pair": f"{retail}<-ProZorro", "coef": rp_coef, "pvalue": rp_p, "n_obs": rp_n},
                {"product": prod, "standardized_type": std, "pair": "ProZorro<-ProducerUA", "coef": pp_coef, "pvalue": pp_p, "n_obs": pp_n},
                {"product": prod, "standardized_type": std, "pair": f"{retail}<-ProducerUA", "coef": rr_coef, "pvalue": rr_p, "n_obs": rr_n},
                {"product": prod, "standardized_type": std, "pair": f"{retail}<-EU", "coef": eu_coef, "pvalue": eu_p, "n_obs": rp_n},
                {"product": prod, "standardized_type": std, "pair": f"{retail}<-CME", "coef": cme_coef, "pvalue": cme_p, "n_obs": rp_n},
            ]
        )
    return pd.DataFrame(summary_rows), pd.DataFrame(details_rows)


def build_prozorro_regional_modules(prozorro_clean: pd.DataFrame, all_daily: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if prozorro_clean.empty:
        return pd.DataFrame(), pd.DataFrame()
    pz = prozorro_clean.copy()
    pz["date"] = pd.to_datetime(pz["date"], errors="coerce")
    pz["price"] = pd.to_numeric(pz["price"], errors="coerce")
    pz = pz.dropna(subset=["date", "price"]).copy()
    pz = pz[pz["price"] > 0]
    if pz.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Region x product matrix.
    mat = pz.groupby(["region", "product"], as_index=False).agg(
        price_median=("price", "median"),
        price_mean=("price", "mean"),
        price_std=("price", "std"),
        n_obs=("price", "size"),
    )
    mat["cv"] = mat["price_std"] / mat["price_mean"].replace(0, np.nan)

    eff = _effective_daily_prices(all_daily)
    bench = (
        eff[eff["source"].isin(["ProducerUA", "EU", "CME"])]
        .groupby(["date", "standardized_type", "source"], as_index=False)["price_eff"]
        .mean()
        .pivot_table(index=["date", "standardized_type"], columns="source", values="price_eff", aggfunc="mean")
        .reset_index()
    ) if not eff.empty else pd.DataFrame(columns=["date", "standardized_type", "ProducerUA", "EU", "CME"])

    rows = []
    for prod, g in pz.groupby("product", dropna=False):
        std = g["standardized_type"].mode().iloc[0] if "standardized_type" in g.columns and not g["standardized_type"].dropna().empty else "other"
        agg = g.groupby(["date", "region"], as_index=False)["price"].median().rename(columns={"price": "proz_price"})
        agg = agg.merge(bench[bench["standardized_type"] == std], on="date", how="left")
        for c in ["proz_price", "ProducerUA", "EU", "CME"]:
            if c in agg.columns:
                agg[f"log_{c}"] = np.where(pd.to_numeric(agg[c], errors="coerce") > 0, np.log(pd.to_numeric(agg[c], errors="coerce")), np.nan)
                agg[f"dlog_{c}"] = agg.groupby("region", dropna=False)[f"log_{c}"].diff()
        use_cols = [c for c in ["dlog_ProducerUA", "dlog_EU", "dlog_CME"] if c in agg.columns]
        if not use_cols:
            continue
        reg_dummies = pd.get_dummies(agg["region"].astype(str), prefix="region", drop_first=True)
        model_df = pd.concat([agg[["dlog_proz_price"] + use_cols], reg_dummies], axis=1).dropna()
        if len(model_df) < 35:
            continue
        try:
            fit = sm.OLS(model_df["dlog_proz_price"], sm.add_constant(model_df.drop(columns=["dlog_proz_price"]))).fit(cov_type="HC1")
            rows.append(
                {
                    "product": prod,
                    "standardized_type": std,
                    "n_obs": int(len(model_df)),
                    "coef_dlog_producer": float(fit.params.get("dlog_ProducerUA", np.nan)),
                    "p_dlog_producer": float(fit.pvalues.get("dlog_ProducerUA", np.nan)),
                    "coef_dlog_eu": float(fit.params.get("dlog_EU", np.nan)),
                    "p_dlog_eu": float(fit.pvalues.get("dlog_EU", np.nan)),
                    "coef_dlog_cme": float(fit.params.get("dlog_CME", np.nan)),
                    "p_dlog_cme": float(fit.pvalues.get("dlog_CME", np.nan)),
                    "r2": float(fit.rsquared),
                }
            )
        except Exception:
            continue
    return pd.DataFrame(rows), mat


def build_tests_summary(tests_df: pd.DataFrame) -> pd.DataFrame:
    if tests_df.empty:
        return pd.DataFrame()
    out = (
        tests_df.groupby(["source", "product", "standardized_type"], as_index=False)
        .agg(
            n_variants=("series_variant", "nunique"),
            i1_share=("integration_class", lambda s: float((s == "I(1)-like").mean())),
            ac_high_share=("ac_risk", lambda s: float((s == "high").mean())),
            het_high_share=("het_risk", lambda s: float((s == "high").mean())),
            non_normal_share=("non_normal_risk", lambda s: float((s == "high").mean())),
            stability_high_share=("stability_risk_class", lambda s: float((s == "high").mean())),
            recommended_model_family=("recommended_model_family", lambda s: s.mode().iloc[0] if not s.mode().empty else ""),
            recommended_action=("recommended_action", lambda s: s.mode().iloc[0] if not s.mode().empty else ""),
            avoid=("avoid", lambda s: s.mode().iloc[0] if not s.mode().empty else ""),
        )
    )
    return out.sort_values(["source", "product"])


def save_plots_pdf(
    pdf_path: Path,
    corr_matrix: pd.DataFrame,
    tests_df: pd.DataFrame,
    model_plan: pd.DataFrame,
    stargazer_like: pd.DataFrame,
) -> Optional[Path]:
    import os
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/mplconfig")
    os.environ.setdefault("XDG_CACHE_HOME", "/tmp")
    Path(os.environ["MPLCONFIGDIR"]).mkdir(parents=True, exist_ok=True)
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
    except Exception as e:
        print(f"[PLOTS_PDF] Skipped: matplotlib unavailable ({e})")
        return None

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    with PdfPages(pdf_path) as pdf:
        # Cover
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        ax.text(0.02, 0.92, "RW2 Results Plots + Tables", fontsize=20, weight="bold", transform=ax.transAxes)
        ax.text(0.02, 0.84, "Contains: correlation matrix, diagnostics distributions, model availability, stargazer-like summary.", fontsize=11, transform=ax.transAxes)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Correlation heatmap
        if not corr_matrix.empty and "source" in corr_matrix.columns:
            cm = corr_matrix.set_index("source")
            fig, ax = plt.subplots(figsize=(10, 8))
            m = cm.to_numpy(dtype=float)
            im = ax.imshow(m, cmap="coolwarm", vmin=-1, vmax=1)
            ax.set_xticks(range(len(cm.columns)))
            ax.set_xticklabels(cm.columns, rotation=45, ha="right")
            ax.set_yticks(range(len(cm.index)))
            ax.set_yticklabels(cm.index)
            ax.set_title("Correlation Matrix (Pearson, lag=0)")
            for i in range(m.shape[0]):
                for j in range(m.shape[1]):
                    if np.isfinite(m[i, j]):
                        ax.text(j, i, f"{m[i, j]:.2f}", ha="center", va="center", fontsize=7)
            fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

        # Diagnostics distribution plots
        if not tests_df.empty:
            fig, axes = plt.subplots(2, 2, figsize=(12, 8))
            counts_int = tests_df["integration_class"].value_counts(dropna=False)
            axes[0, 0].bar(counts_int.index.astype(str), counts_int.values)
            axes[0, 0].set_title("Integration Class")
            counts_ac = tests_df["ac_risk"].value_counts(dropna=False)
            axes[0, 1].bar(counts_ac.index.astype(str), counts_ac.values)
            axes[0, 1].set_title("Autocorrelation Risk")
            counts_het = tests_df["het_risk"].value_counts(dropna=False)
            axes[1, 0].bar(counts_het.index.astype(str), counts_het.values)
            axes[1, 0].set_title("Heteroskedasticity Risk")
            counts_st = tests_df["stability_risk_class"].value_counts(dropna=False)
            axes[1, 1].bar(counts_st.index.astype(str), counts_st.values)
            axes[1, 1].set_title("Stability Risk")
            plt.tight_layout()
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

        # Model availability chart
        if not model_plan.empty and "possible_models" in model_plan.columns:
            model_tokens = []
            for x in model_plan["possible_models"].dropna().astype(str):
                model_tokens.extend([t.strip() for t in x.split(",") if t.strip()])
            if model_tokens:
                vc = pd.Series(model_tokens).value_counts()
                fig, ax = plt.subplots(figsize=(10, 6))
                ax.bar(vc.index, vc.values)
                ax.set_title("Model Availability by Standardized Type")
                ax.set_ylabel("Count")
                ax.tick_params(axis="x", rotation=30)
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

        # Stargazer-like table preview
        if not stargazer_like.empty:
            prev = stargazer_like.head(30).fillna("")
            fig, ax = plt.subplots(figsize=(14, 8))
            ax.axis("off")
            ax.set_title("Stargazer-like Summary (top rows)")
            txt = prev.to_string(index=False, max_colwidth=24)
            ax.text(0.01, 0.95, txt, transform=ax.transAxes, va="top", family="monospace", fontsize=7)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)
    return pdf_path


def group_stats(df: pd.DataFrame, value_col: str, by_cols: List[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    rows = []
    for keys, g in df.groupby(by_cols, dropna=False):
        stats = _desc_one(g[value_col])
        row = {c: v for c, v in zip(by_cols, keys if isinstance(keys, tuple) else (keys,))}
        rows.append({**row, **stats})
    return pd.DataFrame(rows)


def multi_metric_product_stats(df: pd.DataFrame, metrics: List[str], source_name: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    rows = []
    for metric in metrics:
        if metric not in df.columns:
            continue
        for (prod, std), g in df.groupby(["product", "standardized_type"], dropna=False):
            stats = _desc_one(pd.to_numeric(g[metric], errors="coerce"))
            rows.append(
                {
                    "source": source_name,
                    "product": prod,
                    "standardized_type": std,
                    "metric": metric,
                    **stats,
                }
            )
    return pd.DataFrame(rows)


def prepare_weekly_for_tests(daily_df: pd.DataFrame, value_col: str = "price_pchip") -> pd.DataFrame:
    if daily_df.empty:
        return pd.DataFrame()
    d = daily_df.copy()
    d["week"] = pd.to_datetime(d["date"]).dt.to_period("W-MON").dt.start_time
    blocks = []
    for var, col in [("real", "price_real"), ("linear", "price_linear"), ("pchip", "price_pchip")]:
        if col not in d.columns:
            continue
        out = (
            d.groupby(["source", "product", "standardized_type", "week"], as_index=False)[col]
            .mean()
            .rename(columns={col: "price_weekly"})
        )
        out["series_variant"] = var
        blocks.append(out)
    return pd.concat(blocks, ignore_index=True) if blocks else pd.DataFrame()


def stability_flag(y: pd.Series) -> Tuple[float, float, int]:
    s = pd.to_numeric(y, errors="coerce").dropna()
    if len(s) < 20:
        return np.nan, np.nan, 0
    mid = len(s) // 2
    m1 = s.iloc[:mid].mean()
    m2 = s.iloc[mid:].mean()
    sd = s.std(ddof=1)
    drift = (m2 - m1) / sd if sd > 0 else 0.0
    flag = int(abs(drift) > 0.5)
    return float(m1), float(m2), flag


def run_tests_for_series(y: pd.Series, min_obs: int = 24) -> Dict[str, float]:
    s = pd.to_numeric(y, errors="coerce").dropna()
    if len(s) < min_obs:
        return {
            "n_obs": int(len(s)),
            "adf_p": np.nan,
            "kpss_p": np.nan,
            "ljungbox_p": np.nan,
            "bp_p": np.nan,
            "white_p": np.nan,
            "jb_p": np.nan,
            "stability_flag": 0,
            "stability_drift": np.nan,
        }

    if s.nunique(dropna=True) <= 1:
        return {
            "n_obs": int(len(s)),
            "adf_p": np.nan,
            "kpss_p": np.nan,
            "ljungbox_p": np.nan,
            "bp_p": np.nan,
            "white_p": np.nan,
            "jb_p": np.nan,
            "stability_flag": 0,
            "stability_drift": 0.0,
        }

    try:
        adf_p = adfuller(s, autolag="AIC")[1]
    except Exception:
        adf_p = np.nan
    try:
        kpss_p = kpss(s, regression="c", nlags="auto")[1]
    except Exception:
        kpss_p = np.nan

    s_pos = s.where(s > 0)
    dy = np.log(s_pos).diff().replace([np.inf, -np.inf], np.nan).dropna()
    if len(dy) < 10:
        lb_p = np.nan
        bp_p = np.nan
        white_p = np.nan
        jb_p = np.nan
    else:
        lb_p = float(acorr_ljungbox(dy, lags=[min(10, max(2, len(dy) // 5))], return_df=True)["lb_pvalue"].iloc[0])

        # Minimal lag-regression for residual diagnostics
        X = pd.DataFrame({"lag1": dy.shift(1), "lag2": dy.shift(2)}).dropna()
        yreg = dy.loc[X.index]
        if len(X) > 8:
            try:
                fit = sm.OLS(yreg, sm.add_constant(X)).fit()
                bp_p = float(het_breuschpagan(fit.resid, fit.model.exog)[1])
                white_p = float(het_white(fit.resid, fit.model.exog)[1])
                jb_p = float(jarque_bera(fit.resid)[1])
            except Exception:
                bp_p = np.nan
                white_p = np.nan
                jb_p = np.nan
        else:
            bp_p = np.nan
            white_p = np.nan
            jb_p = np.nan

    _, _, stab_flag = stability_flag(s)
    drift = stability_flag(s)[1] - stability_flag(s)[0] if len(s) >= 20 else np.nan

    return {
        "n_obs": int(len(s)),
        "adf_p": float(adf_p),
        "kpss_p": float(kpss_p) if pd.notna(kpss_p) else np.nan,
        "ljungbox_p": lb_p,
        "bp_p": bp_p,
        "white_p": white_p,
        "jb_p": jb_p,
        "stability_flag": int(stab_flag),
        "stability_drift": float(drift) if pd.notna(drift) else np.nan,
    }


def action_label_from_tests(r: pd.Series) -> str:
    notes = []
    adf_p = r.get("adf_p")
    kpss_p = r.get("kpss_p")
    lb_p = r.get("ljungbox_p")
    bp_p = r.get("bp_p")
    white_p = r.get("white_p")

    if pd.notna(adf_p) and pd.notna(kpss_p):
        if adf_p > 0.05 and kpss_p < 0.05:
            notes.append("non-stationary: difference/cointegration")
        elif adf_p < 0.05 and kpss_p > 0.05:
            notes.append("stationary: level models admissible")
        else:
            notes.append("mixed stationarity: check transformations")

    if pd.notna(lb_p) and lb_p < 0.05:
        notes.append("autocorrelation: add lags/respecify")
    if (pd.notna(bp_p) and bp_p < 0.05) or (pd.notna(white_p) and white_p < 0.05):
        notes.append("heteroskedasticity: robust SE")
    if int(r.get("stability_flag", 0)) == 1:
        notes.append("stability risk: rolling/break splits")

    return " | ".join(notes) if notes else "ok"


def compute_correlations(all_daily: pd.DataFrame) -> pd.DataFrame:
    if all_daily.empty:
        return pd.DataFrame()

    rows = []

    # Within-series: real vs linear vs pchip
    for (src, prod), g in all_daily.groupby(["source", "product"], dropna=False):
        g = g.sort_values("date")
        for a, b in [("price_real", "price_linear"), ("price_real", "price_pchip"), ("price_linear", "price_pchip")]:
            pair = g[[a, b]].dropna()
            if len(pair) >= 12:
                rows.append(
                    {
                        "corr_type": "within_variant",
                        "source_left": src,
                        "source_right": src,
                        "product": prod,
                        "lag": 0,
                        "freq": "daily",
                        "series_left": a,
                        "series_right": b,
                        "pearson": float(pair[a].corr(pair[b], method="pearson")),
                        "spearman": float(pair[a].corr(pair[b], method="spearman")),
                    }
                )

    # Between sources at product level, daily lags 0/7/14
    corr_base = all_daily.copy()
    corr_base["price_eff"] = corr_base["price_pchip"].where(corr_base["price_pchip"].notna(), corr_base["price_real"])
    lag_days = [0, 7, 14]
    pivot = corr_base.pivot_table(index=["date", "product", "standardized_type"], columns="source", values="price_eff", aggfunc="mean").reset_index()
    source_cols = [c for c in pivot.columns if c not in {"date", "product", "standardized_type"}]

    for prod, gp in pivot.groupby("product", dropna=False):
        gp = gp.sort_values("date")
        for i in range(len(source_cols)):
            for j in range(i + 1, len(source_cols)):
                s1 = source_cols[i]
                s2 = source_cols[j]
                base = gp[["date", s1, s2]].copy()
                for lag in lag_days:
                    tmp = base.copy()
                    if lag > 0:
                        tmp[s2] = tmp[s2].shift(lag)
                    pair = tmp[[s1, s2]].dropna()
                    if len(pair) < 20:
                        continue
                    rows.append(
                        {
                            "corr_type": "between_sources",
                            "source_left": s1,
                            "source_right": s2,
                            "product": prod,
                            "lag": lag,
                            "freq": "daily",
                            "series_left": "price_pchip",
                            "series_right": "price_pchip",
                            "pearson": float(pair[s1].corr(pair[s2], method="pearson")),
                            "spearman": float(pair[s1].corr(pair[s2], method="spearman")),
                        }
                    )

    # Weekly lagged correlations 0/1/2 weeks
    weekly = corr_base.copy()
    weekly["week"] = pd.to_datetime(weekly["date"]).dt.to_period("W-MON").dt.start_time
    wp = (
        weekly.groupby(["week", "product", "standardized_type", "source"], as_index=False)["price_eff"]
        .mean()
        .pivot_table(index=["week", "product", "standardized_type"], columns="source", values="price_eff", aggfunc="mean")
        .reset_index()
    )
    wsource_cols = [c for c in wp.columns if c not in {"week", "product", "standardized_type"}]
    lag_w = [0, 1, 2]
    for prod, gp in wp.groupby("product", dropna=False):
        gp = gp.sort_values("week")
        for i in range(len(wsource_cols)):
            for j in range(i + 1, len(wsource_cols)):
                s1 = wsource_cols[i]
                s2 = wsource_cols[j]
                base = gp[[s1, s2]].copy()
                for lag in lag_w:
                    tmp = base.copy()
                    if lag > 0:
                        tmp[s2] = tmp[s2].shift(lag)
                    pair = tmp[[s1, s2]].dropna()
                    if len(pair) < 12:
                        continue
                    rows.append(
                        {
                            "corr_type": "between_sources",
                            "source_left": s1,
                            "source_right": s2,
                            "product": prod,
                            "lag": lag,
                            "freq": "weekly",
                            "series_left": "price_pchip",
                            "series_right": "price_pchip",
                            "pearson": float(pair[s1].corr(pair[s2], method="pearson")),
                            "spearman": float(pair[s1].corr(pair[s2], method="spearman")),
                        }
                    )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["corr_type", "product", "source_left", "source_right", "lag"])


def rolling_mae_lastvalue(s: pd.Series, holdout: int = 30) -> float:
    y = pd.to_numeric(s, errors="coerce").dropna()
    if len(y) < holdout + 5:
        return np.nan
    y = y.iloc[-(holdout + 60) :]
    errs = []
    for t in range(1, min(holdout, len(y) - 1) + 1):
        y_true = y.iloc[-t]
        y_pred = y.iloc[-t - 1]
        errs.append(abs(y_true - y_pred))
    return float(np.mean(errs)) if errs else np.nan


def forecast_next_30(s: pd.Series) -> Tuple[float, str]:
    y = pd.to_numeric(s, errors="coerce").dropna()
    if len(y) < 10:
        return np.nan, "insufficient"

    # Simple robust baseline: weighted blend of last value and last-30 mean.
    last_val = float(y.iloc[-1])
    last30 = float(y.tail(30).mean())
    return float(0.6 * last_val + 0.4 * last30), "blend_last_last30"


def build_price_forecasts(all_daily: pd.DataFrame) -> pd.DataFrame:
    if all_daily.empty:
        return pd.DataFrame()

    rows = []
    long_daily = long_price_series(all_daily)
    if long_daily.empty:
        return pd.DataFrame()

    # Product-first forecasts by available time-series variant.
    for (src, prod, std, var), g in long_daily.groupby(["source", "product", "standardized_type", "series_variant"], dropna=False):
        g = g.sort_values("date")
        y = g["price"]
        if y.dropna().shape[0] < 20:
            continue
        fc, model_name = forecast_next_30(y)
        mae = rolling_mae_lastvalue(y, holdout=30)
        last_date = pd.to_datetime(g["date"]).max()
        for horizon_days in [7, 30]:
            for step in range(1, horizon_days + 1):
                rows.append(
                    {
                        "level": "product",
                        "source": src,
                        "product": prod,
                        "standardized_type": std,
                        "series_variant": var,
                        "horizon_days": horizon_days,
                        "forecast_step": step,
                        "forecast_date": last_date + pd.Timedelta(days=step),
                        "forecast_price": fc,
                        "rolling_mae": mae,
                        "model": model_name,
                    }
                )

    # Secondary view by standardized_type.
    for (src, std, var), g in long_daily.groupby(["source", "standardized_type", "series_variant"], dropna=False):
        ydf = g.groupby("date", as_index=False)["price"].mean().sort_values("date")
        y = ydf["price"]
        if y.dropna().shape[0] < 20:
            continue
        fc, model_name = forecast_next_30(y)
        mae = rolling_mae_lastvalue(y, holdout=30)
        last_date = pd.to_datetime(ydf["date"]).max()
        for horizon_days in [7, 30]:
            for step in range(1, horizon_days + 1):
                rows.append(
                    {
                        "level": "standardized_type",
                        "source": src,
                        "product": "",
                        "standardized_type": std,
                        "series_variant": var,
                        "horizon_days": horizon_days,
                        "forecast_step": step,
                        "forecast_date": last_date + pd.Timedelta(days=step),
                        "forecast_price": fc,
                        "rolling_mae": mae,
                        "model": model_name,
                    }
                )

    out = pd.DataFrame(rows)
    return out.sort_values(["level", "source", "standardized_type", "product"]) if not out.empty else out


def build_silpo_discount_forecasts(
    silpo_clean: pd.DataFrame,
    eu_daily: pd.DataFrame,
    producer_daily: pd.DataFrame,
) -> pd.DataFrame:
    if silpo_clean.empty:
        return pd.DataFrame()

    s = silpo_clean.copy()
    s["date"] = pd.to_datetime(s["date"], errors="coerce")
    s = s.dropna(subset=["date"]).copy()

    agg = (
        s.groupby(["date", "product", "standardized_type"], as_index=False)
        .agg(
            price_current=("price_current", "median"),
            discount_dummy_bulk=("discount_dummy_bulk", "mean"),
            discount_dummy_discount=("discount_dummy_discount", "mean"),
            discount_dummy_regular=("discount_dummy_regular", "mean"),
            discount_present=("discount_present", "mean"),
            discount_depth=("discount_depth", "mean"),
        )
        .sort_values(["product", "date"])
    )
    agg["baseline_price"] = np.where(
        agg["discount_depth"].notna() & (agg["discount_depth"] >= 0) & (agg["discount_depth"] < 95),
        agg["price_current"] / (1.0 - agg["discount_depth"] / 100.0),
        agg["price_current"],
    )
    agg["effective_price"] = agg["price_current"]

    # Determinants data
    eu_col = "price_pchip" if "price_pchip" in eu_daily.columns else "price_real"
    prod_col = "price_pchip" if "price_pchip" in producer_daily.columns else "price_real"
    if "price_pchip" in eu_daily.columns and "price_real" in eu_daily.columns:
        eu_daily = eu_daily.copy()
        eu_daily["price_eff"] = eu_daily["price_pchip"].where(eu_daily["price_pchip"].notna(), eu_daily["price_real"])
        eu_col = "price_eff"
    if "price_pchip" in producer_daily.columns and "price_real" in producer_daily.columns:
        producer_daily = producer_daily.copy()
        producer_daily["price_eff"] = producer_daily["price_pchip"].where(producer_daily["price_pchip"].notna(), producer_daily["price_real"])
        prod_col = "price_eff"

    eu_prod = eu_daily.groupby(["date", "product"], as_index=False)[eu_col].mean().rename(columns={eu_col: "eu_price"}) if not eu_daily.empty else pd.DataFrame(columns=["date", "product", "eu_price"])
    prod_prod = producer_daily.groupby(["date", "product"], as_index=False)[prod_col].mean().rename(columns={prod_col: "producer_price"}) if not producer_daily.empty else pd.DataFrame(columns=["date", "product", "producer_price"])

    out_rows = []
    for (prod, std), g in agg.groupby(["product", "standardized_type"], dropna=False):
        g = g.sort_values("date").copy()
        g = g.merge(eu_prod, on=["date", "product"], how="left")
        g = g.merge(prod_prod, on=["date", "product"], how="left")

        g["d_eu"] = np.log(g["eu_price"]).diff()
        g["d_prod"] = np.log(g["producer_price"]).diff()
        g["d_own"] = np.log(g["price_current"]).diff()
        g["dow"] = g["date"].dt.dayofweek
        g["month"] = g["date"].dt.month

        # Forecast core discount metrics
        targets = [
            "discount_dummy_bulk",
            "discount_dummy_discount",
            "discount_dummy_regular",
            "discount_present",
            "discount_depth",
            "baseline_price",
            "effective_price",
        ]

        for t in targets:
            y = pd.to_numeric(g[t], errors="coerce")
            fc, model_name = forecast_next_30(y)
            mae = rolling_mae_lastvalue(y, holdout=30)
            last_date = pd.to_datetime(g["date"]).max()
            for horizon_days in [7, 30]:
                for step in range(1, horizon_days + 1):
                    out_rows.append(
                        {
                            "section": "forecast",
                            "product": prod,
                            "standardized_type": std,
                            "target": t,
                            "horizon_days": horizon_days,
                            "forecast_step": step,
                            "forecast_date": last_date + pd.Timedelta(days=step),
                            "forecast_value": fc,
                            "rolling_mae": mae,
                            "model": model_name,
                            "term": "",
                            "coef": np.nan,
                            "pvalue": np.nan,
                        }
                    )

        # Determinants regressions for present/depth
        det = g[["discount_present", "discount_depth", "d_eu", "d_prod", "d_own", "dow", "month"]].copy()
        det = det.dropna()
        if len(det) >= 25:
            X = sm.add_constant(det[["d_eu", "d_prod", "d_own", "dow", "month"]])

            # Presence model (logit)
            yb = (det["discount_present"] > 0.05).astype(int)
            try:
                fit_b = sm.Logit(yb, X).fit(disp=False)
                for term in ["d_eu", "d_prod", "d_own", "dow", "month"]:
                    out_rows.append(
                        {
                            "section": "determinants_present",
                            "product": prod,
                            "standardized_type": std,
                            "target": "discount_present",
                            "horizon_days": 0,
                            "forecast_step": 0,
                            "forecast_date": pd.NaT,
                            "forecast_value": np.nan,
                            "rolling_mae": np.nan,
                            "model": "logit",
                            "term": term,
                            "coef": float(fit_b.params.get(term, np.nan)),
                            "pvalue": float(fit_b.pvalues.get(term, np.nan)),
                        }
                    )
            except Exception:
                pass

            # Depth model (OLS)
            yd = det["discount_depth"]
            try:
                fit_d = sm.OLS(yd, X).fit(cov_type="HC1")
                for term in ["d_eu", "d_prod", "d_own", "dow", "month"]:
                    out_rows.append(
                        {
                            "section": "determinants_depth",
                            "product": prod,
                            "standardized_type": std,
                            "target": "discount_depth",
                            "horizon_days": 0,
                            "forecast_step": 0,
                            "forecast_date": pd.NaT,
                            "forecast_value": np.nan,
                            "rolling_mae": float(np.sqrt(np.mean(fit_d.resid**2))),
                            "model": "ols_hc1",
                            "term": term,
                            "coef": float(fit_d.params.get(term, np.nan)),
                            "pvalue": float(fit_d.pvalues.get(term, np.nan)),
                        }
                    )
            except Exception:
                pass

    out = pd.DataFrame(out_rows)
    return out.sort_values(["section", "product", "target", "term"]) if not out.empty else out


def build_model_eligibility(
    tests_df: pd.DataFrame,
    weekly_df: pd.DataFrame,
    silpo_clean: pd.DataFrame,
    eu_weekly: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    weekly_pchip = weekly_df[weekly_df["series_variant"] == "pchip"].copy() if "series_variant" in weekly_df.columns else weekly_df

    products = sorted([p for p in weekly_pchip["product"].dropna().unique() if str(p).strip()])
    for prod in products:
        subset = weekly_pchip[weekly_pchip["product"] == prod]
        std = subset["standardized_type"].dropna().mode().iloc[0] if subset["standardized_type"].notna().any() else "other"

        # Aggregate product weekly for UA side
        ua = (
            subset[subset["source"].isin(["ProducerUA", "ConsumerUA", "ProZorro", "Silpo", "Novus"])]
            .groupby("week", as_index=False)["price_weekly"]
            .mean()
            .rename(columns={"price_weekly": "ua_price"})
        )
        bm = (
            eu_weekly[eu_weekly["product"] == prod][["week", "price_weekly"]]
            .rename(columns={"price_weekly": "eu_price"})
            .copy()
        )
        pair = ua.merge(bm, on="week", how="inner").dropna()

        cointegr_ok = 0
        ecm_vecm_ok = 0
        if len(pair) >= 35:
            y = np.log(pair["ua_price"])
            x = np.log(pair["eu_price"])
            try:
                adf_y = adfuller(y, autolag="AIC")[1]
                adf_dy = adfuller(y.diff().dropna(), autolag="AIC")[1]
                coint_p = coint(y, x)[1]
                cointegr_ok = int((adf_y > 0.05) and (adf_dy < 0.05) and (coint_p < 0.10))
                ecm_vecm_ok = int(coint_p < 0.10 and len(pair) >= 50)
            except Exception:
                cointegr_ok = 0
                ecm_vecm_ok = 0

        tsub = tests_df[(tests_df["product"] == prod) & (tests_df["series_variant"] == "pchip")] if "series_variant" in tests_df.columns else tests_df[tests_df["product"] == prod]
        robust_needed = int(((tsub["bp_p"] < 0.05) | (tsub["white_p"] < 0.05)).fillna(False).any()) if not tsub.empty else 0
        unstable = int((tsub["stability_flag"] == 1).any()) if not tsub.empty else 0
        rec_family = tsub["recommended_model_family"].mode().iloc[0] if (not tsub.empty and "recommended_model_family" in tsub.columns) else ""
        avoid = tsub["avoid"].mode().iloc[0] if (not tsub.empty and "avoid" in tsub.columns) else ""
        diag_note = tsub["diag_note"].mode().iloc[0] if (not tsub.empty and "diag_note" in tsub.columns) else ""

        sil = silpo_clean[silpo_clean["product"] == prod] if not silpo_clean.empty else pd.DataFrame()
        promo_share = float(pd.to_numeric(sil.get("discount_present", pd.Series(dtype=float)), errors="coerce").fillna(0).mean()) if not sil.empty else 0.0
        promo_confound = "high" if promo_share >= 0.25 else "low"

        rows.append(
            {
                "product": prod,
                "standardized_type": std,
                "eligible_for_cointegration": cointegr_ok,
                "ecm_vecm_feasible": ecm_vecm_ok,
                "need_robust_se": robust_needed,
                "promotions_confounding": promo_confound,
                "stability_risk": unstable,
                "recommended_model_family": rec_family,
                "avoid": avoid,
                "diag_note": diag_note,
                "note": "product-first admissibility summary",
            }
        )

    return pd.DataFrame(rows).sort_values(["standardized_type", "product"]) if rows else pd.DataFrame()


def read_sheet_or_empty(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    return xls.parse(sheet_name) if sheet_name in xls.sheet_names else pd.DataFrame()


def prep_producer_consumer(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    req = {
        "date": ["date", "дата"],
        "ua_product": ["ua_product", "product", "товар", "розріз"],
        "price_linear": ["price_linear"],
        "price_pchip": ["price_pchip", "price_chip"],
    }
    cmap = require_columns(df, req, source)
    out = df.rename(columns={
        cmap["date"]: "date",
        cmap["ua_product"]: "raw_product",
        cmap["price_linear"]: "price_linear_raw",
        cmap["price_pchip"]: "price_pchip_raw",
    }).copy()
    price_real_col = find_column(df, ["price_real", "price"])
    if price_real_col:
        out["price_real_raw"] = pd.to_numeric(df[price_real_col], errors="coerce")
    else:
        out["price_real_raw"] = np.nan

    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date"]).copy()
    out["product"] = out["raw_product"].map(lambda x: detect_product_name(x))
    out["standardized_type"] = out["raw_product"].map(lambda x: detect_standardized_type(x))
    out["brand"] = ""
    out["region"] = df[find_column(df, ["region", "регіон", "територіальний розріз"])].astype(str) if find_column(df, ["region", "регіон", "територіальний розріз"]) else "Україна"

    # Producer/Consumer are pre-processed in UAH and provide linear/pchip directly.
    price_linear = pd.to_numeric(out["price_linear_raw"], errors="coerce")
    pchip = pd.to_numeric(out["price_pchip_raw"], errors="coerce")
    price_real = pd.to_numeric(out["price_real_raw"], errors="coerce")
    out["price"] = price_linear.where(price_linear.notna(), pchip).where(lambda s: s.notna(), price_real)
    out["price_real_input"] = price_real
    out["price_linear_input"] = price_linear
    out["price_pchip_input"] = pchip
    out["unit_ok"] = 1
    out["source"] = source
    return out[
        [
            "source",
            "date",
            "raw_product",
            "product",
            "standardized_type",
            "brand",
            "region",
            "price",
            "price_real_input",
            "price_linear_input",
            "price_pchip_input",
            "unit_ok",
        ]
    ]


def prep_eu(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    req = {"date": ["date", "дата"], "product": ["product"]}
    cmap = require_columns(df, req, "EU")

    out = df.rename(columns={cmap["date"]: "date", cmap["product"]: "raw_product"}).copy()
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date"]).copy()
    out["product"] = out["raw_product"].map(lambda x: detect_product_name(x))
    out["standardized_type"] = out["raw_product"].map(lambda x: detect_standardized_type(x))
    out["brand"] = ""
    out["region"] = "EU"

    price_uah_col = find_column(df, ["price (uah/kg)", "price uah/kg", "uah/kg"])
    if price_uah_col is None:
        raise ValueError("[EU] Missing required column: 'Price (UAH/kg)'.")
    out["price"] = pd.to_numeric(df[price_uah_col], errors="coerce")

    out["unit_ok"] = 1
    out["source"] = "EU"
    return out[["source", "date", "raw_product", "product", "standardized_type", "brand", "region", "price", "unit_ok"]]


def _unit_to_norm_multiplier(unit: str) -> Tuple[Optional[float], Optional[str]]:
    u = ntext(unit)
    if u in KG_UNITS:
        return 1.0, "kg"
    if u in G_UNITS:
        return 0.001, "kg"
    if u in L_UNITS:
        return 1.0, "liter"
    if u in ML_UNITS:
        return 0.001, "liter"
    if u in PCS_UNITS:
        return None, "piece"
    return None, None


def prep_prozorro(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    req = {
        "date": ["дата", "date"],
        "product": ["product", "профіль", "profile"],
        "title": ["товар", "title", "назва"],
        "unit_price": ["ціна за одиницю", "unit price"],
        "region": ["регіон організатора", "region"],
        "expected": ["очікувана вартість", "expected"],
        "sum_initial": ["сума договорів початкова", "sum contracts initial"],
        "sum_current": ["сума договорів поточна", "sum contracts current"],
    }
    cmap = require_columns(df, req, "ProZorro")

    out = df.rename(
        columns={
            cmap["date"]: "date",
            cmap["product"]: "raw_product",
            cmap["title"]: "title",
            cmap["unit_price"]: "unit_price",
            cmap["region"]: "region",
            cmap["expected"]: "expected",
            cmap["sum_initial"]: "sum_initial",
            cmap["sum_current"]: "sum_current",
        }
    ).copy()
    qty_col = find_column(df, ["кількість", "quantity"])
    unit_col = find_column(df, ["одиниця виміру", "unit"])
    out["qty"] = pd.to_numeric(df[qty_col], errors="coerce") if qty_col else np.nan
    out["unit"] = df[unit_col].astype(str) if unit_col else ""

    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date"]).copy()
    out["product"] = out.apply(lambda r: detect_product_name(r.get("raw_product"), r.get("title")), axis=1)
    out["standardized_type"] = out.apply(lambda r: detect_standardized_type(r.get("raw_product"), r.get("title")), axis=1)
    out["brand"] = ""

    out["unit_price"] = pd.to_numeric(out["unit_price"], errors="coerce")
    out["expected"] = pd.to_numeric(out["expected"], errors="coerce")
    out["sum_initial"] = pd.to_numeric(out["sum_initial"], errors="coerce")
    out["sum_current"] = pd.to_numeric(out["sum_current"], errors="coerce")

    # Use pre-transformed UAH unit price directly (no currency conversion here).
    out["price"] = out["unit_price"]

    # fallback from expected / qty if unit price missing
    fallback_unit_price = out["expected"] / out["qty"].replace(0, np.nan)
    out["price"] = out["price"].where(out["price"].notna(), fallback_unit_price)

    out["unit_ok"] = out["price"].notna().astype(int)
    out["source"] = "ProZorro"

    return out[
        [
            "source",
            "date",
            "raw_product",
            "title",
            "product",
            "standardized_type",
            "brand",
            "region",
            "qty",
            "unit",
            "unit_price",
            "expected",
            "sum_initial",
            "sum_current",
            "price",
            "unit_ok",
        ]
    ]


def prep_retail(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    req = {
        "date": ["date", "дата"],
        "title": ["product_title", "title", "назва товару"],
        "product": ["product", "product category", "категорія"],
        "brand": ["brand", "бренд"],
        "price_current": ["price_current", "current price", "ціна поточна"],
        "unit_price": ["unit_price", "price per unit"],
    }
    cmap = require_columns(df, req, source)
    ts_col = find_column(df, ["timestamp", "time", "час"])

    out = df.rename(
        columns={
            cmap["date"]: "date",
            cmap["title"]: "title",
            cmap["product"]: "raw_product",
            cmap["brand"]: "brand",
            cmap["price_current"]: "price_current",
            cmap["unit_price"]: "unit_price",
        }
    ).copy()
    out["timestamp"] = pd.to_datetime(df[ts_col], errors="coerce") if ts_col else pd.NaT

    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date"]).copy()

    sku_col = find_column(df, ["id", "sku", "артикул", "product_id"])
    if sku_col:
        out["sku_id"] = df.loc[out.index, sku_col].astype(str)
    else:
        out["sku_id"] = out["title"].astype(str).str.lower().str.replace(r"\s+", " ", regex=True)

    # Keep latest within day key; for Novus prioritize (date,sku_id) dedupe.
    out = out.sort_values(["date", "sku_id", "timestamp"])
    if source == "Novus":
        out = out.drop_duplicates(subset=["date", "sku_id"], keep="last")
    else:
        out = out.drop_duplicates(subset=["date", "raw_product", "title", "brand"], keep="last")

    out["product"] = out.apply(lambda r: detect_product_name(r.get("raw_product"), r.get("title")), axis=1)
    out["standardized_type"] = out.apply(lambda r: detect_standardized_type(r.get("raw_product"), r.get("title")), axis=1)

    out["price_current"] = pd.to_numeric(out["price_current"], errors="coerce")
    out["unit_price"] = pd.to_numeric(out["unit_price"], errors="coerce")

    # Try pack extraction if unit_price missing
    pack_info = out["title"].map(parse_pack_size_kg_or_l)
    out["pack_qty_norm"] = pack_info.map(lambda x: x[0])
    out["pack_unit_norm"] = pack_info.map(lambda x: x[1])
    out["pack_qty"] = out["pack_qty_norm"]
    out["pack_unit"] = out["pack_unit_norm"]

    # Prefer provided unit_price in UAH; fallback to price_current level if missing.
    out["price"] = out["unit_price"]
    out["price"] = out["price"].where(out["price"].notna(), out["price_current"])
    out["unit_price_uah_per_kg_or_l"] = out["price"]

    out["unit_ok"] = out["price"].notna().astype(int)
    out["region"] = "Україна"

    # discount fields
    for c in ["discount_dummy_bulk", "discount_dummy_discount", "discount_dummy_regular", "discount_present", "discount_depth", "price_old"]:
        found = find_column(df, [c])
        if found:
            out[c] = pd.to_numeric(df.loc[out.index, found], errors="coerce")
        else:
            out[c] = np.nan

    # Generate discount_present/depth if missing
    if out["discount_present"].isna().all():
        out["discount_present"] = (pd.to_numeric(out["discount_dummy_discount"], errors="coerce").fillna(0) > 0).astype(int)

    if out["discount_depth"].isna().all():
        if out["price_old"].notna().any():
            out["discount_depth"] = np.where(
                out["price_old"] > 0,
                (out["price_old"] - out["price_current"]) / out["price_old"] * 100.0,
                np.nan,
            )
        else:
            out["discount_depth"] = 0.0

    out["source"] = source

    cols = [
        "source",
        "sku_id",
        "timestamp",
        "date",
        "title",
        "raw_product",
        "product",
        "standardized_type",
        "brand",
        "region",
        "price_current",
        "price_old",
        "price",
        "unit_price_uah_per_kg_or_l",
        "pack_qty",
        "pack_unit",
        "unit_ok",
        "discount_dummy_bulk",
        "discount_dummy_discount",
        "discount_dummy_regular",
        "discount_present",
        "discount_depth",
    ]
    return out[cols]


def prep_cme(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    req = {
        "date": ["date", "дата"],
        "price": ["cme iii uah", "price", "ціна"],
    }
    cmap = require_columns(df, req, "CME")
    out = df.rename(columns={cmap["date"]: "date", cmap["price"]: "price"}).copy()
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out["price"] = pd.to_numeric(out["price"], errors="coerce")
    out = out.dropna(subset=["date"]).copy()
    out["raw_product"] = "CME Class III"
    out["product"] = "Молоко питне"
    out["standardized_type"] = "milk"
    out["brand"] = ""
    out["region"] = "US"
    out["unit_ok"] = 1
    out["source"] = "CME"
    return out[["source", "date", "raw_product", "product", "standardized_type", "brand", "region", "price", "unit_ok"]]


def ensure_numeric_price(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out["price"] = pd.to_numeric(out.get("price"), errors="coerce")
    return out


def categorisation_table(datasets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    summary_rows = []
    for name, df in datasets.items():
        if df.empty:
            summary_rows.append(
                {
                    "dataset": name,
                    "rows_total": 0,
                    "classified_product_rows": 0,
                    "classified_stdtype_rows": 0,
                    "other_or_unknown_rows": 0,
                    "coverage_product": np.nan,
                    "coverage_standardized_type": np.nan,
                }
            )
            continue

        prod_ok = (df["product"].astype(str) != "Інше/невідомо").sum()
        std_ok = (df["standardized_type"].astype(str) != "other").sum()
        n = len(df)
        summary_rows.append(
            {
                "dataset": name,
                "rows_total": n,
                "classified_product_rows": int(prod_ok),
                "classified_stdtype_rows": int(std_ok),
                "other_or_unknown_rows": int(n - min(prod_ok, std_ok)),
                "coverage_product": float(prod_ok / n) if n else np.nan,
                "coverage_standardized_type": float(std_ok / n) if n else np.nan,
            }
        )

    mapping_rows = []
    for name, df in datasets.items():
        if df.empty:
            continue
        cols = [c for c in ["raw_product", "title", "product", "standardized_type", "brand"] if c in df.columns]
        g = (
            df[cols]
            .fillna("")
            .groupby(cols, dropna=False, as_index=False)
            .size()
            .rename(columns={"size": "rows"})
        )
        g.insert(0, "dataset", name)
        mapping_rows.append(g)

    detail_rows = []
    for name, df in datasets.items():
        if df.empty:
            continue
        sample_unknown = df[(df["product"] == "Інше/невідомо") | (df["standardized_type"] == "other")].head(15)
        for _, r in sample_unknown.iterrows():
            detail_rows.append(
                {
                    "dataset": name,
                    "raw_product": r.get("raw_product", ""),
                    "title": r.get("title", ""),
                    "product": r.get("product", ""),
                    "standardized_type": r.get("standardized_type", ""),
                    "brand": r.get("brand", ""),
                    "note": "sample unresolved",
                }
            )

    summary = pd.DataFrame(summary_rows)
    mapping = pd.concat(mapping_rows, ignore_index=True) if mapping_rows else pd.DataFrame()
    details = pd.DataFrame(detail_rows)
    if mapping.empty and details.empty:
        return summary
    chunks = [summary]
    if not mapping.empty:
        chunks.append(pd.DataFrame([{"dataset": "--- mapping raw->product below ---"}]))
        chunks.append(mapping)
    if not details.empty:
        chunks.append(pd.DataFrame([{"dataset": "--- unresolved samples below ---"}]))
        chunks.append(details)
    return pd.concat(chunks, ignore_index=True, sort=False)


def build_tests_table(weekly_df: pd.DataFrame, min_obs: int) -> pd.DataFrame:
    if weekly_df.empty:
        return pd.DataFrame()

    rows = []
    for (src, prod, std, variant), g in weekly_df.groupby(
        ["source", "product", "standardized_type", "series_variant"], dropna=False
    ):
        y = g.sort_values("week")["price_weekly"]
        tests = run_tests_for_series(y, min_obs=min_obs)
        row = {
            "source": src,
            "product": prod,
            "standardized_type": std,
            "series_variant": variant,
            **tests,
        }
        rows.append(row)

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["action_label"] = out.apply(action_label_from_tests, axis=1)
    diag = out.apply(interpret_diagnostics, axis=1, result_type="expand")
    out = pd.concat([out, diag], axis=1)
    return out.sort_values(["source", "product", "series_variant"])


def interpret_diagnostics(row: pd.Series) -> Dict[str, str]:
    adf_p = row.get("adf_p")
    kpss_p = row.get("kpss_p")
    lb_p = row.get("ljungbox_p")
    bp_p = row.get("bp_p")
    white_p = row.get("white_p")
    jb_p = row.get("jb_p")
    stab = int(row.get("stability_flag", 0))

    if pd.notna(adf_p) and pd.notna(kpss_p):
        if adf_p > 0.05 and kpss_p < 0.05:
            integration_class = "I(1)-like"
        elif adf_p < 0.05 and kpss_p > 0.05:
            integration_class = "I(0)"
        else:
            integration_class = "ambiguous"
    else:
        integration_class = "ambiguous"

    ac_risk = "high" if pd.notna(lb_p) and lb_p < 0.05 else "low"
    het_risk = "high" if ((pd.notna(bp_p) and bp_p < 0.05) or (pd.notna(white_p) and white_p < 0.05)) else "low"
    non_normal_risk = "high" if pd.notna(jb_p) and jb_p < 0.05 else "low"
    stability_risk = "high" if stab == 1 else "low"

    rec = []
    if integration_class in {"I(1)-like", "ambiguous"}:
        rec.append("Use differences or cointegration framework")
    if ac_risk == "high":
        rec.append("Increase lags / AR structure")
    if het_risk == "high":
        rec.append("Use robust SE / HAC")
    if stability_risk == "high":
        rec.append("Use rolling / split sample / add break dummies")

    pattern = (
        integration_class == "I(1)-like"
        and ac_risk == "high"
        and het_risk == "high"
        and non_normal_risk == "high"
        and stability_risk == "high"
    )
    if pattern:
        model_family = "ARDL/ECM/VECM (cointegration framework) + HAC SE + rolling robustness"
        avoid = "OLS in levels without cointegration; naive asymmetry without promo controls"
        note = "pattern consistent with price series; interpret coefficients as dynamic/conditional"
    else:
        model_family = "Diagnostics-dependent"
        avoid = "Unvalidated level OLS under integration uncertainty"
        note = "Follow diagnostics-driven specification and robust inference."

    return {
        "integration_class": integration_class,
        "ac_risk": ac_risk,
        "het_risk": het_risk,
        "non_normal_risk": non_normal_risk,
        "stability_risk_class": stability_risk,
        "recommended_action": " | ".join(rec) if rec else "No immediate adjustment",
        "recommended_model_family": model_family,
        "avoid": avoid,
        "diag_note": note,
    }


def build_dataset_registry(raw_by_sheet: Dict[str, pd.DataFrame], cleaned: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    sheet_map = {
        "ProducerUA": "Producer_UA",
        "ConsumerUA": "Consumer_UA",
        "EU": "Europe",
        "ProZorro": "Prozorro",
        "Silpo": "Silpo",
        "Novus": "Novus",
        "CME": "CME III",
    }
    for name, sheet in sheet_map.items():
        raw = raw_by_sheet.get(sheet, pd.DataFrame())
        cln = cleaned.get(name, pd.DataFrame())
        rows.append(
            {
                "name": name,
                "sheet_name": sheet,
                "date_column": "date" if "date" in cln.columns else "Date",
                "price_column": "price",
                "product_column": "product",
                "extra_columns": ", ".join([c for c in ["region", "brand", "unit_ok", "expected", "sum_initial", "sum_current", "discount_present", "discount_depth"] if c in cln.columns]),
                "rows_raw": len(raw),
                "rows_clean": len(cln),
            }
        )
    return pd.DataFrame(rows)


def build_product_mapping_table(cleaned: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for source, df in cleaned.items():
        if df.empty:
            continue
        label_col = "raw_product" if "raw_product" in df.columns else "product"
        g = (
            df[[label_col, "product", "standardized_type"]]
            .fillna("")
            .groupby([label_col, "product", "standardized_type"], as_index=False)
            .size()
            .rename(columns={label_col: "raw_label", "size": "rows"})
        )
        g.insert(0, "source", source)
        rows.append(g)
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def build_model_series(weekly_df: pd.DataFrame) -> pd.DataFrame:
    if weekly_df.empty:
        return pd.DataFrame()
    selected = []
    for source in weekly_df["source"].dropna().unique():
        v = "pchip" if source in {"ProducerUA", "ConsumerUA", "FarmGateUA_initial", "FarmGateUA_filled"} else "real"
        ss = weekly_df[(weekly_df["source"] == source) & (weekly_df["series_variant"] == v)].copy()
        if ss.empty:
            ss = weekly_df[(weekly_df["source"] == source) & (weekly_df["series_variant"] == "real")].copy()
        selected.append(ss)
    out = pd.concat(selected, ignore_index=True) if selected else pd.DataFrame()
    if out.empty:
        return out
    out = (
        out.groupby(["standardized_type", "source", "week"], as_index=False)["price_weekly"]
        .mean()
        .rename(columns={"price_weekly": "price"})
    )
    return out.sort_values(["standardized_type", "source", "week"])


def build_model_plan(mapping_table: pd.DataFrame, model_series: pd.DataFrame) -> pd.DataFrame:
    if model_series.empty:
        return pd.DataFrame()
    rows = []
    for std in sorted(model_series["standardized_type"].dropna().unique()):
        ss = model_series[model_series["standardized_type"] == std]
        sources = sorted(ss["source"].dropna().unique().tolist())
        has = {s: (s in sources) for s in ["ProducerUA", "ConsumerUA", "EU", "CME", "ProZorro", "Silpo", "Novus"]}
        ua_core = "ProducerUA" + ("+ConsumerUA" if has["ConsumerUA"] else "")
        benchmark = "+".join([x for x in ["EU", "CME"] if has[x]])
        procurement = "ProZorro(unit_price+expected+initial+current+region)" if has["ProZorro"] else ""
        retail = " + ".join([x for x in ["Silpo(price_type+discounts)", "Novus(timestamp_dedupe)"] if (("Silpo" in x and has["Silpo"]) or ("Novus" in x and has["Novus"]))])

        possible = []
        if has["EU"] and (has["ProducerUA"] or has["ConsumerUA"] or has["ProZorro"] or has["Silpo"]):
            possible.append("ARDL")
            possible.append("NARDL")
        if has["EU"] and (has["ProducerUA"] or has["ConsumerUA"]):
            possible.append("ECM")
        if has["EU"] and sum([has["ProducerUA"], has["ConsumerUA"], has["ProZorro"], has["Silpo"], has["Novus"], has["CME"]]) >= 2:
            possible.append("VECM")
        if has["Silpo"] or has["ProZorro"]:
            possible.append("NEIO_proxy")

        rows.append(
            {
                "standardized_type": std,
                "bundle_ua_core": ua_core,
                "bundle_benchmark": benchmark,
                "bundle_procurement": procurement,
                "bundle_retail": retail,
                "available_series": ", ".join(sources),
                "possible_models": ", ".join(sorted(set(possible))),
                "n_weeks_min": int(ss.groupby("source")["week"].nunique().min()),
                "n_weeks_max": int(ss.groupby("source")["week"].nunique().max()),
            }
        )
    return pd.DataFrame(rows)


def _merge_model_pair(model_series: pd.DataFrame, std: str, y_source: str, x_source: str) -> pd.DataFrame:
    y = model_series[(model_series["standardized_type"] == std) & (model_series["source"] == y_source)][["week", "price"]].rename(columns={"price": "y"})
    x = model_series[(model_series["standardized_type"] == std) & (model_series["source"] == x_source)][["week", "price"]].rename(columns={"price": "x"})
    d = y.merge(x, on="week", how="inner").sort_values("week").dropna()
    d = d[(d["y"] > 0) & (d["x"] > 0)]
    return d


def _fit_ardl_like_pair(d: pd.DataFrame, max_p: int = 3, max_q: int = 3) -> Optional[Dict[str, float]]:
    if len(d) < 28:
        return None
    best = None
    for p in range(1, max_p + 1):
        for q in range(0, max_q + 1):
            tmp = d.copy()
            for i in range(1, p + 1):
                tmp[f"y_l{i}"] = np.log(tmp["y"]).shift(i)
            for j in range(0, q + 1):
                tmp[f"x_l{j}"] = np.log(tmp["x"]).shift(j)
            tmp = tmp.dropna()
            if len(tmp) < 20:
                continue
            y = np.log(tmp["y"])
            X = sm.add_constant(tmp[[c for c in tmp.columns if c.startswith("y_l") or c.startswith("x_l")]])
            fit = sm.OLS(y, X).fit()
            if best is None or fit.aic < best["aic"]:
                best = {"fit": fit, "p": p, "q": q, "tmp": tmp, "aic": float(fit.aic), "bic": float(fit.bic)}
    if best is None:
        return None
    fit = best["fit"]
    params = fit.params
    phi = float(np.sum([params.get(f"y_l{i}", 0.0) for i in range(1, best["p"] + 1)]))
    beta = float(np.sum([params.get(f"x_l{j}", 0.0) for j in range(0, best["q"] + 1)]))
    long_run = beta / (1.0 - phi) if abs(1.0 - phi) > 1e-6 else np.nan
    lb = float(acorr_ljungbox(fit.resid, lags=[min(10, max(2, len(fit.resid) // 5))], return_df=True)["lb_pvalue"].iloc[0])
    bp = float(het_breuschpagan(fit.resid, fit.model.exog)[1])
    wh = float(het_white(fit.resid, fit.model.exog)[1])
    return {
        "p": best["p"],
        "q": best["q"],
        "aic": best["aic"],
        "bic": best["bic"],
        "short_run_coef_x0": float(params.get("x_l0", np.nan)),
        "long_run_coef": float(long_run) if pd.notna(long_run) else np.nan,
        "lb_p": lb,
        "bp_p": bp,
        "white_p": wh,
        "n_obs": int(len(best["tmp"])),
    }


def build_ardl_summary(model_series: pd.DataFrame) -> pd.DataFrame:
    if model_series.empty:
        return pd.DataFrame()
    rows = []
    for std in sorted(model_series["standardized_type"].dropna().unique()):
        sources = set(model_series[model_series["standardized_type"] == std]["source"].unique().tolist())
        benchmark = "EU" if "EU" in sources else ("CME" if "CME" in sources else None)
        if benchmark is None:
            continue
        for y_source in ["ProducerUA", "ConsumerUA", "ProZorro", "Silpo", "Novus"]:
            if y_source not in sources:
                continue
            d = _merge_model_pair(model_series, std, y_source, benchmark)
            if len(d) < 28:
                continue
            fit = _fit_ardl_like_pair(d)
            if fit is None:
                continue
            try:
                coint_p = float(coint(np.log(d["y"]), np.log(d["x"]))[1])
            except Exception:
                coint_p = np.nan
            rows.append(
                {
                    "standardized_type": std,
                    "y_series_source": y_source,
                    "x_series_sources": benchmark,
                    "frequency": "weekly",
                    "sample_period": f"{d['week'].min().date()}..{d['week'].max().date()}",
                    "lags_selected": f"p={fit['p']}, q={fit['q']}",
                    "short_run_coef": fit["short_run_coef_x0"],
                    "long_run_coef": fit["long_run_coef"],
                    "coint_or_bounds_p": coint_p,
                    "diagnostics_flags": f"LB={fit['lb_p']:.3g}; BP={fit['bp_p']:.3g}; White={fit['white_p']:.3g}",
                    "eligible_for_ecm_form": "yes" if pd.notna(coint_p) and coint_p < 0.10 else "no",
                    "interpretation_note": "ARDL captures short- and long-run transmission under I(0)/I(1)-mix assumptions.",
                }
            )
    return pd.DataFrame(rows)


def build_ecm_summary(model_series: pd.DataFrame, ardl_summary: pd.DataFrame) -> pd.DataFrame:
    if model_series.empty or ardl_summary.empty:
        return pd.DataFrame()
    rows = []
    for _, r in ardl_summary.iterrows():
        if str(r.get("eligible_for_ecm_form", "no")) != "yes":
            continue
        std = r["standardized_type"]
        y_source = r["y_series_source"]
        x_source = r["x_series_sources"]
        d = _merge_model_pair(model_series, std, y_source, x_source)
        if len(d) < 30:
            continue
        lny = np.log(d["y"])
        lnx = np.log(d["x"])
        try:
            long_fit = sm.OLS(lny, sm.add_constant(lnx)).fit()
        except Exception:
            continue
        ect = long_fit.resid.shift(1)
        dy = lny.diff()
        dx = lnx.diff()
        tmp = pd.DataFrame({"dy": dy, "dx": dx, "ect": ect}).dropna()
        if len(tmp) < 20:
            continue
        fit = sm.OLS(tmp["dy"], sm.add_constant(tmp[["dx", "ect"]])).fit(cov_type="HAC", cov_kwds={"maxlags": 2})
        rows.append(
            {
                "standardized_type": std,
                "y_series_source": y_source,
                "x_series_sources": x_source,
                "frequency": "weekly",
                "sample_period": f"{d['week'].min().date()}..{d['week'].max().date()}",
                "lags_selected": "dy(1), dx(0), ect(1)",
                "short_run_coef": float(fit.params.get("dx", np.nan)),
                "long_run_coef": float(long_fit.params.get("x", np.nan)) if "x" in long_fit.params else float(long_fit.params.iloc[-1]),
                "ect_coef": float(fit.params.get("ect", np.nan)),
                "ect_pvalue": float(fit.pvalues.get("ect", np.nan)),
                "diagnostics_flags": "HAC SE applied",
                "interpretation_note": "Negative significant ECT indicates convergence to long-run equilibrium.",
            }
        )
    return pd.DataFrame(rows)


def build_nardl_summary(model_series: pd.DataFrame) -> pd.DataFrame:
    if model_series.empty:
        return pd.DataFrame()
    rows = []
    for std in sorted(model_series["standardized_type"].dropna().unique()):
        sources = set(model_series[model_series["standardized_type"] == std]["source"].unique().tolist())
        benchmark = "EU" if "EU" in sources else ("CME" if "CME" in sources else None)
        if benchmark is None:
            continue
        for y_source in ["ProducerUA", "ConsumerUA", "ProZorro", "Silpo"]:
            if y_source not in sources:
                continue
            d = _merge_model_pair(model_series, std, y_source, benchmark)
            if len(d) < 35:
                continue
            lny = np.log(d["y"])
            lnx = np.log(d["x"])
            dx = lnx.diff().dropna()
            x_pos = dx.clip(lower=0).cumsum()
            x_neg = dx.clip(upper=0).cumsum()
            dy = lny.diff()
            tmp = pd.DataFrame({"dy": dy, "dy_l1": dy.shift(1), "x_pos": x_pos, "x_neg": x_neg}).dropna()
            if len(tmp) < 20:
                continue
            fit = sm.OLS(tmp["dy"], sm.add_constant(tmp[["dy_l1", "x_pos", "x_neg"]])).fit(cov_type="HAC", cov_kwds={"maxlags": 2})
            try:
                short_p = float(fit.f_test("x_pos = x_neg").pvalue)
            except Exception:
                short_p = np.nan
            lvl = pd.DataFrame({"y": lny.loc[tmp.index], "x_pos": x_pos.loc[tmp.index], "x_neg": x_neg.loc[tmp.index]}).dropna()
            long_p = np.nan
            long_diff = np.nan
            if len(lvl) > 20:
                lfit = sm.OLS(lvl["y"], sm.add_constant(lvl[["x_pos", "x_neg"]])).fit()
                try:
                    long_p = float(lfit.f_test("x_pos = x_neg").pvalue)
                except Exception:
                    long_p = np.nan
                long_diff = float(lfit.params.get("x_pos", np.nan) - lfit.params.get("x_neg", np.nan))
            rows.append(
                {
                    "standardized_type": std,
                    "y_series_source": y_source,
                    "x_series_sources": benchmark,
                    "frequency": "weekly",
                    "sample_period": f"{d['week'].min().date()}..{d['week'].max().date()}",
                    "lags_selected": "dy(1), x_pos, x_neg",
                    "short_run_coef": float(fit.params.get("x_pos", np.nan)),
                    "long_run_coef": long_diff,
                    "asymmetry_short_p": short_p,
                    "asymmetry_long_p": long_p,
                    "diagnostics_flags": "HAC SE applied",
                    "interpretation_note": "Low p-values indicate asymmetric transmission to positive vs negative shocks.",
                }
            )
    return pd.DataFrame(rows)


def build_vecm_summary(model_series: pd.DataFrame) -> pd.DataFrame:
    if model_series.empty:
        return pd.DataFrame()
    systems = {
        "System A": ["ProducerUA", "ConsumerUA", "EU"],
        "System B": ["ProZorro", "ProducerUA", "EU"],
        "System C": ["Silpo", "Novus", "EU"],
        "System D": ["ProducerUA", "EU", "CME"],
    }
    rows = []
    for std in sorted(model_series["standardized_type"].dropna().unique()):
        sub = model_series[model_series["standardized_type"] == std]
        for sys_name, cols in systems.items():
            wide = (
                sub[sub["source"].isin(cols)]
                .pivot_table(index="week", columns="source", values="price", aggfunc="mean")
                .sort_index()
            )
            if not set(cols).issubset(set(wide.columns)):
                continue
            wide = wide[cols].dropna()
            wide = wide[(wide > 0).all(axis=1)]
            if len(wide) < 40:
                continue
            lw = np.log(wide)
            try:
                sel = select_order(lw, maxlags=4, deterministic="ci")
                k = sel.aic if sel.aic is not None else 2
                k_diff = max(1, int(k) - 1)
                rank = int(select_coint_rank(lw, det_order=0, k_ar_diff=k_diff, signif=0.10).rank)
                if rank < 1:
                    continue
                fit = VECM(lw, k_ar_diff=k_diff, coint_rank=rank, deterministic="ci").fit()
                cvec = "; ".join([f"{v:.3f}" for v in fit.beta[:, 0]])
                alpha_mean = float(np.nanmean(np.abs(fit.alpha[:, 0])))
                rows.append(
                    {
                        "standardized_type": std,
                        "system": sys_name,
                        "y_series_source": cols[0],
                        "x_series_sources": ", ".join(cols[1:]),
                        "frequency": "weekly",
                        "sample_period": f"{wide.index.min().date()}..{wide.index.max().date()}",
                        "lags_selected": f"k_ar_diff={k_diff}",
                        "cointegration_rank": rank,
                        "cointegration_vectors": cvec,
                        "adjustment_alpha_abs_mean": alpha_mean,
                        "diagnostics_flags": "VECM estimated",
                        "interpretation_note": "Rank>0 supports long-run co-movement in the system.",
                    }
                )
            except Exception:
                continue
    return pd.DataFrame(rows)


def build_neio_proxy_summary(
    silpo_clean: pd.DataFrame,
    prozorro_clean: pd.DataFrame,
    nardl_summary: pd.DataFrame,
    model_series: pd.DataFrame,
) -> pd.DataFrame:
    stds = sorted(model_series["standardized_type"].dropna().unique().tolist()) if not model_series.empty else []
    rows = []
    for std in stds:
        asym = np.nan
        if not nardl_summary.empty:
            ss = nardl_summary[nardl_summary["standardized_type"] == std]
            if not ss.empty:
                asym = float(np.nanmean(np.abs(pd.to_numeric(ss["long_run_coef"], errors="coerce"))))

        promo = np.nan
        if not silpo_clean.empty:
            sil = silpo_clean[silpo_clean["standardized_type"] == std].copy()
            if not sil.empty:
                promo = float(pd.to_numeric(sil["discount_present"], errors="coerce").fillna(0).mean())

        disp = np.nan
        if not prozorro_clean.empty:
            pz = prozorro_clean[prozorro_clean["standardized_type"] == std].copy()
            if not pz.empty:
                tmp = pz.groupby(["date", "region"], as_index=False)["price"].median()
                d = tmp.groupby("date")["price"].agg(["mean", "std"]).dropna()
                disp = float((d["std"] / d["mean"].replace(0, np.nan)).mean()) if not d.empty else np.nan

        # Benchmark link with promo intensity (weekly)
        coef = np.nan
        pval = np.nan
        try:
            if not silpo_clean.empty and not model_series.empty:
                sil = silpo_clean[silpo_clean["standardized_type"] == std].copy()
                if not sil.empty:
                    sil["week"] = pd.to_datetime(sil["date"]).dt.to_period("W-MON").dt.start_time
                    promo_w = sil.groupby("week", as_index=False)["discount_present"].mean()
                    bm = model_series[(model_series["standardized_type"] == std) & (model_series["source"].isin(["EU", "CME"]))].groupby("week", as_index=False)["price"].mean()
                    m = promo_w.merge(bm, on="week", how="inner").sort_values("week")
                    m["d_bm"] = np.log(m["price"]).diff()
                    m = m.dropna()
                    if len(m) > 15:
                        fit = sm.OLS(m["discount_present"], sm.add_constant(m[["d_bm"]])).fit(cov_type="HC1")
                        coef = float(fit.params.get("d_bm", np.nan))
                        pval = float(fit.pvalues.get("d_bm", np.nan))
        except Exception:
            pass

        rows.append(
            {
                "standardized_type": std,
                "asymmetry_strength_index": asym,
                "promo_intensity_index": promo,
                "dispersion_index_prozorro_region": disp,
                "benchmark_shock_link_coef": coef,
                "benchmark_shock_link_pvalue": pval,
                "interpretation_note": "Proxy evidence only; not a structural conduct estimate.",
            }
        )
    return pd.DataFrame(rows)


def write_tests_sections(ws, tests_df: pd.DataFrame, start_row: int = 3) -> None:
    ws.delete_rows(1, ws.max_row)
    row = start_row
    header = [
        "source",
        "product",
        "standardized_type",
        "series_variant",
        "n_obs",
        "adf_p",
        "kpss_p",
        "ljungbox_p",
        "bp_p",
        "white_p",
        "jb_p",
        "stability_flag",
        "stability_drift",
        "integration_class",
        "ac_risk",
        "het_risk",
        "non_normal_risk",
        "stability_risk_class",
        "recommended_action",
        "recommended_model_family",
        "avoid",
        "diag_note",
        "action_label",
    ]

    for src in DATASET_ORDER:
        block = tests_df[tests_df["source"] == src].copy() if not tests_df.empty else pd.DataFrame()

        ws.cell(row=row, column=1, value=f"SECTION: {src}")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for c, h in enumerate(header, start=1):
            ws.cell(row=row, column=c, value=h)
            ws.cell(row=row, column=c).font = Font(bold=True)
        row += 1

        if block.empty:
            ws.cell(row=row, column=1, value="No data")
            row += 2
            continue

        block = block[header]
        for _, r in block.iterrows():
            for c, h in enumerate(header, start=1):
                ws.cell(row=row, column=c, value=r[h])
            row += 1
        row += 1


def write_table_with_note(writer, sheet_name: str, df: pd.DataFrame, note: str) -> None:
    safe_df = df if df is not None else pd.DataFrame()
    safe_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
    ws = writer.book[sheet_name]
    ws["A1"] = note
    ws["A1"].font = Font(bold=True, color="000000")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, min(14, safe_df.shape[1] + 2)))
    ws.row_dimensions[1].height = 36


def style_readme_sheet(ws) -> None:
    ws["A1"] = "RW2 Full Pipeline README"
    ws["A1"].font = Font(bold=True, size=14)

    rule_text = (
        "INTERPRETATION RULES:\n"
        "ADF p>0.05 & KPSS p<0.05 -> series likely non-stationary -> use differences or test cointegration.\n"
        "Cointegration supported -> ECM/VECM admissible.\n"
        "Ljung-Box p<0.05 -> autocorrelation -> add lags / change spec.\n"
        "BP/White p<0.05 -> heteroskedasticity -> robust SE / respecify.\n"
        "Stability flag -> rolling / breaks / split sample.\n"
        "For Silpo discounts, separate baseline price and promo processes.\n"
        "These patterns are expected for price series; they indicate correct specification needs, not 'bad data'."
    )
    ws["B2"] = rule_text
    ws["B2"].font = Font(color="FFFFFF", bold=True)
    ws["B2"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["B"].width = 95
    ws.row_dimensions[2].height = 130


def _pdf_preview(df: pd.DataFrame, max_rows: int = 32, max_cols: int = 10) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame({"info": ["No data rows in this sheet."]})
    cols = list(df.columns[:max_cols])
    out = df[cols].head(max_rows).copy()
    out = out.replace({np.nan: ""})
    if len(df) > max_rows:
        marker = {c: "" for c in out.columns}
        marker[out.columns[0]] = f"... truncated: {len(df) - max_rows} more rows in XLSX"
        out = pd.concat([out, pd.DataFrame([marker])], ignore_index=True)
    return out


def save_pdf_combo(
    xlsx_path: Path,
    pdf_path: Path,
    sheet_notes: Dict[str, str],
    sheet_order: Optional[List[str]] = None,
) -> Optional[Path]:
    import os
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/mplconfig")
    os.environ.setdefault("XDG_CACHE_HOME", "/tmp")
    Path(os.environ["MPLCONFIGDIR"]).mkdir(parents=True, exist_ok=True)
    try:
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
    except Exception as e:
        print(f"[PDF] Skipped: matplotlib not available ({e})")
        return None

    xls = pd.ExcelFile(xlsx_path)
    ordered_sheets = sheet_order if sheet_order is not None else xls.sheet_names
    ordered_sheets = [s for s in ordered_sheets if s in xls.sheet_names]
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    with PdfPages(pdf_path) as pdf:
        # Cover page
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        ax.text(0.02, 0.90, "RW2 Combined Results PDF", fontsize=20, weight="bold", transform=ax.transAxes)
        ax.text(0.02, 0.84, f"Source workbook: {xlsx_path.name}", fontsize=11, transform=ax.transAxes)
        ax.text(0.02, 0.80, f"Sheets included: {len(ordered_sheets)}", fontsize=11, transform=ax.transAxes)
        ax.text(0.02, 0.74, "This PDF is a compact preview. Full tables remain in XLSX.", fontsize=11, transform=ax.transAxes)
        ax.text(0.02, 0.68, "Each page shows top rows/columns for quick review.", fontsize=11, transform=ax.transAxes)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        for sheet in ordered_sheets:
            # Most sheets use header at row 3 because row 1 is note and row 3 is table header.
            try:
                df = xls.parse(sheet, header=2)
            except Exception:
                try:
                    df = xls.parse(sheet)
                except Exception:
                    df = pd.DataFrame()
            prev = _pdf_preview(df, max_rows=30, max_cols=10)

            fig, ax = plt.subplots(figsize=(16, 9))
            ax.axis("off")
            note = sheet_notes.get(sheet, "")
            ax.text(0.01, 1.05, f"{sheet}", fontsize=14, weight="bold", transform=ax.transAxes)
            if note:
                ax.text(0.01, 1.00, note, fontsize=9, transform=ax.transAxes, wrap=True)
            # Fast text rendering is much lighter than matplotlib tables for large multi-page reports.
            txt = prev.astype(str).to_string(index=False, max_colwidth=28)
            ax.text(
                0.01,
                0.93,
                txt,
                transform=ax.transAxes,
                fontsize=7.5,
                family="monospace",
                va="top",
            )
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)
    return pdf_path


def save_full_report_pdf(
    xlsx_path: Path,
    pdf_path: Path,
    sheet_notes: Dict[str, str],
) -> Optional[Path]:
    import os

    os.environ.setdefault("MPLCONFIGDIR", "/tmp/mplconfig")
    os.environ.setdefault("XDG_CACHE_HOME", "/tmp")
    Path(os.environ["MPLCONFIGDIR"]).mkdir(parents=True, exist_ok=True)
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
    except Exception as e:
        print(f"[PDF] Skipped full report: matplotlib unavailable ({e})")
        return None

    if not xlsx_path.exists():
        return None
    xls = pd.ExcelFile(xlsx_path)
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    with PdfPages(pdf_path) as pdf:
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        ax.text(0.02, 0.92, "RW2 Full Modeling + Diagnostics + Decomposition Report", fontsize=18, weight="bold", transform=ax.transAxes)
        ax.text(0.02, 0.86, f"Workbook: {xlsx_path.name}", fontsize=11, transform=ax.transAxes)
        ax.text(0.02, 0.82, f"Sheets: {len(xls.sheet_names)} (one page per sheet)", fontsize=11, transform=ax.transAxes)
        ax.text(0.02, 0.76, "Each page: short interpretation + table preview + chart from available numeric/time fields.", fontsize=10, transform=ax.transAxes)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        for sheet in xls.sheet_names:
            try:
                df = xls.parse(sheet, header=2)
            except Exception:
                try:
                    df = xls.parse(sheet)
                except Exception:
                    df = pd.DataFrame()

            fig = plt.figure(figsize=(11.69, 8.27))
            gs = fig.add_gridspec(2, 2, height_ratios=[0.22, 0.78], width_ratios=[1.15, 0.85])
            ax_head = fig.add_subplot(gs[0, :])
            ax_tbl = fig.add_subplot(gs[1, 0])
            ax_plot = fig.add_subplot(gs[1, 1])

            ax_head.axis("off")
            ax_tbl.axis("off")
            note = sheet_notes.get(sheet, "Module output table.")
            ax_head.text(0.0, 0.86, sheet, fontsize=14, weight="bold", transform=ax_head.transAxes)
            ax_head.text(0.0, 0.20, note[:450], fontsize=9, transform=ax_head.transAxes, wrap=True)

            prev = _pdf_preview(df, max_rows=20, max_cols=9)
            txt = prev.astype(str).to_string(index=False, max_colwidth=24)
            ax_tbl.text(0.0, 1.0, txt, va="top", family="monospace", fontsize=7)

            # Generic chart selection logic.
            plotted = False
            if df is not None and not df.empty:
                d = df.copy()
                date_col = _find_date_col(d)
                num_idx = [i for i in range(d.shape[1]) if pd.api.types.is_numeric_dtype(d.iloc[:, i])]
                if date_col is not None and num_idx:
                    date_idx = next((i for i, c in enumerate(d.columns) if c == date_col), None)
                    if date_idx is not None:
                        dd = pd.DataFrame({"__date__": d.iloc[:, date_idx]})
                        plot_cols: List[Tuple[str, str]] = []
                        for i in num_idx[:2]:
                            cname = str(d.columns[i])
                            key = f"__y{i}__"
                            dd[key] = pd.to_numeric(d.iloc[:, i], errors="coerce")
                            plot_cols.append((key, cname))
                        dd["__date__"] = pd.to_datetime(dd["__date__"], errors="coerce")
                        dd = dd.dropna(subset=["__date__"]).sort_values("__date__")
                    if not dd.empty:
                        for key, lbl in plot_cols:
                            yy = pd.to_numeric(dd[key], errors="coerce")
                            if yy.notna().sum() > 2:
                                ax_plot.plot(dd["__date__"], yy, label=lbl, linewidth=1.2)
                                plotted = True
                        if plotted:
                            ax_plot.set_title("Time-Series View")
                            ax_plot.legend(fontsize=7)
                            ax_plot.tick_params(axis="x", rotation=35, labelsize=7)
                            ax_plot.tick_params(axis="y", labelsize=7)
                if not plotted and num_idx:
                    s = pd.to_numeric(d.iloc[:, num_idx[0]], errors="coerce").dropna()
                    if len(s) > 2:
                        ax_plot.hist(s.values, bins=20, color="#4C78A8", alpha=0.85)
                        ax_plot.set_title(f"Distribution: {str(d.columns[num_idx[0]])}")
                        plotted = True
            if not plotted:
                ax_plot.axis("off")
                ax_plot.text(0.05, 0.5, "No plottable numeric/time data", fontsize=9)

            fig.tight_layout()
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)
    return pdf_path


def terminal_dataset_summary(
    name: str,
    df: pd.DataFrame,
    daily_df: Optional[pd.DataFrame] = None,
    date_col: str = "date",
    prod_col: str = "product",
) -> str:
    if df.empty:
        return f"- {name}: rows=0"
    d = df.copy()
    d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
    rows = len(d)
    dmin = d[date_col].min()
    dmax = d[date_col].max()
    miss = float(d["price"].isna().mean()) if "price" in d.columns else np.nan
    imp = np.nan
    if daily_df is not None and not daily_df.empty:
        imp = float(pd.to_numeric(daily_df.get("imputed_flag_pchip"), errors="coerce").mean())
    top = ", ".join(d[prod_col].astype(str).value_counts().head(3).index.tolist()) if prod_col in d.columns else ""
    return f"- {name}: rows={rows}, date=[{dmin.date()}..{dmax.date()}], missing={miss:.3f}, imputed={imp:.3f}, top_products={top}"


def print_terminal_report(
    datasets_clean: Dict[str, pd.DataFrame],
    all_daily: pd.DataFrame,
    corr_df: pd.DataFrame,
    tests_df: pd.DataFrame,
    eligibility: pd.DataFrame,
    output_file: Path,
) -> None:
    print("\n=== RW2 PIPELINE REPORT ===")
    print(f"Output: {output_file}")
    print("\n[Datasets]")
    for name in ["ProducerUA", "ConsumerUA", "EU", "ProZorro", "Silpo", "Novus", "CME"]:
        df = datasets_clean.get(name, pd.DataFrame())
        dd = all_daily[all_daily["source"] == name] if not all_daily.empty else pd.DataFrame()
        print(terminal_dataset_summary(name, df, daily_df=dd))

    print("\n[Top-5 correlations between sources]")
    if corr_df.empty:
        print("- no eligible pairs")
    else:
        top5 = corr_df[corr_df["corr_type"] == "between_sources"].copy()
        top5["abs_pearson"] = top5["pearson"].abs()
        top5 = top5.sort_values("abs_pearson", ascending=False).head(5)
        for _, r in top5.iterrows():
            print(
                f"- {r['product']}: {r['source_left']} vs {r['source_right']} lag={int(r['lag'])}d "
                f"pearson={r['pearson']:.3f} spearman={r['spearman']:.3f}"
            )

    print("\n[Thesis eligibility summary]")
    if eligibility.empty:
        print("- no products with sufficient data")
    else:
        n = len(eligibility)
        ecm = int((eligibility["ecm_vecm_feasible"] == 1).sum())
        robust = int((eligibility["need_robust_se"] == 1).sum())
        unstable = int((eligibility["stability_risk"] == 1).sum())
        promo_high = int((eligibility["promotions_confounding"] == "high").sum())
        print(f"- products_total={n}")
        print(f"- ecm_vecm_feasible={ecm}")
        print(f"- need_robust_se={robust}")
        print(f"- stability_risk={unstable}")
        print(f"- promo_confounding_high={promo_high}")

    print("\n[Addressing diagnostics]")
    if tests_df.empty:
        print("- no diagnostics rows")
    else:
        patt = tests_df[
            (tests_df["integration_class"] == "I(1)-like")
            & (tests_df["ac_risk"] == "high")
            & (tests_df["het_risk"] == "high")
            & (tests_df["non_normal_risk"] == "high")
            & (tests_df["stability_risk_class"] == "high")
        ]
        print(f"- pattern_rows={len(patt)}")
        if len(patt) > 0:
            rec = patt["recommended_model_family"].mode().iloc[0]
            avoid = patt["avoid"].mode().iloc[0]
            note = patt["diag_note"].mode().iloc[0]
            print(f"- recommended_model_family={rec}")
            print(f"- avoid={avoid}")
            print(f"- note={note}")


def run_pipeline(cfg: Config) -> None:
    cfg.output_file.parent.mkdir(parents=True, exist_ok=True)

    # STEP 0: discovery + fail-fast checks
    workbook = discover_input_workbook(cfg.model_dir, cfg.input_file)
    discovered = discover_input_files(cfg.model_dir)

    xls = pd.ExcelFile(workbook)
    required_sheets = ["CME III", "Consumer_UA", "Producer_UA", "Prozorro", "Europe", "Silpo", "Novus"]
    missing_sheets = [s for s in required_sheets if s not in xls.sheet_names]
    if missing_sheets:
        raise ValueError(f"Missing critical sheets in {workbook}: {missing_sheets}")

    raw_producer = read_sheet_or_empty(xls, "Producer_UA")
    raw_consumer = read_sheet_or_empty(xls, "Consumer_UA")
    raw_prozorro = read_sheet_or_empty(xls, "Prozorro")
    raw_silpo = read_sheet_or_empty(xls, "Silpo")
    raw_novus = read_sheet_or_empty(xls, "Novus")
    raw_eu = read_sheet_or_empty(xls, "Europe")
    raw_cme = read_sheet_or_empty(xls, "CME III")

    # STEP 1 + 2: classification and unit logic
    producer = prep_producer_consumer(raw_producer, "ProducerUA")
    consumer = prep_producer_consumer(raw_consumer, "ConsumerUA")
    prozorro = prep_prozorro(raw_prozorro)
    silpo = prep_retail(raw_silpo, "Silpo")
    novus = prep_retail(raw_novus, "Novus")
    eu = prep_eu(raw_eu)
    cme = prep_cme(raw_cme)

    # STEP 3: standardization gate in UAH-native scale (no FX conversion)
    all_dates = pd.concat(
        [
            producer.get("date", pd.Series(dtype="datetime64[ns]")),
            consumer.get("date", pd.Series(dtype="datetime64[ns]")),
            prozorro.get("date", pd.Series(dtype="datetime64[ns]")),
            silpo.get("date", pd.Series(dtype="datetime64[ns]")),
            novus.get("date", pd.Series(dtype="datetime64[ns]")),
            eu.get("date", pd.Series(dtype="datetime64[ns]")),
            cme.get("date", pd.Series(dtype="datetime64[ns]")),
        ],
        ignore_index=True,
    )
    producer = ensure_numeric_price(producer)
    consumer = ensure_numeric_price(consumer)
    prozorro = ensure_numeric_price(prozorro)
    silpo = ensure_numeric_price(silpo)
    novus = ensure_numeric_price(novus)
    eu = ensure_numeric_price(eu)
    cme = ensure_numeric_price(cme)

    producer_pre_scale = producer.copy()
    consumer_pre_scale = consumer.copy()
    prozorro_pre_scale = prozorro.copy()
    silpo_pre_scale = silpo.copy()
    novus_pre_scale = novus.copy()
    eu_pre_scale = eu.copy()
    cme_pre_scale = cme.copy()

    # Explicit derived variables in UAH-native scale.
    if not prozorro.empty:
        prozorro["prozorro_unit_price_uah"] = pd.to_numeric(prozorro["price"], errors="coerce")
        prozorro["savings_rate"] = (
            pd.to_numeric(prozorro.get("expected"), errors="coerce")
            - pd.to_numeric(prozorro.get("sum_current"), errors="coerce")
        ) / pd.to_numeric(prozorro.get("expected"), errors="coerce").replace(0, np.nan)
        lvl = prozorro.groupby(["region", "product"], as_index=False).agg(
            region_price_level_median=("price", "median"),
            region_price_level_mean=("price", "mean"),
            region_cv=("price", lambda s: float(pd.to_numeric(s, errors="coerce").std() / pd.to_numeric(s, errors="coerce").mean()) if pd.to_numeric(s, errors="coerce").mean() not in [0, np.nan] else np.nan),
            region_iqr=("price", lambda s: float(pd.to_numeric(s, errors="coerce").quantile(0.75) - pd.to_numeric(s, errors="coerce").quantile(0.25))),
        )
        prozorro = prozorro.merge(lvl, on=["region", "product"], how="left")

    if not silpo.empty:
        silpo["baseline_price"] = np.where(
            pd.to_numeric(silpo.get("price_old"), errors="coerce").notna()
            & (pd.to_numeric(silpo.get("price_old"), errors="coerce") > 0)
            & (pd.to_numeric(silpo.get("price_current"), errors="coerce") < pd.to_numeric(silpo.get("price_old"), errors="coerce")),
            pd.to_numeric(silpo.get("price_old"), errors="coerce"),
            pd.to_numeric(silpo.get("price_current"), errors="coerce"),
        )

    datasets_clean = {
        "ProducerUA": producer,
        "ConsumerUA": consumer,
        "EU": eu,
        "ProZorro": prozorro,
        "Silpo": silpo,
        "Novus": novus,
        "CME": cme,
    }
    raw_by_sheet = {
        "Producer_UA": raw_producer,
        "Consumer_UA": raw_consumer,
        "Europe": raw_eu,
        "Prozorro": raw_prozorro,
        "Silpo": raw_silpo,
        "Novus": raw_novus,
        "CME III": raw_cme,
    }

    step_frames: List[Dict[str, object]] = [
        {
            "dataset_name": "ProducerUA",
            "sheet_name": "Producer_UA",
            "step_id": "S1_parse_classify",
            "step_description": "Date parsing, mapping to product/standardized_type, base clean.",
            "before_df": raw_producer,
            "after_df": producer_pre_scale,
        },
        {
            "dataset_name": "ProducerUA",
            "sheet_name": "Producer_UA",
            "step_id": "S2_uah_alignment",
            "step_description": "UAH-native alignment by date and numeric coercion.",
            "before_df": producer_pre_scale,
            "after_df": producer,
        },
        {
            "dataset_name": "ConsumerUA",
            "sheet_name": "Consumer_UA",
            "step_id": "S1_parse_classify",
            "step_description": "Date parsing, mapping to product/standardized_type, base clean.",
            "before_df": raw_consumer,
            "after_df": consumer_pre_scale,
        },
        {
            "dataset_name": "ConsumerUA",
            "sheet_name": "Consumer_UA",
            "step_id": "S2_uah_alignment",
            "step_description": "UAH-native alignment by date and numeric coercion.",
            "before_df": consumer_pre_scale,
            "after_df": consumer,
        },
        {
            "dataset_name": "ProZorro",
            "sheet_name": "Prozorro",
            "step_id": "S1_parse_classify_units",
            "step_description": "Date parsing, NLP categorization, unit extraction and unit price harmonization.",
            "before_df": raw_prozorro,
            "after_df": prozorro_pre_scale,
        },
        {
            "dataset_name": "ProZorro",
            "sheet_name": "Prozorro",
            "step_id": "S2_uah_metrics",
            "step_description": "UAH procurement metrics derivation (unit/current/initial/expected).",
            "before_df": prozorro_pre_scale,
            "after_df": prozorro,
        },
        {
            "dataset_name": "Silpo",
            "sheet_name": "Silpo",
            "step_id": "S1_parse_classify_discounts",
            "step_description": "Date parsing, categorization, baseline and discount regime fields.",
            "before_df": raw_silpo,
            "after_df": silpo_pre_scale,
        },
        {
            "dataset_name": "Silpo",
            "sheet_name": "Silpo",
            "step_id": "S2_uah_alignment",
            "step_description": "UAH-native alignment and baseline price update.",
            "before_df": silpo_pre_scale,
            "after_df": silpo,
        },
        {
            "dataset_name": "Novus",
            "sheet_name": "Novus",
            "step_id": "S1_parse_classify_dedupe",
            "step_description": "Date parsing, NLP categorization and latest-timestamp deduplication per (date,id).",
            "before_df": raw_novus,
            "after_df": novus_pre_scale,
        },
        {
            "dataset_name": "Novus",
            "sheet_name": "Novus",
            "step_id": "S2_uah_alignment",
            "step_description": "UAH-native alignment and unit-price extraction.",
            "before_df": novus_pre_scale,
            "after_df": novus,
        },
        {
            "dataset_name": "EU",
            "sheet_name": "Europe",
            "step_id": "S1_parse_classify",
            "step_description": "Date parsing, product mapping, using Price (UAH/kg).",
            "before_df": raw_eu,
            "after_df": eu_pre_scale,
        },
        {
            "dataset_name": "EU",
            "sheet_name": "Europe",
            "step_id": "S2_uah_alignment",
            "step_description": "UAH-native alignment by date.",
            "before_df": eu_pre_scale,
            "after_df": eu,
        },
        {
            "dataset_name": "CME",
            "sheet_name": "CME III",
            "step_id": "S1_parse",
            "step_description": "Date parsing and base CME series clean.",
            "before_df": raw_cme,
            "after_df": cme_pre_scale,
        },
        {
            "dataset_name": "CME",
            "sheet_name": "CME III",
            "step_id": "S2_uah_alignment",
            "step_description": "UAH-native alignment by date.",
            "before_df": cme_pre_scale,
            "after_df": cme,
        },
    ]

    # STEP 4: daily real/linear/pchip by product x source
    daily_parts = []
    for src in ["ProducerUA", "ConsumerUA", "EU", "ProZorro", "Silpo", "Novus", "CME"]:
        dd = datasets_clean[src]
        dv_input = dd.copy()
        dv = daily_variants_for_dataset(dv_input, source=src)
        if not dv.empty:
            daily_parts.append(dv)
    all_daily = pd.concat(daily_parts, ignore_index=True) if daily_parts else pd.DataFrame()
    all_daily, outliers_summary = winsorize_daily_variants(all_daily, lower_q=0.01, upper_q=0.99)
    series_catalog = build_series_catalog(long_price_series(all_daily))
    log_variants = build_log_variant_table(all_daily)
    transformation_ledger, transformation_stats, transformation_tests = build_transformation_outputs(
        step_frames
        + [
            {
                "dataset_name": "ALL",
                "sheet_name": "all_daily",
                "step_id": "S3_interpolation_variants",
                "step_description": "Built real/linear/pchip daily variants per source-product.",
                "before_df": pd.DataFrame(),
                "after_df": all_daily.rename(columns={"price_real": "price"}),
            },
            {
                "dataset_name": "ALL",
                "sheet_name": "all_daily",
                "step_id": "S4_outlier_policy",
                "step_description": "Applied winsorization policy with outlier flags.",
                "before_df": all_daily.rename(columns={"price_real": "price"}),
                "after_df": all_daily.rename(columns={"price_winsor_real": "price"}),
            },
            {
                "dataset_name": "ALL",
                "sheet_name": "log_variants",
                "step_id": "S5_log_transform",
                "step_description": "Generated log and dlog variants for modeling admissibility.",
                "before_df": all_daily.rename(columns={"price_real": "price"}),
                "after_df": log_variants.rename(columns={"price": "price"}),
            },
        ],
        min_obs=cfg.min_obs_tests,
    )

    # STEP 5: descriptive stats (product-first)
    desc_producer = build_descriptive_stats(all_daily[all_daily["source"] == "ProducerUA"], "ProducerUA") if not all_daily.empty else pd.DataFrame()
    desc_consumer = build_descriptive_stats(all_daily[all_daily["source"] == "ConsumerUA"], "ConsumerUA") if not all_daily.empty else pd.DataFrame()
    desc_eu = build_descriptive_stats(all_daily[all_daily["source"] == "EU"], "EU") if not all_daily.empty else pd.DataFrame()
    desc_cme = build_descriptive_stats(all_daily[all_daily["source"] == "CME"], "CME") if not all_daily.empty else pd.DataFrame()
    desc_prozorro = build_descriptive_stats(all_daily[all_daily["source"] == "ProZorro"], "ProZorro") if not all_daily.empty else pd.DataFrame()
    desc_novus = build_descriptive_stats(all_daily[all_daily["source"] == "Novus"], "Novus") if not all_daily.empty else pd.DataFrame()
    desc_silpo = build_descriptive_stats(all_daily[all_daily["source"] == "Silpo"], "Silpo") if not all_daily.empty else pd.DataFrame()
    prozorro_metric_stats = multi_metric_product_stats(
        prozorro,
        metrics=["unit_price", "expected", "sum_initial", "sum_current", "price"],
        source_name="ProZorro",
    )
    series_producer = long_price_series(all_daily[all_daily["source"] == "ProducerUA"].copy()) if not all_daily.empty else pd.DataFrame()
    series_consumer = long_price_series(all_daily[all_daily["source"] == "ConsumerUA"].copy()) if not all_daily.empty else pd.DataFrame()
    series_eu = long_price_series(all_daily[all_daily["source"] == "EU"].copy()) if not all_daily.empty else pd.DataFrame()
    series_cme = long_price_series(all_daily[all_daily["source"] == "CME"].copy()) if not all_daily.empty else pd.DataFrame()

    # Additional slices
    proz_by_region = group_stats(prozorro, "price", ["region", "product", "standardized_type"]) if not prozorro.empty else pd.DataFrame()
    novus_by_brand = group_stats(novus, "price", ["brand", "product", "standardized_type"]) if not novus.empty else pd.DataFrame()
    silpo_by_brand = group_stats(silpo, "price", ["brand", "product", "standardized_type"]) if not silpo.empty else pd.DataFrame()
    standardisation_policy = build_standardisation_policy(datasets_clean)
    brand_io_metrics = build_brand_io_metrics(silpo, novus)
    brand_econ_metrics = build_brand_economic_metrics(silpo, novus, all_daily)
    lag_matrix_all = build_lag_matrix_all(all_daily)
    lag_best, lag_profiles = build_lag_outputs(lag_matrix_all)
    decomp_all, decomp_summary, decomp_sheet_tables, decomp_index = build_decomposition_tables(all_daily)
    before_after_ln_all, before_after_ln_tables, before_after_ln_index = build_before_after_ln_tables(all_daily)
    overlay_all, overlay_tables, overlay_index = build_overlay_tables(all_daily)
    corr_modules = build_correlation_modules(all_daily, prozorro, silpo, novus)

    # STEP 6: correlations
    corr_df = compute_correlations(all_daily)
    corr_matrix = build_corr_matrix(corr_df)

    # STEP 7: tests
    weekly_all_df = prepare_weekly_for_tests(all_daily)
    tests_df = build_tests_table(weekly_all_df, min_obs=cfg.min_obs_tests)
    stationarity_cols = [
        "source",
        "product",
        "standardized_type",
        "series_variant",
        "n_obs",
        "adf_p",
        "kpss_p",
        "integration_class",
        "recommended_action",
    ]
    autocorr_cols = [
        "source",
        "product",
        "standardized_type",
        "series_variant",
        "ljungbox_p",
        "ac_risk",
        "recommended_action",
    ]
    heterosk_cols = [
        "source",
        "product",
        "standardized_type",
        "series_variant",
        "bp_p",
        "white_p",
        "het_risk",
        "recommended_action",
    ]
    normality_cols = [
        "source",
        "product",
        "standardized_type",
        "series_variant",
        "jb_p",
        "non_normal_risk",
        "recommended_action",
    ]
    stability_cols = [
        "source",
        "product",
        "standardized_type",
        "series_variant",
        "drift_metric",
        "stability_flag",
        "stability_risk_class",
        "recommended_action",
        "recommended_model_family",
        "avoid",
        "diag_note",
    ]
    test_stationarity = (
        tests_df[[c for c in stationarity_cols if c in tests_df.columns]].copy() if not tests_df.empty else pd.DataFrame()
    )
    test_autocorr = tests_df[[c for c in autocorr_cols if c in tests_df.columns]].copy() if not tests_df.empty else pd.DataFrame()
    test_heterosk = tests_df[[c for c in heterosk_cols if c in tests_df.columns]].copy() if not tests_df.empty else pd.DataFrame()
    test_normality = (
        tests_df[[c for c in normality_cols if c in tests_df.columns]].copy() if not tests_df.empty else pd.DataFrame()
    )
    test_stability = (
        tests_df[[c for c in stability_cols if c in tests_df.columns]].copy() if not tests_df.empty else pd.DataFrame()
    )

    # Bundles/model plan and model modules
    dataset_registry = build_dataset_registry(raw_by_sheet, datasets_clean)
    product_mapping_table = build_product_mapping_table(datasets_clean)
    # Hard gate for level-model families: unit_ok==1 only.
    weekly_level_df = prepare_weekly_for_tests(all_daily[all_daily["unit_ok"] == 1].copy()) if not all_daily.empty else pd.DataFrame()
    model_series = build_model_series(weekly_level_df)
    model_plan = build_model_plan(product_mapping_table, model_series)
    ardl_summary = build_ardl_summary(model_series)
    ecm_summary = build_ecm_summary(model_series, ardl_summary)
    nardl_summary = build_nardl_summary(model_series)
    vecm_summary = build_vecm_summary(model_series)
    neio_proxy_summary = build_neio_proxy_summary(silpo, prozorro, nardl_summary, model_series)
    msem_summary = build_msem_summary(all_daily)
    prozorro_regional_fe = build_prozorro_regional_fe(prozorro, model_series)
    silpo_occurrence, silpo_depth, silpo_transmission = build_silpo_discount_modules(silpo, novus, all_daily, brand_io_metrics)
    matrices_heatmap = build_matrices_heatmap_ready(model_plan, brand_io_metrics, ardl_summary, neio_proxy_summary, lag_matrix_all)
    short_run_summary, short_run_details = build_short_run_models(all_daily, lag_best)
    chain_summary, chain_details = build_chain_effects(all_daily, lag_best)
    prozorro_regional_models, region_product_matrix = build_prozorro_regional_modules(prozorro, all_daily)
    if short_run_summary.empty:
        short_run_summary = pd.DataFrame(
            [
                {
                    "product": "n/a",
                    "standardized_type": "n/a",
                    "y_source": "n/a",
                    "n_obs": 0,
                    "r2": np.nan,
                    "adj_r2": np.nan,
                    "interpretation_note": "No short-run retail model passed minimum overlap/lag-data requirements.",
                }
            ]
        )
    if prozorro_regional_models.empty:
        prozorro_regional_models = pd.DataFrame(
            [
                {
                    "product": "n/a",
                    "standardized_type": "n/a",
                    "n_obs": 0,
                    "interpretation_note": "No ProZorro regional FE model passed minimum data requirements.",
                }
            ]
        )
    if region_product_matrix.empty:
        region_product_matrix = pd.DataFrame(
            [{"region": "n/a", "product": "n/a", "price_median": np.nan, "note": "No region-product matrix available in current sample."}]
        )
    stargazer_like = build_stargazer_like(
        ardl_summary,
        ecm_summary,
        nardl_summary,
        vecm_summary,
        prozorro_regional_fe,
        silpo_transmission,
    )

    # STEP 8: forecasts
    forecast_prices = build_price_forecasts(all_daily)
    eu_daily = all_daily[all_daily["source"] == "EU"]
    producer_daily = all_daily[all_daily["source"] == "ProducerUA"]
    forecast_discounts = build_silpo_discount_forecasts(silpo, eu_daily, producer_daily)

    # Eligibility
    eu_weekly = weekly_level_df[(weekly_level_df["source"] == "EU") & (weekly_level_df["series_variant"] == "real")]
    eligibility = build_model_eligibility(tests_df, weekly_level_df, silpo, eu_weekly)
    tests_summary = build_tests_summary(tests_df)

    # Detail tables (per product/standardized_type modules).
    ecm_detail_tables: Dict[str, pd.DataFrame] = {}
    vecm_detail_tables: Dict[str, pd.DataFrame] = {}
    nardl_detail_tables: Dict[str, pd.DataFrame] = {}
    retail_trans_tables: Dict[str, pd.DataFrame] = {}
    chain_effect_tables: Dict[str, pd.DataFrame] = {}

    if not ecm_summary.empty:
        for std, g in ecm_summary.groupby("standardized_type", dropna=False):
            ecm_detail_tables[_excel_safe_sheet_name("ECM_Details", str(std))] = g.copy()
    if not vecm_summary.empty:
        for std, g in vecm_summary.groupby("standardized_type", dropna=False):
            vecm_detail_tables[_excel_safe_sheet_name("VECM_Details", str(std))] = g.copy()
    if not nardl_summary.empty:
        for std, g in nardl_summary.groupby("standardized_type", dropna=False):
            nardl_detail_tables[_excel_safe_sheet_name("NARDL_Details", str(std))] = g.copy()
    if not short_run_details.empty:
        for prod, g in short_run_details.groupby("product", dropna=False):
            retail_trans_tables[_excel_safe_sheet_name("Retail_Transmission", str(prod))] = g.copy()
    if not chain_details.empty:
        for prod, g in chain_details.groupby("product", dropna=False):
            chain_effect_tables[_excel_safe_sheet_name("Chain_Effects", str(prod))] = g.copy()
    if not retail_trans_tables:
        retail_trans_tables["Retail_Transmission_NoData"] = pd.DataFrame(
            [{"note": "No short-run retail transmission model met minimum sample/quality conditions in current run."}]
        )

    # Categorization report
    cat_table = categorisation_table(datasets_clean)

    # README base table
    readme_df = pd.DataFrame(
        {
            "item": [
                "input_workbook",
                "discovered_xlsx_files",
                "period_min",
                "period_max",
                "currency_mode",
                "notes",
            ],
            "value": [
                str(workbook),
                "; ".join([p.name for p in discovered]),
                str(pd.to_datetime(all_dates, errors="coerce").min().date()) if not all_dates.empty else "",
                str(pd.to_datetime(all_dates, errors="coerce").max().date()) if not all_dates.empty else "",
                "UAH-native (no FX transformation)",
                "Product-first pipeline completed",
            ],
        }
    )

    sheet_notes = {
        "00_README": "Overview: inputs, period coverage, and interpretation rules for tests/models.",
        "01_Categorisation": "Check classification quality: raw labels -> product -> standardized_type -> brand. Low coverage means mapping should be improved.",
        "02_Prozorro_Clean": "Cleaned ProZorro records with unified `price` and unit validity flag (`unit_ok`).",
        "03_Prozorro_Stats": "Descriptive stats for ProZorro time-series and contract metrics; compare spread/volatility across products.",
        "04_Prozorro_ByRegion": "Regional distribution of ProZorro prices; use for heterogeneity checks.",
        "05_Novus_Clean": "Cleaned Novus records with unified `price`; latest observation retained for same day-item.",
        "06_Novus_Stats": "Descriptive stats for Novus product series.",
        "07_Novus_ByBrand": "Brand-level Novus dispersion and price levels.",
        "08_Silpo_Clean": "Cleaned Silpo records with price/promo fields and unified `price`.",
        "09_Silpo_Stats": "Descriptive stats for Silpo product series and promo-sensitive variability.",
        "10_Silpo_ByBrand": "Brand-level Silpo price behavior summary.",
        "11_Series_ProducerUA": "Daily ProducerUA time-series by variant. `real` observed; `linear`/`pchip` are interpolated variants.",
        "12_Series_ConsumerUA": "Daily ConsumerUA time-series by variant. `real` observed; `linear`/`pchip` are interpolated variants.",
        "13_Series_EU": "Daily EU series as observed (`real`) in unified `price` scale.",
        "14_Series_CME": "Daily CME series as observed (`real`) in unified `price` scale.",
        "15_Correlations": "Pearson/Spearman and lagged correlations. High absolute values indicate stronger co-movement, not causality.",
        "Corr_Matrix": "Source-level Pearson correlation matrix (lag=0), averaged over available product pairs.",
        "Stargazer_Like": "Consolidated model-results table (stargazer-like) across ARDL/ECM/NARDL/VECM and key robustness modules.",
        "16_Tests": "Interpretation: ADF p>0.05 & KPSS p<0.05 -> non-stationary; Ljung-Box p<0.05 -> autocorrelation; BP/White p<0.05 -> robust SE; stability_flag=1 -> potential breaks.",
        "Test_Stationarity": "Stationarity diagnostics: ADF + KPSS with integration class and action hints.",
        "Test_Autocorr": "Autocorrelation diagnostics: Ljung-Box p-values and lag-structure recommendations.",
        "Test_Heterosk": "Heteroskedasticity diagnostics: Breusch-Pagan/White with robust-inference recommendations.",
        "Test_Normality": "Normality diagnostics: Jarque-Bera and implications for robust inference/bootstrapping.",
        "Test_Stability": "Stability/drift diagnostics and split-sample or rolling-estimation recommendations.",
        "17_Forecasts_Prices": "Daily forecast paths for 7-day and 30-day horizons by product and standardized_type.",
        "18_Forecasts_Discounts": "Daily 7/30-day discount forecasts plus determinants (coef/pvalue) for present/depth models.",
        "19_Model_Eligibility": "Product-level modeling admissibility: cointegration/ECM-VECM feasibility, robust SE need, promo confounding, stability risk.",
        "20_Model_Plan": "Model bundles by standardized_type and admissible model families based on available series.",
        "21_Dataset_Registry": "Registry of source sheets and mapped operational columns used by the runner.",
        "22_Product_Mapping": "Mapping matrix snapshot: source raw labels mapped to product and standardized_type.",
        "23_ARDL_Summary": "ARDL module summary: selected lags, short/long-run coefficients, diagnostics, and ECM eligibility.",
        "24_ECM_Summary": "ECM module summary: long-run relation and error-correction adjustment speed.",
        "25_NARDL_Summary": "NARDL asymmetry module: short-run and long-run asymmetry tests.",
        "26_VECM_Summary": "VECM systems by standardized_type: rank, vectors, adjustment coefficients.",
        "27_NEIO_Proxy_Summary": "NEIO-style proxy indices: asymmetry, promo intensity, procurement dispersion, benchmark-shock linkage.",
        "NEIO_Proxy_Summary": "NEIO-style proxy indices: asymmetry, promo intensity, procurement dispersion, benchmark-shock linkage.",
        "Standardisation_UnitPolicy": "Hard gate: level/cointegration models should use only unit_ok==1. Currency is UAH-native across all datasets.",
        "Outliers_Winsorisation": "Outlier policy and trimming shares by product×source×variant (winsor 1%/99%).",
        "Brand_IO_Metrics": "IO structure in retail: HHI, top3 share, private label share, SKU count.",
        "Brand_Economic_Metrics": "Brand premium, promo intensity, volatility, and brand-level pass-through proxies.",
        "Series_Catalog": "Catalog of all modeled time series with variant, period, missingness, and block assignment.",
        "Correlations_All": "All pairwise correlations and lagged relationships across market layers.",
        "LagMatrix_All": "Lag scan (1..30 days) for layer pairs; best lag markers included.",
        "Tests_All": "All diagnostics with interpretation and recommended actions.",
        "ARDL_Summary": "ARDL results for admissible bundles (short-run and implied long-run effects).",
        "NARDL_Summary": "Asymmetric transmission (positive vs negative shocks) for admissible bundles.",
        "ECM_Summary": "Error-correction results where long-run co-movement is supported.",
        "VECM_Summary": "System-level cointegration dynamics (Johansen/VECM).",
        "MSEM_Summary": "MSEM-like latent factor summary for the short retail intersection window.",
        "Silpo_Discounts_Occurrence": "Discount occurrence models with EU/Producer/competitor/seasonality/brand IO controls.",
        "Silpo_Discounts_Depth": "Discount depth models conditional on discount occurrence.",
        "Silpo_Transmission_PromoCtrl": "Pass-through comparison without vs with promo controls in Silpo baseline dynamics.",
        "Prozorro_Regional_FE": "Regional FE model for ProZorro transmission and procurement heterogeneity.",
        "Model_Eligibility_Summary": "Admissible model family by product with diagnostics-driven restrictions.",
        "Matrices_HeatmapReady": "Heatmap-ready long matrices for coverage, lags, IO structure, and transmission strength.",
        "Transformation_Ledger": "Transformation ledger with row/missing/unit/outlier changes per step and interpretation notes.",
        "Transformation_Stats": "Descriptive statistics after each transformation step.",
        "Transformation_Tests": "Diagnostics after each transformation step with model-choice implications.",
        "Log_Variants_All": "Long table of price, log_price, and dlog_price across real/linear/pchip variants.",
        "Decomposition_All": "Trend/seasonal/residual decomposition components for each source-product series.",
        "Decomposition_Summary": "Variance shares and seasonal strength from decomposition.",
        "Decomposition_Index": "Map of decomposition module sheets by source and product.",
        "BeforeAfterLN_All": "Before-ln and after-ln series with rolling trends.",
        "BeforeAfterLN_Index": "Map of per-series before/after-ln sheets.",
        "Overlay_All": "Intersection overlay data across market layers by product.",
        "Overlay_Index": "Map of overlay chart sheets by product and window.",
        "LagMatrix_ByProduct": "Best lag per product and source pair from 1..30-day lag scan.",
        "LagProfiles_ByProduct": "Full lag profiles (1..30 days) for each product/source pair.",
        "Corr_Regions_Prozorro": "Region-level correlations for ProZorro series (log and dlog).",
        "Corr_Brands_Retail": "Brand-level correlations for retail series (log and dlog).",
        "Corr_Sources_ByProduct": "Cross-source correlations by product for log and dlog series.",
        "Tests_All_Full": "Complete diagnostics table across all series and variants.",
        "Tests_Summary": "Condensed diagnostics implication table per source/product.",
        "Models_LongRun_Summary": "Long-run transmission module summary (ARDL/ECM/NARDL/VECM).",
        "Models_ShortRun_Summary": "Short-run retail transmission module summary on intersection windows.",
        "Chain_Effects_Summary": "Retail↔ProZorro↔Producer pass-through comparison with benchmark exposures.",
        "Prozorro_Regional_Models": "Region-aware transmission models for ProZorro with benchmark and producer shocks.",
        "Prozorro_Regional_Effects_Summary": "Region×product price level and dispersion matrix.",
    }
    for src in DATASET_ORDER:
        sheet_notes[f"Corr_Products_{src}"] = f"Within-{src} product correlation matrix in log and dlog space."

    # Write all to one XLSX (write to /tmp first to avoid FS timeout, then copy to final path)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir="/tmp") as tf:
        tmp_xlsx = Path(tf.name)

    with pd.ExcelWriter(tmp_xlsx, engine="openpyxl") as writer:
        write_table_with_note(writer, "00_README", readme_df, sheet_notes["00_README"])
        write_table_with_note(writer, "01_Categorisation", cat_table, sheet_notes["01_Categorisation"])
        write_table_with_note(writer, "Transformation_Ledger", transformation_ledger, sheet_notes["Transformation_Ledger"])
        write_table_with_note(writer, "Transformation_Stats", transformation_stats, sheet_notes["Transformation_Stats"])
        write_table_with_note(writer, "Transformation_Tests", transformation_tests, sheet_notes["Transformation_Tests"])
        write_table_with_note(writer, "Log_Variants_All", log_variants, sheet_notes["Log_Variants_All"])
        write_table_with_note(writer, "Decomposition_All", decomp_all, sheet_notes["Decomposition_All"])
        write_table_with_note(writer, "Decomposition_Summary", decomp_summary, sheet_notes["Decomposition_Summary"])
        write_table_with_note(writer, "Decomposition_Index", decomp_index, sheet_notes["Decomposition_Index"])
        write_table_with_note(writer, "BeforeAfterLN_All", before_after_ln_all, sheet_notes["BeforeAfterLN_All"])
        write_table_with_note(writer, "BeforeAfterLN_Index", before_after_ln_index, sheet_notes["BeforeAfterLN_Index"])
        write_table_with_note(writer, "Overlay_All", overlay_all, sheet_notes["Overlay_All"])
        write_table_with_note(writer, "Overlay_Index", overlay_index, sheet_notes["Overlay_Index"])

        write_table_with_note(writer, "02_Prozorro_Clean", prozorro, sheet_notes["02_Prozorro_Clean"])
        proz_stats = pd.concat(
            [
                desc_prozorro.assign(block="series_real_linear_pchip"),
                prozorro_metric_stats.assign(block="contract_metrics"),
            ],
            ignore_index=True,
            sort=False,
        )
        write_table_with_note(writer, "03_Prozorro_Stats", proz_stats, sheet_notes["03_Prozorro_Stats"])
        write_table_with_note(writer, "04_Prozorro_ByRegion", proz_by_region, sheet_notes["04_Prozorro_ByRegion"])

        write_table_with_note(writer, "05_Novus_Clean", novus, sheet_notes["05_Novus_Clean"])
        write_table_with_note(writer, "06_Novus_Stats", desc_novus, sheet_notes["06_Novus_Stats"])
        write_table_with_note(writer, "07_Novus_ByBrand", novus_by_brand, sheet_notes["07_Novus_ByBrand"])

        write_table_with_note(writer, "08_Silpo_Clean", silpo, sheet_notes["08_Silpo_Clean"])
        write_table_with_note(writer, "09_Silpo_Stats", desc_silpo, sheet_notes["09_Silpo_Stats"])
        write_table_with_note(writer, "10_Silpo_ByBrand", silpo_by_brand, sheet_notes["10_Silpo_ByBrand"])

        write_table_with_note(writer, "11_Series_ProducerUA", series_producer, sheet_notes["11_Series_ProducerUA"])
        write_table_with_note(writer, "12_Series_ConsumerUA", series_consumer, sheet_notes["12_Series_ConsumerUA"])
        write_table_with_note(writer, "13_Series_EU", series_eu, sheet_notes["13_Series_EU"])
        write_table_with_note(writer, "14_Series_CME", series_cme, sheet_notes["14_Series_CME"])

        write_table_with_note(writer, "15_Correlations", corr_df, sheet_notes["15_Correlations"])
        write_table_with_note(writer, "Corr_Matrix", corr_matrix, sheet_notes["Corr_Matrix"])
        write_table_with_note(writer, "Stargazer_Like", stargazer_like, sheet_notes["Stargazer_Like"])
        for src in DATASET_ORDER:
            write_table_with_note(writer, f"Corr_Products_{src}", corr_modules.get(f"Corr_Products_{src}", pd.DataFrame()), sheet_notes[f"Corr_Products_{src}"])
        write_table_with_note(writer, "Corr_Regions_Prozorro", corr_modules.get("Corr_Regions_Prozorro", pd.DataFrame()), sheet_notes["Corr_Regions_Prozorro"])
        write_table_with_note(writer, "Corr_Brands_Retail", corr_modules.get("Corr_Brands_Retail", pd.DataFrame()), sheet_notes["Corr_Brands_Retail"])
        write_table_with_note(writer, "Corr_Sources_ByProduct", corr_modules.get("Corr_Sources_ByProduct", pd.DataFrame()), sheet_notes["Corr_Sources_ByProduct"])
        write_table_with_note(writer, "LagMatrix_ByProduct", lag_best, sheet_notes["LagMatrix_ByProduct"])
        write_table_with_note(writer, "LagProfiles_ByProduct", lag_profiles, sheet_notes["LagProfiles_ByProduct"])

        # write tests placeholder; then overwrite with sectioned layout below interpretation row
        write_table_with_note(writer, "16_Tests", pd.DataFrame(), sheet_notes["16_Tests"])
        write_table_with_note(writer, "Test_Stationarity", test_stationarity, sheet_notes["Test_Stationarity"])
        write_table_with_note(writer, "Test_Autocorr", test_autocorr, sheet_notes["Test_Autocorr"])
        write_table_with_note(writer, "Test_Heterosk", test_heterosk, sheet_notes["Test_Heterosk"])
        write_table_with_note(writer, "Test_Normality", test_normality, sheet_notes["Test_Normality"])
        write_table_with_note(writer, "Test_Stability", test_stability, sheet_notes["Test_Stability"])

        write_table_with_note(writer, "17_Forecasts_Prices", forecast_prices, sheet_notes["17_Forecasts_Prices"])
        write_table_with_note(writer, "18_Forecasts_Discounts", forecast_discounts, sheet_notes["18_Forecasts_Discounts"])
        write_table_with_note(writer, "19_Model_Eligibility", eligibility, sheet_notes["19_Model_Eligibility"])
        write_table_with_note(writer, "20_Model_Plan", model_plan, sheet_notes["20_Model_Plan"])
        write_table_with_note(writer, "21_Dataset_Registry", dataset_registry, sheet_notes["21_Dataset_Registry"])
        write_table_with_note(writer, "22_Product_Mapping", product_mapping_table, sheet_notes["22_Product_Mapping"])
        write_table_with_note(writer, "23_ARDL_Summary", ardl_summary, sheet_notes["23_ARDL_Summary"])
        write_table_with_note(writer, "24_ECM_Summary", ecm_summary, sheet_notes["24_ECM_Summary"])
        write_table_with_note(writer, "25_NARDL_Summary", nardl_summary, sheet_notes["25_NARDL_Summary"])
        write_table_with_note(writer, "26_VECM_Summary", vecm_summary, sheet_notes["26_VECM_Summary"])
        write_table_with_note(writer, "27_NEIO_Proxy_Summary", neio_proxy_summary, sheet_notes["27_NEIO_Proxy_Summary"])
        # Requested thesis spec sheets
        write_table_with_note(writer, "Standardisation_UnitPolicy", standardisation_policy, sheet_notes["Standardisation_UnitPolicy"])
        write_table_with_note(writer, "Outliers_Winsorisation", outliers_summary, sheet_notes["Outliers_Winsorisation"])
        write_table_with_note(writer, "Brand_IO_Metrics", brand_io_metrics, sheet_notes["Brand_IO_Metrics"])
        write_table_with_note(writer, "Brand_Economic_Metrics", brand_econ_metrics, sheet_notes["Brand_Economic_Metrics"])
        write_table_with_note(writer, "Series_Catalog", series_catalog, sheet_notes["Series_Catalog"])
        write_table_with_note(writer, "Correlations_All", corr_df, sheet_notes["Correlations_All"])
        write_table_with_note(writer, "LagMatrix_All", lag_matrix_all, sheet_notes["LagMatrix_All"])
        write_table_with_note(writer, "Tests_All", tests_df, sheet_notes["Tests_All"])
        write_table_with_note(writer, "Tests_All_Full", tests_df, sheet_notes["Tests_All_Full"])
        write_table_with_note(writer, "Tests_Summary", tests_summary, sheet_notes["Tests_Summary"])
        write_table_with_note(writer, "ARDL_Summary", ardl_summary, sheet_notes["ARDL_Summary"])
        write_table_with_note(writer, "NARDL_Summary", nardl_summary, sheet_notes["NARDL_Summary"])
        write_table_with_note(writer, "ECM_Summary", ecm_summary, sheet_notes["ECM_Summary"])
        write_table_with_note(writer, "VECM_Summary", vecm_summary, sheet_notes["VECM_Summary"])
        write_table_with_note(writer, "NEIO_Proxy_Summary", neio_proxy_summary, sheet_notes["NEIO_Proxy_Summary"])
        longrun_summary = pd.concat(
            [
                ardl_summary.assign(model_family="ARDL"),
                ecm_summary.assign(model_family="ECM"),
                nardl_summary.assign(model_family="NARDL"),
                vecm_summary.assign(model_family="VECM"),
            ],
            ignore_index=True,
            sort=False,
        )
        write_table_with_note(writer, "Models_LongRun_Summary", longrun_summary, sheet_notes["Models_LongRun_Summary"])
        write_table_with_note(writer, "Models_ShortRun_Summary", short_run_summary, sheet_notes["Models_ShortRun_Summary"])
        write_table_with_note(writer, "Chain_Effects_Summary", chain_summary, sheet_notes["Chain_Effects_Summary"])
        write_table_with_note(writer, "MSEM_Summary", msem_summary, sheet_notes["MSEM_Summary"])
        write_table_with_note(writer, "Silpo_Discounts_Occurrence", silpo_occurrence, sheet_notes["Silpo_Discounts_Occurrence"])
        write_table_with_note(writer, "Silpo_Discounts_Depth", silpo_depth, sheet_notes["Silpo_Discounts_Depth"])
        write_table_with_note(writer, "Silpo_Transmission_PromoCtrl", silpo_transmission, sheet_notes["Silpo_Transmission_PromoCtrl"])
        write_table_with_note(writer, "Prozorro_Regional_FE", prozorro_regional_fe, sheet_notes["Prozorro_Regional_FE"])
        write_table_with_note(writer, "Prozorro_Regional_Models", prozorro_regional_models, sheet_notes["Prozorro_Regional_Models"])
        write_table_with_note(
            writer,
            "Prozorro_Regional_Effects_Summary",
            region_product_matrix,
            sheet_notes["Prozorro_Regional_Effects_Summary"],
        )
        write_table_with_note(writer, "Model_Eligibility_Summary", eligibility, sheet_notes["Model_Eligibility_Summary"])
        write_table_with_note(writer, "Matrices_HeatmapReady", matrices_heatmap, sheet_notes["Matrices_HeatmapReady"])

        for sheet_name, df in decomp_sheet_tables.items():
            write_table_with_note(writer, sheet_name, df, "Decomposition components for this source/product series.")
            sheet_notes[sheet_name] = "Decomposition components for this source/product series."
        for sheet_name, df in before_after_ln_tables.items():
            write_table_with_note(writer, sheet_name, df, "Before-ln and after-ln chart data with rolling trends.")
            sheet_notes[sheet_name] = "Before-ln and after-ln chart data with rolling trends."
        for sheet_name, df in overlay_tables.items():
            write_table_with_note(writer, sheet_name, df, "Aligned intersection overlay across available sources.")
            sheet_notes[sheet_name] = "Aligned intersection overlay across available sources."
        for sheet_name, df in ecm_detail_tables.items():
            write_table_with_note(writer, sheet_name, df, "Detailed ECM table for this product block.")
            sheet_notes[sheet_name] = "Detailed ECM table for this product block."
        for sheet_name, df in vecm_detail_tables.items():
            write_table_with_note(writer, sheet_name, df, "Detailed VECM table for this product block.")
            sheet_notes[sheet_name] = "Detailed VECM table for this product block."
        for sheet_name, df in nardl_detail_tables.items():
            write_table_with_note(writer, sheet_name, df, "Detailed NARDL table for this product block.")
            sheet_notes[sheet_name] = "Detailed NARDL table for this product block."
        for sheet_name, df in retail_trans_tables.items():
            write_table_with_note(writer, sheet_name, df, "Retail short-run transmission coefficients for this product.")
            sheet_notes[sheet_name] = "Retail short-run transmission coefficients for this product."
        for sheet_name, df in chain_effect_tables.items():
            write_table_with_note(writer, sheet_name, df, "Chain effects Retail↔ProZorro↔Producer for this product.")
            sheet_notes[sheet_name] = "Chain effects Retail↔ProZorro↔Producer for this product."

        wb = writer.book
        style_readme_sheet(wb["00_README"])
        write_tests_sections(wb["16_Tests"], tests_df, start_row=3)
        ws16 = wb["16_Tests"]
        ws16["A1"] = sheet_notes["16_Tests"]
        ws16["A1"].font = Font(bold=True, color="000000")
        ws16["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws16.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
        ws16.row_dimensions[1].height = 36

    shutil.copy2(tmp_xlsx, cfg.output_file)
    try:
        tmp_xlsx.unlink(missing_ok=True)
    except Exception:
        pass

    separate_dir = cfg.output_file.parent / f"{cfg.output_file.stem}_separate"
    _export_separate_csvs(
        separate_dir,
        {
            "Transformation_Ledger": transformation_ledger,
            "Transformation_Stats": transformation_stats,
            "Transformation_Tests": transformation_tests,
            "Log_Variants_All": log_variants,
            "Decomposition_All": decomp_all,
            "Decomposition_Summary": decomp_summary,
            "BeforeAfterLN_All": before_after_ln_all,
            "Overlay_All": overlay_all,
            "Correlations_All": corr_df,
            "Corr_Matrix": corr_matrix,
            "Corr_Sources_ByProduct": corr_modules.get("Corr_Sources_ByProduct", pd.DataFrame()),
            "Corr_Regions_Prozorro": corr_modules.get("Corr_Regions_Prozorro", pd.DataFrame()),
            "Corr_Brands_Retail": corr_modules.get("Corr_Brands_Retail", pd.DataFrame()),
            "Stargazer_Like": stargazer_like,
            "Tests_All": tests_df,
            "Tests_Summary": tests_summary,
            "Test_Stationarity": test_stationarity,
            "Test_Autocorr": test_autocorr,
            "Test_Heterosk": test_heterosk,
            "Test_Normality": test_normality,
            "Test_Stability": test_stability,
            "LagMatrix_ByProduct": lag_best,
            "LagProfiles_ByProduct": lag_profiles,
            "Model_Eligibility_Summary": eligibility,
            "Model_Plan": model_plan,
            "ARDL_Summary": ardl_summary,
            "ECM_Summary": ecm_summary,
            "NARDL_Summary": nardl_summary,
            "VECM_Summary": vecm_summary,
            "Models_ShortRun_Summary": short_run_summary,
            "Chain_Effects_Summary": chain_summary,
            "MSEM_Summary": msem_summary,
            "NEIO_Proxy_Summary": neio_proxy_summary,
            "Silpo_Discounts_Occurrence": silpo_occurrence,
            "Silpo_Discounts_Depth": silpo_depth,
            "Silpo_Transmission_PromoCtrl": silpo_transmission,
            "Prozorro_Regional_FE": prozorro_regional_fe,
            "Prozorro_Regional_Models": prozorro_regional_models,
            "Prozorro_Regional_Effects_Summary": region_product_matrix,
        },
    )
    print(f"[EXPORT] Separate CSV tables: {separate_dir}")

    pdf_out = cfg.pdf_output if cfg.pdf_output is not None else cfg.output_file.with_name("rw2_full_report.pdf")
    pdf_saved = save_full_report_pdf(cfg.output_file, pdf_out, sheet_notes)
    if pdf_saved is not None:
        print(f"[PDF] Saved full report: {pdf_saved}")
    else:
        print("[PDF] Full PDF report was not generated in this environment.")

    print_terminal_report(datasets_clean, all_daily, corr_df, tests_df, eligibility, cfg.output_file)


def parse_args() -> Config:
    parser = argparse.ArgumentParser(description="RW2 thesis-grade pipeline: product-first, one XLSX output.")
    parser.add_argument("--model-dir", type=Path, default=DEFAULT_MODEL_DIR, help="Directory with source files")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Main input workbook")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output workbook path (rw2_full_output.xlsx)")
    parser.add_argument("--pdf-output", type=Path, default=None, help="Optional full PDF report path (default: rw2_full_report.pdf)")
    args = parser.parse_args()

    out = args.output
    if out.suffix.lower() != ".xlsx":
        out = out / "rw2_full_output.xlsx"

    if out.name != "rw2_full_output.xlsx":
        out = out.parent / "rw2_full_output.xlsx"

    return Config(
        model_dir=args.model_dir,
        input_file=args.input,
        output_file=out,
        pdf_output=args.pdf_output,
    )


if __name__ == "__main__":
    cfg = parse_args()
    run_pipeline(cfg)
