#!/usr/bin/env python3
"""Product-first analytics pipeline on full_uah_unified_output.xlsx."""

from __future__ import annotations

import argparse
import re
import shutil
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl.styles import Alignment, Font, PatternFill
from statsmodels.stats.diagnostic import acorr_ljungbox
from statsmodels.tsa.stattools import adfuller

DEFAULT_INPUT = Path("/Users/getapple/Documents/KSE/Master Thesis/Main materials/Model/results/full_uah_unified_output.xlsx")
DEFAULT_OUTPUT = Path("/Users/getapple/Documents/KSE/Master Thesis/Main materials/Model/results/full_uah_unified_output.xlsx")
RAW_FALLBACK = Path("/Users/getapple/Documents/KSE/Master Thesis/Data/for model/full_uah.xlsx")


def ntext(v: object) -> str:
    if pd.isna(v):
        return ""
    return re.sub(r"\\s+", " ", str(v).strip().lower())


def detect_product(raw: object) -> str:
    s = ntext(raw)
    rules = [
        ("масл|butter", "Масло вершкове"),
        ("сметан|sour cream", "Сметана"),
        ("вершк|cream", "Вершки"),
        ("кисломолоч|творог|cottage", "Сир кисломолочний"),
        ("сир|cheese|gouda|edam|emmental|cheddar", "Сир твердий"),
        ("молок|milk|кефір|kefir|ряжан|smp|wmp", "Молоко питне"),
        ("йогурт|yogurt|десерт", "Йогурт/десерт"),
    ]
    for pat, out in rules:
        if re.search(pat, s):
            return out
    return "Інше/невідомо"


def detect_standardized_type(product: object) -> str:
    s = ntext(product)
    if any(k in s for k in ["масло", "butter"]):
        return "butter"
    if any(k in s for k in ["сметана", "sour"]):
        return "sour_cream"
    if any(k in s for k in ["вершки", "cream"]):
        return "cream"
    if any(k in s for k in ["кисломол", "творог", "cottage"]):
        return "cottage_cheese"
    if any(k in s for k in ["сир", "cheese", "gouda", "edam", "emmental", "cheddar"]):
        return "hard_cheese"
    if any(k in s for k in ["молоко", "milk", "кефір", "kefir", "ряжан", "smp", "wmp"]):
        return "milk"
    if any(k in s for k in ["йогурт", "yogurt", "десерт"]):
        return "yogurt_dessert"
    return "other"


def _to_datetime(df: pd.DataFrame, col: str) -> pd.Series:
    return pd.to_datetime(df[col], errors="coerce")


def load_raw_book(path: Path) -> dict[str, pd.DataFrame]:
    xls = pd.ExcelFile(path)
    wanted = ["CME III", "Consumer_UA", "Producer_UA", "Prozorro", "Europe", "Silpo", "Novus"]
    return {name: xls.parse(name) if name in xls.sheet_names else pd.DataFrame() for name in wanted}


def prep_consumer_or_producer(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    out = pd.DataFrame(
        {
            "date": _to_datetime(df, "date"),
            "source": source,
            "product_raw": df.get("ua_product"),
            "region": df.get("region"),
            "price_uah": pd.to_numeric(df.get("price_pchip"), errors="coerce"),
            "unit": df.get("unit"),
            "method": df.get("method"),
            "brand": np.nan,
        }
    )
    return out.dropna(subset=["date", "price_uah"])


def prep_cme(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    out = pd.DataFrame(
        {
            "date": _to_datetime(df, "Date"),
            "source": "CME",
            "product_raw": "Class III Milk",
            "region": "US",
            "price_uah": pd.to_numeric(df.get("CME III UAH"), errors="coerce"),
            "unit": "UAH/index",
            "method": "source_provided",
            "brand": np.nan,
        }
    )
    return out.dropna(subset=["date", "price_uah"])


def prep_europe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    out = pd.DataFrame(
        {
            "date": _to_datetime(df, "date"),
            "source": "EU",
            "product_raw": df.get("Product"),
            "region": df.get("Country"),
            "price_uah": pd.to_numeric(df.get("Price (UAH/kg)"), errors="coerce"),
            "unit": "UAH/kg",
            "method": "source_provided",
            "brand": np.nan,
        }
    )
    return out.dropna(subset=["date", "price_uah"])


def prep_prozorro(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    out = pd.DataFrame(
        {
            "date": _to_datetime(df, "Дата"),
            "source": "ProZorro",
            "product_raw": df.get("Товар").fillna(df.get("Product")),
            "region": df.get("Регіон організатора"),
            "price_uah": pd.to_numeric(df.get("Ціна за одиницю"), errors="coerce"),
            "unit": df.get("Одиниця виміру"),
            "method": "source_provided",
            "brand": np.nan,
        }
    )
    return out.dropna(subset=["date", "price_uah"])


def prep_retail(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    product_col = "Product" if "Product" in df.columns else "product"
    out = pd.DataFrame(
        {
            "date": _to_datetime(df, "date"),
            "source": source,
            "product_raw": df.get("product_title").fillna(df.get(product_col)).fillna(df.get("product_name")),
            "region": pd.NA,
            "price_uah": pd.to_numeric(df.get("unit_price"), errors="coerce").fillna(pd.to_numeric(df.get("price_current"), errors="coerce")),
            "unit": df.get("unit_std").fillna("standardized_unit"),
            "method": "unit_price_or_current",
            "brand": df.get("brand").fillna("Unknown"),
            "discount_dummy_bulk": pd.to_numeric(df.get("discount_dummy_bulk"), errors="coerce"),
            "discount_dummy_discount": pd.to_numeric(df.get("discount_dummy_discount"), errors="coerce"),
            "discount_dummy_regular": pd.to_numeric(df.get("discount_dummy_regular"), errors="coerce"),
            "discount_value": pd.to_numeric(df.get("discount_value"), errors="coerce"),
        }
    )
    return out.dropna(subset=["date", "price_uah"])


def build_unified_panel_from_raw(raw_path: Path) -> pd.DataFrame:
    book = load_raw_book(raw_path)
    parts = [
        prep_cme(book["CME III"]),
        prep_consumer_or_producer(book["Consumer_UA"], "ConsumerUA"),
        prep_consumer_or_producer(book["Producer_UA"], "ProducerUA"),
        prep_prozorro(book["Prozorro"]),
        prep_europe(book["Europe"]),
        prep_retail(book["Silpo"], "Silpo"),
        prep_retail(book["Novus"], "Novus"),
    ]
    panel = pd.concat(parts, ignore_index=True)
    panel["product"] = panel["product_raw"].map(detect_product)
    panel["standardized_type"] = panel["product"].map(detect_standardized_type)
    panel["year_month"] = panel["date"].dt.to_period("M").astype(str)
    panel["week_start"] = panel["date"].dt.to_period("W-MON").dt.start_time
    panel = panel.sort_values(["date", "source", "product"], na_position="last").reset_index(drop=True)
    return panel


def load_unified_panel_from_workbook(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    if "unified_panel" not in xls.sheet_names:
        raise ValueError(f"Workbook does not contain 'unified_panel': {path}")

    # Auto-detect header row:
    # - standard workbook: header is row 1
    # - workbook with interpretation note: header is row 3 (startrow=2 in writer)
    probe = pd.read_excel(path, sheet_name="unified_panel", header=None, nrows=6)
    header_row = 0
    for i in range(len(probe)):
        vals = {ntext(v) for v in probe.iloc[i].tolist() if pd.notna(v)}
        if "date" in vals and "source" in vals:
            header_row = i
            break

    panel = pd.read_excel(path, sheet_name="unified_panel", header=header_row)
    panel["date"] = pd.to_datetime(panel.get("date"), errors="coerce")
    panel["price_uah"] = pd.to_numeric(panel.get("price_uah"), errors="coerce")
    if "product" not in panel.columns:
        panel["product"] = panel.get("product_raw", pd.Series(index=panel.index)).map(detect_product)
    if "standardized_type" not in panel.columns:
        panel["standardized_type"] = panel["product"].map(detect_standardized_type)
    if "source" not in panel.columns:
        panel["source"] = "Unknown"
    if "brand" not in panel.columns:
        panel["brand"] = "Unknown"
    panel["year_month"] = pd.to_datetime(panel["date"], errors="coerce").dt.to_period("M").astype(str)
    panel["week_start"] = pd.to_datetime(panel["date"], errors="coerce").dt.to_period("W-MON").dt.start_time
    return panel.dropna(subset=["date", "price_uah"]).copy()


def build_daily_variants(panel: pd.DataFrame) -> pd.DataFrame:
    rows = []
    gcols = ["source", "product", "standardized_type"]
    interp_sources = {"consumer_ua", "producer_ua", "europe", "consumerua", "producerua", "eu"}
    for keys, g in panel.groupby(gcols, dropna=False):
        g = g.sort_values("date")
        daily_idx = pd.date_range(g["date"].min(), g["date"].max(), freq="D")
        s = g.groupby("date", as_index=True)["price_uah"].median().reindex(daily_idx)
        real = s.copy()
        src = ntext(keys[0])

        # Interpolation variants are admissible only for Consumer_UA / Producer_UA / Europe.
        if src in interp_sources:
            linear = s.interpolate(method="linear", limit_direction="both")
            try:
                pchip = s.interpolate(method="pchip", limit_direction="both")
            except Exception:
                pchip = linear.copy()
        else:
            linear = real.copy()
            pchip = real.copy()

        out = pd.DataFrame({"date": daily_idx, "price_real": real.values, "price_linear": linear.values, "price_pchip": pchip.values})
        out["source"], out["product"], out["standardized_type"] = keys
        if src in interp_sources:
            out["imputed_flag_linear"] = out["price_real"].isna() & out["price_linear"].notna()
            out["imputed_flag_pchip"] = out["price_real"].isna() & out["price_pchip"].notna()
        else:
            out["imputed_flag_linear"] = False
            out["imputed_flag_pchip"] = False
        rows.append(out)

    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def descriptive_stats(daily: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (src, prod, st), g in daily.groupby(["source", "product", "standardized_type"], dropna=False):
        for vcol in ["price_real", "price_linear", "price_pchip"]:
            x = pd.to_numeric(g[vcol], errors="coerce")
            x_valid = x.dropna()
            if x_valid.empty:
                continue
            ret = x_valid.pct_change().dropna()
            rows.append(
                {
                    "source": src,
                    "product": prod,
                    "standardized_type": st,
                    "series_variant": vcol,
                    "count": int(x_valid.size),
                    "missing": int(x.isna().sum()),
                    "imputed_share": float(g["imputed_flag_linear"].mean()) if vcol == "price_linear" else float(g["imputed_flag_pchip"].mean()) if vcol == "price_pchip" else 0.0,
                    "mean": float(x_valid.mean()),
                    "median": float(x_valid.median()),
                    "std": float(x_valid.std(ddof=1)) if x_valid.size > 1 else np.nan,
                    "min": float(x_valid.min()),
                    "max": float(x_valid.max()),
                    "q05": float(x_valid.quantile(0.05)),
                    "q25": float(x_valid.quantile(0.25)),
                    "q75": float(x_valid.quantile(0.75)),
                    "q95": float(x_valid.quantile(0.95)),
                    "skew": float(x_valid.skew()) if x_valid.size > 2 else np.nan,
                    "kurtosis": float(x_valid.kurt()) if x_valid.size > 3 else np.nan,
                    "cv": float(x_valid.std(ddof=1) / x_valid.mean()) if x_valid.mean() != 0 else np.nan,
                    "rolling_vol_7": float(ret.rolling(7).std().mean()) if ret.size >= 7 else np.nan,
                    "rolling_vol_30": float(ret.rolling(30).std().mean()) if ret.size >= 30 else np.nan,
                }
            )
    return pd.DataFrame(rows)


def correlations(daily: pd.DataFrame) -> pd.DataFrame:
    piv = daily.pivot_table(index=["date", "product"], columns="source", values="price_pchip", aggfunc="mean").reset_index()
    pairs = [
        ("ProducerUA", "ConsumerUA"),
        ("ProZorro", "ProducerUA"),
        ("ProZorro", "ConsumerUA"),
        ("Silpo", "ProducerUA"),
        ("Novus", "ProducerUA"),
        ("EU", "ProducerUA"),
        ("CME", "ProducerUA"),
    ]
    lags = [0, 7, 14]

    out = []
    for prod, g in piv.groupby("product", dropna=False):
        g = g.sort_values("date")
        for a, b in pairs:
            if a not in g.columns or b not in g.columns:
                continue
            for lag in lags:
                bser = g[b].shift(lag) if lag else g[b]
                dd = pd.DataFrame({"a": g[a], "b": bser}).dropna()
                if len(dd) < 10:
                    continue
                out.append(
                    {
                        "product": prod,
                        "source_a": a,
                        "source_b": b,
                        "lag_days": lag,
                        "pearson": float(dd["a"].corr(dd["b"], method="pearson")),
                        "spearman": float(dd["a"].corr(dd["b"], method="spearman")),
                        "n": int(len(dd)),
                    }
                )
    return pd.DataFrame(out)


def run_tests(daily: pd.DataFrame) -> pd.DataFrame:
    out = []
    weekly = daily.copy()
    weekly["week"] = weekly["date"].dt.to_period("W-MON").dt.start_time
    weekly = weekly.groupby(["source", "product", "standardized_type", "week"], as_index=False).agg(
        price_real=("price_real", "median"), price_linear=("price_linear", "median"), price_pchip=("price_pchip", "median")
    )

    for (src, prod, st), g in weekly.groupby(["source", "product", "standardized_type"], dropna=False):
        for v in ["price_real", "price_linear", "price_pchip"]:
            y = pd.to_numeric(g[v], errors="coerce").dropna()
            if len(y) < 24 or y.nunique() <= 1:
                out.append({"source": src, "product": prod, "standardized_type": st, "series_variant": v, "n_obs": int(len(y)), "adf_p": np.nan, "ljung_box_p": np.nan, "action_label": "insufficient_or_constant"})
                continue
            try:
                adf_p = float(adfuller(y, autolag="AIC")[1])
            except Exception:
                adf_p = np.nan
            dy = np.log(y.where(y > 0)).diff().replace([np.inf, -np.inf], np.nan).dropna()
            if len(dy) >= 10:
                lb = float(acorr_ljungbox(dy, lags=[min(8, max(2, len(dy) // 5))], return_df=True)["lb_pvalue"].iloc[0])
            else:
                lb = np.nan
            label = "ok"
            if pd.notna(adf_p) and adf_p > 0.05:
                label = "non_stationary_diff_or_coint"
            if pd.notna(lb) and lb < 0.05:
                label = f"{label};autocorr_add_lags" if label != "ok" else "autocorr_add_lags"
            out.append({"source": src, "product": prod, "standardized_type": st, "series_variant": v, "n_obs": int(len(y)), "adf_p": adf_p, "ljung_box_p": lb, "action_label": label})
    return pd.DataFrame(out)


def intersection_window(daily: pd.DataFrame) -> tuple[pd.Timestamp, pd.Timestamp]:
    spans = (
        daily.groupby("source", dropna=False)["date"]
        .agg(["min", "max"])
        .dropna()
        .reset_index()
    )
    if spans.empty:
        return pd.NaT, pd.NaT
    inter_start = spans["min"].max()
    inter_end = spans["max"].min()
    return inter_start, inter_end


def price_forecast(daily: pd.DataFrame, horizon_days: int = 7) -> pd.DataFrame:
    rows = []
    for (src, prod, st), g in daily.groupby(["source", "product", "standardized_type"], dropna=False):
        g = g.sort_values("date").copy()
        y = pd.to_numeric(g["price_pchip"], errors="coerce")
        if y.notna().sum() < 40:
            continue
        g["y"] = y
        g["lag1"] = g["y"].shift(1)
        g["lag7"] = g["y"].shift(7)
        g["dow"] = g["date"].dt.dayofweek
        m = g.dropna(subset=["y", "lag1", "lag7"])
        if len(m) < 40:
            continue
        X = sm.add_constant(pd.get_dummies(m[["lag1", "lag7", "dow"]], columns=["dow"], drop_first=True)).astype(float)
        fit = sm.OLS(m["y"], X).fit(cov_type="HC1")
        pred = fit.predict(X)
        mae = float(np.mean(np.abs(m["y"] - pred)))

        # Recursive next-7-days forecast with explicit forecast_date.
        history = m[["date", "y"]].copy()
        last_date = history["date"].max()
        for step in range(1, horizon_days + 1):
            fdate = last_date + pd.Timedelta(days=step)
            lag1 = history["y"].iloc[-1] if len(history) >= 1 else np.nan
            lag7 = history["y"].iloc[-7] if len(history) >= 7 else lag1
            new_row = pd.DataFrame({"lag1": [lag1], "lag7": [lag7], "dow": [fdate.dayofweek]})
            new_X = pd.get_dummies(new_row, columns=["dow"], drop_first=True)
            new_X = new_X.reindex(columns=[c for c in X.columns if c != "const"], fill_value=0)
            new_X = sm.add_constant(new_X, has_constant="add").astype(float)
            yhat = float(fit.predict(new_X).iloc[0])
            history = pd.concat([history, pd.DataFrame({"date": [fdate], "y": [yhat]})], ignore_index=True)
            rows.append(
                {
                    "source": src,
                    "product": prod,
                    "standardized_type": st,
                    "scope": "product",
                    "forecast_date": fdate,
                    "forecast_horizon_day": step,
                    "forecast_price_uah": yhat,
                    "n_obs_train": int(len(m)),
                    "model": "AR_lag1_lag7_dow",
                    "rolling_mae": mae,
                }
            )

    # Dataset-level fallback: always forecast next 7 days for each source
    for src, g in daily.groupby("source", dropna=False):
        g = g.sort_values("date").copy()
        s = g.groupby("date", as_index=False)["price_pchip"].median().rename(columns={"price_pchip": "y"})
        s["y"] = pd.to_numeric(s["y"], errors="coerce")
        s = s.dropna(subset=["y"])
        if len(s) < 14:
            continue
        s["lag1"] = s["y"].shift(1)
        s["lag7"] = s["y"].shift(7)
        s["dow"] = s["date"].dt.dayofweek
        m = s.dropna(subset=["y", "lag1", "lag7"])
        if len(m) < 14:
            continue
        X = sm.add_constant(pd.get_dummies(m[["lag1", "lag7", "dow"]], columns=["dow"], drop_first=True)).astype(float)
        fit = sm.OLS(m["y"], X).fit(cov_type="HC1")
        mae = float(np.mean(np.abs(m["y"] - fit.predict(X))))

        history = m[["date", "y"]].copy()
        last_date = history["date"].max()
        for step in range(1, horizon_days + 1):
            fdate = last_date + pd.Timedelta(days=step)
            lag1 = history["y"].iloc[-1] if len(history) >= 1 else np.nan
            lag7 = history["y"].iloc[-7] if len(history) >= 7 else lag1
            new_row = pd.DataFrame({"lag1": [lag1], "lag7": [lag7], "dow": [fdate.dayofweek]})
            new_X = pd.get_dummies(new_row, columns=["dow"], drop_first=True)
            new_X = new_X.reindex(columns=[c for c in X.columns if c != "const"], fill_value=0)
            new_X = sm.add_constant(new_X, has_constant="add").astype(float)
            yhat = float(fit.predict(new_X).iloc[0])
            history = pd.concat([history, pd.DataFrame({"date": [fdate], "y": [yhat]})], ignore_index=True)
            rows.append(
                {
                    "source": src,
                    "product": "ALL_PRODUCTS",
                    "standardized_type": "all",
                    "scope": "dataset",
                    "forecast_date": fdate,
                    "forecast_horizon_day": step,
                    "forecast_price_uah": yhat,
                    "n_obs_train": int(len(m)),
                    "model": "AR_lag1_lag7_dow_dataset",
                    "rolling_mae": mae,
                }
            )
    return pd.DataFrame(rows)


def terminal_report(panel: pd.DataFrame, daily: pd.DataFrame, corr: pd.DataFrame, tests: pd.DataFrame, fcst: pd.DataFrame, out_path: Path) -> None:
    print("\\n=== UNIFIED PANEL REPORT ===")
    for src, g in panel.groupby("source", dropna=False):
        top = g["product"].value_counts().head(3).index.tolist()
        print(
            f"- {src}: rows={len(g):,}, date=[{g['date'].min().date()}..{g['date'].max().date()}], "
            f"missing_price={g['price_uah'].isna().mean():.2%}, top_products={', '.join(top)}"
        )

    if not daily.empty:
        imp = daily.groupby("source", dropna=False)[["imputed_flag_linear", "imputed_flag_pchip"]].mean().reset_index()
        print("\\n[Imputed share]")
        for _, r in imp.iterrows():
            print(f"- {r['source']}: linear={r['imputed_flag_linear']:.2%}, pchip={r['imputed_flag_pchip']:.2%}")

    if not corr.empty:
        top5 = corr.assign(abs_pearson=lambda d: d["pearson"].abs()).sort_values("abs_pearson", ascending=False).head(5)
        print("\\n[Top-5 correlations]")
        for _, r in top5.iterrows():
            print(f"- {r['product']}: {r['source_a']} vs {r['source_b']} lag={int(r['lag_days'])}d pearson={r['pearson']:.3f}")

    if not tests.empty:
        print("\\n[Eligibility quick view]")
        t = tests.groupby("product", dropna=False).agg(non_stationary=("action_label", lambda s: int(s.astype(str).str.contains("non_stationary").any()))).reset_index()
        for _, r in t.head(10).iterrows():
            print(f"- {r['product']}: non_stationary_flag={r['non_stationary']}")

    print(f"\\nSaved: {out_path}")
    if not fcst.empty:
        print(f"Forecast rows: {len(fcst):,} (next 7 days with forecast_date)")


def run(input_file: Path, output_file: Path, raw_fallback: Path) -> None:
    if input_file.exists():
        panel = load_unified_panel_from_workbook(input_file)
    elif raw_fallback.exists():
        panel = build_unified_panel_from_raw(raw_fallback)
    else:
        raise FileNotFoundError(f"Neither {input_file} nor {raw_fallback} exists")

    panel["product"] = panel["product"].fillna(panel.get("product_raw", pd.Series(index=panel.index)).map(detect_product))
    panel["standardized_type"] = panel["standardized_type"].fillna(panel["product"].map(detect_standardized_type))

    daily = build_daily_variants(panel)
    stats = descriptive_stats(daily)
    inter_start, inter_end = intersection_window(daily)
    daily_model = daily[(daily["date"] >= inter_start) & (daily["date"] <= inter_end)].copy() if pd.notna(inter_start) and pd.notna(inter_end) and inter_start <= inter_end else daily.iloc[0:0].copy()

    corr = correlations(daily_model)
    tests = run_tests(daily_model)
    fcst = price_forecast(daily_model, horizon_days=7)

    summary_by_source = (
        panel.groupby("source", dropna=False)
        .agg(
            n_obs=("price_uah", "size"),
            n_products=("product", "nunique"),
            n_regions=("region", "nunique"),
            date_min=("date", "min"),
            date_max=("date", "max"),
            price_mean=("price_uah", "mean"),
            price_median=("price_uah", "median"),
        )
        .reset_index()
        .sort_values("source")
    )

    summary_by_month = (
        panel.groupby(["source", "year_month"], dropna=False)
        .agg(n_obs=("price_uah", "size"), avg_price_uah=("price_uah", "mean"), min_price_uah=("price_uah", "min"), max_price_uah=("price_uah", "max"))
        .reset_index()
        .sort_values(["source", "year_month"])
    )

    sheet_notes = {
        "unified_panel": "Interpretation: unified raw analytical panel. Each row is an observed price point by date/source/product. Use this as the base table before interpolation and tests.",
        "summary_by_source": "Interpretation: source-level coverage and central tendency. Compare n_obs/date ranges to assess representativeness and possible sample imbalance.",
        "summary_by_month": "Interpretation: monthly dynamics by source. Use avg/min/max to detect regime shifts, seasonality, and abrupt level changes.",
        "daily_real_linear_pchip": "Interpretation: real/linear/pchip variants. For Consumer_UA/Producer_UA/Europe interpolation fills gaps; for other datasets linear and pchip equal real by design.",
        "stats_product_series": "Interpretation: product-first descriptive statistics across variants. High CV/rolling_vol indicates instability; high imputed_share means stronger dependence on constructed values.",
        "correlations": "Interpretation: cross-source comovement by product and lag. Positive high correlations suggest transmission; lag with highest correlation indicates delayed pass-through.",
        "tests": "TEST INTERPRETATION: ADF p>0.05 => likely non-stationary (difference or test cointegration). Ljung-Box p<0.05 => autocorrelation (add lags / respecify). action_label provides the recommended next step.",
        "forecast_prices": "Interpretation: 7-day ahead forecasts on intersection window. Each row is one forecasted date. scope=dataset is aggregate per source; scope=product is product-level where sample is sufficient.",
    }

    def write_with_note(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ws = writer.book[sheet_name]
        note = sheet_notes.get(sheet_name, "")
        ws["A1"] = note
        ws["A1"].font = Font(bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        merge_to_col = max(6, df.shape[1] if not df.empty else 6)
        # Convert column index to Excel letter (supports 1..702 here, more than enough for this workbook)
        q, r = divmod(merge_to_col - 1, 26)
        end_col = (chr(64 + q) if q else "") + chr(65 + r)
        ws.merge_cells(f"A1:{end_col}1")
        ws.row_dimensions[1].height = 36

    output_file.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir="/tmp") as tf:
        tmp_xlsx = Path(tf.name)
    with pd.ExcelWriter(tmp_xlsx, engine="openpyxl") as writer:
        write_with_note(writer, panel, "unified_panel")
        write_with_note(writer, summary_by_source, "summary_by_source")
        write_with_note(writer, summary_by_month, "summary_by_month")
        write_with_note(writer, daily, "daily_real_linear_pchip")
        write_with_note(writer, stats, "stats_product_series")
        write_with_note(writer, corr, "correlations")
        write_with_note(writer, tests, "tests")
        write_with_note(writer, fcst, "forecast_prices")
    shutil.copy2(tmp_xlsx, output_file)
    try:
        tmp_xlsx.unlink(missing_ok=True)
    except Exception:
        pass

    if pd.notna(inter_start) and pd.notna(inter_end) and inter_start <= inter_end:
        print(f"Modeling intersection window: [{inter_start.date()} .. {inter_end.date()}]")
        print(f"Intersection rows: {len(daily_model):,}")
    else:
        print("Modeling intersection window: not found (empty overlap).")

    terminal_report(panel, daily_model, corr, tests, fcst, output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run product-first analytics on full_uah_unified_output.xlsx")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Path to workbook with sheet unified_panel")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output workbook path")
    parser.add_argument("--raw-fallback", type=Path, default=RAW_FALLBACK, help="Fallback raw full_uah.xlsx if unified workbook is absent")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    run(args.input, args.output, args.raw_fallback)
